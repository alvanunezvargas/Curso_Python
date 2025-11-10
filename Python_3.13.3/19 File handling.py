"""
Manejo de ficheros
Hasta ahora hemos visto diferentes tipos de datos en Python. Normalmente almacenamos nuestros datos en diferentes formatos de
archivo. Además del manejo de archivos, también veremos diferentes formatos de archivo(.txt, .json, .xml, .csv, .tsv, .excel)
en esta sección. Primero, familiaricémonos con el manejo de archivos con formato de archivo común(.txt).

El manejo de archivos es una parte importante de la programación que nos permite crear, leer, actualizar y borrar archivos. En
Python para manejar datos usamos la función incorporada open().

# Sintaxis
open('filename', mode) # mode(r, a, w, x, t,b) puede ser: leer, escribir, actualizar.

"r" - Read - Valor por defecto. Abre un archivo para su lectura, devuelve un error si el archivo no existe
"a" - Append - Abre un fichero para añadir, crea el fichero si no existe.
"w" - Write - Abre un archivo para escritura, crea el archivo si no existe.
"x" - Create - Crea el fichero especificado, devuelve un error si el fichero existe
"t" - Text - Valor por defecto. Modo texto
"b" - Binary - Modo binario (por ejemplo, imágenes)
"""

# Abrir archivos para lectura: El modo de apertura por defecto es lectura, por lo que no tenemos que especificar 'r' o 'rt'.
# He creado y guardado un archivo llamado reading_file_example.txt en el directorio C:/Users/alvan/Downloads. Veamos cómo se abre:

f = open('C:/Users/alvan/Downloads/reading_file_example.txt')
txt = f.read()
print(txt)
"""
This is an example to show how to open a file and read.
This is the second line of the text.I love python
"""
# Como puede ver en el ejemplo anterior, imprimí el archivo abierto y me dio información sobre él. Un archivo abierto tiene
# diferentes métodos de lectura: read(), readline, readlines. Un archivo abierto tiene que ser cerrado con el método close().
# read(): lee todo el texto como cadena.

f = open('C:/Users/alvan/Downloads/reading_file_example.txt')
txt = f.read()
print(type(txt))  # <class 'str'>
print(txt)
"""
This is an example to show how to open a file and read.
This is the second line of the text.I love python
"""
f.close() # (`f.close()`: Cierra el archivo después de leer su contenido utilizando el método `close()` de Python. Es
# importante cerrar el archivo después de leerlo para liberar los recursos del sistema operativo.)

# Si queremos limitar el número de caracteres que queremos leer, podemos limitarlo pasando un valor int al método read(number).

f = open('C:/Users/alvan/Downloads/reading_file_example.txt')
txt = f.read(10)
print(type(txt))  # <class 'str'>
print(txt)  # This is an
f.close()

# readline(): lee sólo la primera línea

f = open('C:/Users/alvan/Downloads/reading_file_example.txt')
line = f.readline()  # <class 'str'>
print(type(line))
print(line)  # This is an example to show how to open a file and read.
f.close()

# readlines(): lee todo el texto línea por línea y devuelve una lista de líneas

f = open('C:/Users/alvan/Downloads/reading_file_example.txt')
lines = f.readlines()
print(type(lines))  # <class 'list'>
print(lines)  # ['This is an example to show how to open a file and read.\n', 'This is the second line of the text.I love python']
f.close()

# Otra forma de obtener todas las líneas como una lista es utilizando splitlines():

f = open('C:/Users/alvan/Downloads/reading_file_example.txt')
lines = f.read().splitlines()
print(type(lines))  # <class 'list'>
print(lines) # ['This is an example to show how to open a file and read.', 'This is the second line of the text.I love python']
f.close()

# Después de abrir un archivo, debemos cerrarlo. Hay una gran tendencia a olvidar cerrarlos. Hay una nueva forma de abrir archivos
# usando with - cierra los archivos por sí mismo. Reescribamos el ejemplo anterior con el método with:

with open('C:/Users/alvan/Downloads/reading_file_example.txt') as f:
    lines = f.read().splitlines()
    print(type(lines))  # <class 'list'>
    print(lines) # ['This is an example to show how to open a file and read.', 'This is the second line of the text.I love python']
    
# Abrir archivos para escribir y actualizar: Para escribir en un archivo existente, debemos añadir un modo como parámetro a
# la función open():

# "a" - append - añadirá al final del archivo, si el archivo no existe crea un nuevo archivo.
# "w" - write - sobrescribirá cualquier contenido existente, si el archivo no existe lo creará.

# Modo de Escritura ('w'): Este modo se utiliza para crear un nuevo archivo o sobrescribir el contenido de un archivo existente.
# Si el archivo no existe, se creará.

with open('C:/Users/alvan/Downloads/reading_file_example.txt', 'w') as archivo:
    archivo.write('Este es el contenido del archivo.\n')
    archivo.write('Puedes agregar mas lineas de texto aqui\n')

# Modo de Escritura de Anexado ('a'): Este modo se utiliza para agregar contenido al final de un archivo existente o crear uno
# nuevo si no existe.

with open('C:/Users/alvan/Downloads/reading_file_example.txt', 'a') as archivo:
    archivo.write('Este texto se agrega al final del archivo.\n')
    archivo.write('Puedes seguir agregando más líneas aquí.\n')
    
# Modo de Lectura y Escritura ('r+'): Este modo se utiliza para abrir un archivo en modo de lectura y escritura al mismo tiempo.
# Puedes leer y escribir en el archivo.

with open('C:/Users/alvan/Downloads/reading_file_example.txt', 'r+') as archivo:
    contenido = archivo.read()
    archivo.write('Este texto se agrega al final del archivo.\n')
    
# Borrar archivos: si queremos eliminar un fichero usamos el módulo os.

import os
os.remove('C:/Users/alvan/Downloads/reading_file_example.txt')

# Si el archivo no existe, el método remove dará un error, por lo que es bueno utilizar una condición como esta:

import os
if os.path.exists('C:/Users/alvan/Downloads/reading_file_example.txt'):
    os.remove('C:/Users/alvan/Downloads/reading_file_example.txt')
else:
    print('Este archivo no existe')  # Este archivo no existe

# Tipos de archivos

# Archivos de Texto (.txt): Son archivos simples que contienen texto sin formato. Puedes leer y escribir en archivos de texto
# utilizando Python fácilmente, como lo vimos en ejemplos anteriores.

# Archivos JSON (.json): JSON es un formato de intercambio de datos que se utiliza para almacenar datos estructurados en texto
# plano. Python tiene una biblioteca incorporada llamada json que te permite trabajar con archivos JSON.

# Puedemos transformar el diccionario person_dct en un formato JSON de la siguiente manera:

person_dct= {
    "name":"Alvaro",
    "country":"Colombia",
    "city":"Armenia",
    "skills":["Matlab", "Quality","Python"]
}
print(type(person_dct))  # <class 'dict'>
import json
person_json = json.dumps(person_dct)
print(person_json)  # {"name": "Alvaro", "country": "Colombia", "city": "Armenia", "skills": ["Matlab", "Quality", "Python"]}
print(type(person_json))  # <class 'str'>

# Puedemos transformar un formato JSON person_json en un diccionario de la siguiente manera:

person_json = '''{
    "name":"Alvaro",
    "country":"Colombia",
    "city":"Armenia",
    "skills":["Matlab", "Quality","Python"]
}'''
print(type(person_json))  # <class 'str'>
import json
person_dct = json.loads(person_json)
print(person_dct)  # {'name': 'Alvaro', 'country': 'Colombia', 'city': 'Armenia', 'skills': ['Matlab', 'Quality', 'Python']}
print(type(person_dct))  # <class 'dict'>

# Guardar como archivo JSON: También podemos guardar nuestros datos como un archivo json. Vamos a guardarlo como un archivo
# json utilizando los siguientes pasos. Para escribir un archivo json, usamos el método json.dump(), puede tomar diccionario,
# archivo de salida, ensure_ascii y sangría.

import json

person = {'name': 'Alvaro', 'country': 'Colombia', 'city': 'Armenia', 'skills': ['Matlab', 'Quality', 'Python']}
with open('C:/Users/alvan/Downloads/json_example.json', 'w', encoding='utf-8') as f:
    json.dump(person, f, ensure_ascii=False, indent=4)

"""
`open('C:/Users/alvan/Downloads/json_example.json', 'w', encoding='utf-8')`: Abre el archivo `json_example.json` que se encuentra
en la ruta `C:/Users/alvan/Downloads/` utilizando la función `open()` de Python. El modo `'w'` indica que el archivo se abrirá en
modo escritura. El argumento `encoding='utf-8'` se utiliza para especificar la codificación de caracteres del archivo.
- `json.dump(person, f, ensure_ascii=False, indent=4)`: Utiliza la función `json.dump()` para escribir el diccionario `person`
en el archivo `json_example.json`. El argumento `ensure_ascii=False` se utiliza para permitir la escritura de caracteres no ASCII
en el archivo. El argumento `indent=4` se utiliza para especificar la cantidad de espacios en blanco que se utilizarán para la
indentación del archivo.
"""

# Archivo con extensión csv: CSV significa valores separados por comas. CSV es un formato de archivo sencillo utilizado para
# almacenar datos tabulares, como una hoja de cálculo o una base de datos. CSV es un formato de datos muy común en la ciencia de
# datos.

import csv

with open('C:/Users/alvan/Downloads/F500.csv', mode='r', newline='') as file:
    lector_csv = csv.reader(file)
    for fila in lector_csv:
        print(fila)

"""
import csv: Importa el módulo `csv` de Python, que proporciona soporte para la lectura y escritura de archivos CSV.

with open('C:/Users/alvan/Downloads/F500.csv', mode='r', newline='') as file: Abre el archivo `F500.csv` que se encuentra en la
ruta `C:/Users/alvan/Downloads/` utilizando la función `open()` de Python. El modo 'r' indica que el archivo se abrirá en modo
lectura. El argumento `newline='' se utiliza para evitar problemas con el manejo de saltos de línea en diferentes sistemas
operativos. `with ... as file:`Utiliza la cláusula `with` para asegurarse de que el archivo se cierre correctamente después de
leer su contenido. El archivo se asigna a la variable `file`.

lector_csv = csv.reader(file): `csv.reader(file)`: Crea un objeto `reader` de CSV utilizando la función `csv.reader()`. El objeto
`reader` se utiliza para leer el contenido del archivo CSV línea por línea. El objeto `reader` se asigna a la variable `lector_csv`.

for fila in lector_csv:
    print(fila)
`for fila in lector_csv:`: Utiliza un bucle `for` para iterar sobre las filas del archivo CSV. Cada fila se almacena en la 
variable `fila`.
`print(fila)`: Imprime cada fila en la consola utilizando la función `print()`. Cada fila se imprime como una lista de valores separados por comas.
"""


import csv
with open('C:/Users/alvan/Downloads/F500.csv') as f:
    csv_reader = csv.reader(f, delimiter=',') 
    line_count = 0
    for row in csv_reader:
        if line_count == 0:
            print(f'Column names are :{", ".join(row)}')
            line_count += 1
        else:
            print(
                f'\t{row[0]} is a teachers. He lives in {row[1]}, {row[2]}.')
            line_count += 1
    print(f'Number of lines:  {line_count}')

"""
`import csv`: Importa el módulo `csv` de Python, que proporciona soporte para la lectura y escritura de archivos CSV.

with open('C:/Users/alvan/Downloads/F500.csv') as f:  `open('C:/Users/alvan/Downloads/F500.csv')`: Abre el archivo `F500.csv` que 
se encuentra en la ruta `C:/Users/alvan/Downloads/` utilizando la función `open()` de Python. El archivo se abre en modo lectura 
por defecto. Utilizamos la cláusula `with` para asegurarnos de que el archivo se cierre correctamente después de leer su contenido. 
El archivo se asigna a la variable `f`.

csv_reader = csv.reader(f, delimiter=',') : `csv.reader(f, delimiter=',')`: Crea un objeto `reader` de CSV utilizando la función 
`csv.reader()`. El objeto `reader` se utiliza para leer el contenido del archivo CSV línea por línea. El argumento `delimiter=','` 
se utiliza para especificar que el separador de campos es la coma. El objeto `reader` se asigna a la variable `csv_reader`.

line_count = 0 : `line_count = 0` Inicializa la variable `line_count` en cero. Esta variable se utiliza para contar el número de 
líneas en el archivo CSV.

for row in csv_reader:
    if line_count == 0:
        print(f'Column names are :{", ".join(row)}')
        line_count += 1
    else:
        print(
            f'\t{row[0]} is a teachers. He lives in {row[1]}, {row[2]}.')
        line_count += 1

 `for row in csv_reader:`: Utiliza un bucle `for` para iterar sobre las filas del archivo CSV. Cada fila se almacena en la 
 variable `row`.
`if line_count == 0:`: Verifica si la variable `line_count` es igual a cero. Si es así, se imprime el encabezado de las columnas.
`print(f'Column names are :{", ".join(row)}')`: Imprime el encabezado de las columnas en la consola utilizando la función `print()`. 
El método `join()` se utiliza para unir los elementos de la lista `row` con una coma.
`else:`: Si la variable `line_count` no es igual a cero, se imprime la información de cada fila.
`print(f'\t{row[0]} is a teachers. He lives in {row[1]}, {row[2]}.')`: Imprime la información de cada fila en la consola utilizando
la función `print()`.
`line_count += 1`: Incrementa la variable `line_count` en uno después de procesar cada fila.

print(f'Number of lines:  {line_count}') : Imprime el número total de líneas en el archivo CSV en la consola utilizando la función `print()`.
"""
# Archivo con extensión xlsx: Para leer archivos excel necesitamos instalar el paquete openpyxl

import openpyxl

excel_book = openpyxl.load_workbook('C:/Users/alvan/Downloads/Modelo Gasolina.xlsx')
print(f'Nombres de las hojas: {excel_book.sheetnames}')  # Nombres de las hojas: ['CRUDO', 'SNV', '2DERIV', 'SNV-2DERIV']
print(f'Número de hojas: {len(excel_book.sheetnames)}')  # Número de hojas: 4


"""
import openpyxl: Importa la biblioteca `openpyxl` de Python, que proporciona soporte para la lectura y escritura de archivos de Excel.

excel_book = openpyxl.load_workbook('C:/Users/alvan/Downloads/Modelo Gasolina.xlsx'): Abre el archivo `Modelo Gasolina.xlsx` que se 
encuentra en la ruta `C:/Users/alvan/Downloads/` utilizando la función `load_workbook()` de `openpyxl`. El archivo se asigna a la variable
`excel_book`.

'print(f'Nombres de las hojas: {excel_book.sheetnames}')`: Imprime los nombres de las hojas de cálculo en el libro de trabajo en la 
consola utilizando la función `print()`.

'print(f'Número de hojas: {len(excel_book.sheetnames)}'): Devuelve el número de hojas de cálculo en el libro de trabajo. La función 
`len()` se utiliza para obtener este número.
"""

# XML (Extensible Markup Language) es un lenguaje de marcado diseñado para almacenar y transportar datos de manera legible tanto por 
# máquinas como por humanos. XML permite a los usuarios definir sus propias etiquetas y estructuras de datos. Esto significa que puedes 
# crear un conjunto personalizado de etiquetas que se adapten a tus necesidades específicas. Los documentos XML son legibles por humanos, 
# lo que facilita su lectura y edición manual. Utiliza etiquetas y atributos para definir la estructura y el contenido de los datos.
# Los datos en un documento XML se organizan en una estructura jerárquica de elementos anidados. Esto permite representar relaciones 
# complejas entre datos.

# Como leer archivos XML: Para leer archivos XML necesitamos instalar el paquete xml.etree.ElementTree

import xml.etree.ElementTree as ET

tree = ET.parse('C:/Users/alvan/Downloads/xml_example.xml')
root = tree.getroot()
print(f'Nombre de la etiqueta raíz: {root.tag}')
print('Attribute:', root.attrib)
for child in root:
    print('field: ', child.tag)

"""
Nombre de la etiqueta raíz: person
Attribute: {'gender': 'male'}
field:  name
field:  country
field:  city
field:  skills
"""

"""
import xml.etree.ElementTree as ET : Importa la biblioteca `xml.etree.ElementTree` de Python, que proporciona soporte para la lectura y 
escritura de archivos XML.

`tree = ET.parse('C:/Users/alvan/Downloads/xml_example.xml')`: Abre el archivo `xml_example.xml` que se encuentra en la ruta `C:/Users/alvan/Downloads/` 
utilizando la función `parse()` de `xml.etree.ElementTree`. El archivo se asigna a la variable `tree`. 
- `root = tree.getroot()`: Devuelve la etiqueta raíz del archivo XML. La función `getroot()` se utiliza para obtener esta etiqueta. 
La etiqueta raíz se asigna a la variable `root`.

print(f'Nombre de la etiqueta raíz: {root.tag}'): '{root.tag}' Devuelve el nombre de la etiqueta raíz del archivo XML. La propiedad `tag` 
se utiliza para obtener este nombre, e imprime el nombre de la etiqueta raíz en la consola utilizando la función `print()`.

print('Attribute:', root.attrib): `root.attrib`: Devuelve un diccionario que contiene los atributos de la etiqueta raíz del archivo XML. 
La propiedad `attrib` se utiliza para obtener este diccionario e emprime los atributos de la etiqueta raíz en la consola utilizando la 
función `print()`.

for child in root:
    print('field: ', child.tag)
`for child in root:`: Utiliza un bucle `for` para iterar sobre los elementos secundarios de la etiqueta raíz del archivo XML. 
Cada elemento secundario se almacena en la variable `child`.
`child.tag`: Devuelve el nombre del elemento secundario. La propiedad `tag` se utiliza para obtener este nombre.
`print('field: ', child.tag)`: Imprime el nombre de cada elemento secundario en la consola utilizando la función `print()`.
"""

# Ejercicios:

# Escribe una función que cuente el número de líneas y el número de palabras del texto ubicado en
# 'C:/Users/alvan/Downloads/obama_speech.txt

def contar_lineas_y_palabras(archivo):
    try:
        with open(archivo, 'r', encoding='utf-8') as file:
            contenido = file.read()
            lineas = contenido.split('\n')
            palabras = contenido.split()

        num_lineas = len(lineas)
        num_palabras = len(palabras)

        return num_lineas, num_palabras
    except FileNotFoundError:
        return "El archivo no se encontró."

archivo_texto = 'C:/Users/alvan/Downloads/obama_speech.txt'

lineas, palabras = contar_lineas_y_palabras(archivo_texto)

if isinstance(lineas, int) and isinstance(palabras, int):
    print(f"Número de líneas: {lineas}")  # Número de líneas: 67
    print(f"Número de palabras: {palabras}") # Número de palabras: 2400
else:
    print(lineas)

"""
def contar_lineas_y_palabras(archivo):
    try:
        with open(archivo, 'r', encoding='utf-8') as file:
            contenido = file.read()
            lineas = contenido.split('\n')
            palabras = contenido.split()

        num_lineas = len(lineas)
        num_palabras = len(palabras)

        return num_lineas, num_palabras
    except FileNotFoundError:
        return "El archivo no se encontró."
```
- `def contar_lineas_y_palabras(archivo):`: Define una función llamada `contar_lineas_y_palabras()` que toma un argumento llamado `archivo`.
- `with open(archivo, 'r', encoding='utf-8') as file:`: Abre el archivo especificado en el argumento `archivo` en modo lectura (`'r'`) y lo 
asigna a la variable `file`.El argumento `encoding='utf-8'` se utiliza para especificar la codificación de caracteres del archivo. La cláusula 
`with` se utiliza para asegurarse de que el archivo se cierre correctamente después de leerlo.
- `contenido = file.read()`: Lee el contenido del archivo y lo asigna a la variable `contenido`.
- `lineas = contenido.split('\n')`: Divide el contenido del archivo en una lista de líneas utilizando el carácter de nueva línea (`'\n'`) 
como separador y lo asigna a la variable `lineas`.
- `palabras = contenido.split()`: Divide el contenido del archivo en una lista de palabras utilizando los espacios en blanco como separadores 
y lo asigna a la variable `palabras`.
- `num_lineas = len(lineas)`: Cuenta el número de elementos en la lista `lineas` y lo asigna a la variable `num_lineas`.
- `num_palabras = len(palabras)`: Cuenta el número de elementos en la lista `palabras` y lo asigna a la variable `num_palabras`.
- `return num_lineas, num_palabras`: Devuelve una tupla que contiene el número de líneas y palabras en el archivo.

except FileNotFoundError: Si se produce error porque el archivo no es encontrado ... 
        return "El archivo no se encontró."

- `archivo_texto = 'C:/Users/alvan/Downloads/obama_speech.txt'`: Asigna la ruta del archivo `obama_speech.txt` a la variable `archivo_texto`.
- `lineas, palabras = contar_lineas_y_palabras(archivo_texto)`: Llama a la función `contar_lineas_y_palabras()` con `archivo_texto` como argumento
y asigna los valores devueltos a las variables `lineas` y `palabras`.

if isinstance(lineas, int) and isinstance(palabras, int):
    print(f"Número de líneas: {lineas}")  
    print(f"Número de palabras: {palabras}") 
else:
    print(lineas)

- `isinstance(lineas, int) and isinstance(palabras, int)`: Verifica si `lineas` y `palabras` son enteros utilizando la función `isinstance()`.
- `print(f"Número de líneas: {lineas}")`: Imprime el número de líneas en el archivo en la consola utilizando la función `print()`.
- `print(f"Número de palabras: {palabras}")`: Imprime el número de palabras en el archivo en la consola utilizando la función `print()`.
- `print(lineas)`: Si `lineas` y `palabras` no son enteros, imprime el mensaje de error devuelto por la función `contar_lineas_y_palabras()` 
en la consola utilizando la función `print()`.
"""
# Escribe una función que cuente el número de líneas y el número de palabras del texto ubicado en
# 'C:/Users/alvan/Downloads/donald_speech.txt

def contar_lineas_y_palabras(archivo):
    try:
        with open(archivo, 'r', encoding='utf-8') as file:
            contenido = file.read()
            lineas = contenido.split('\n')
            palabras = contenido.split()

        num_lineas = len(lineas)
        num_palabras = len(palabras)

        return num_lineas, num_palabras
    except FileNotFoundError:
        return "El archivo no se encontró."

archivo_texto = 'C:/Users/alvan/Downloads/donald_speech.txt'

lineas, palabras = contar_lineas_y_palabras(archivo_texto)

if isinstance(lineas, int) and isinstance(palabras, int):
    print(f"Número de líneas: {lineas}")  # Número de líneas: 49
    print(f"Número de palabras: {palabras}") # Número de palabras: 1259
else:
    print(lineas)
    
# Escribe una función que cuente el número de líneas y el número de palabras del texto ubicado en
# 'C:/Users/alvan/Downloads/melina_trump_speech.txt

def contar_lineas_y_palabras(archivo):
    try:
        with open(archivo, 'r', encoding='utf-8') as file:
            contenido = file.read()
            lineas = contenido.split('\n')
            palabras = contenido.split()

        num_lineas = len(lineas)
        num_palabras = len(palabras)

        return num_lineas, num_palabras
    except FileNotFoundError:
        return "El archivo no se encontró."

archivo_texto = 'C:/Users/alvan/Downloads/melina_trump_speech.txt'

lineas, palabras = contar_lineas_y_palabras(archivo_texto)

if isinstance(lineas, int) and isinstance(palabras, int):
    print(f"Número de líneas: {lineas}")  # Número de líneas: 33
    print(f"Número de palabras: {palabras}") # Número de palabras: 1375
else:
    print(lineas)

# Lea el archivo de datos countries_data.json en el directorio data, cree una función que encuentre los diez idiomas más hablados.
# JSON (JavaScript Object Notation) es un formato de intercambio de datos ligero y ampliamente utilizado. Es fácil de leer y escribir para 
# los seres humanos, y fácil de analizar y generar para las máquinas. JSON se utiliza para transmitir datos estructurados entre un servidor y 
# un cliente, o entre diferentes partes de una aplicación. JSON admite varios tipos de datos, incluyendo números, cadenas, booleanos, objetos, 
# matrices y valores nulos. Esto permite representar datos complejos de manera estructurada.

#Ejemplos de archivos JSON:

{
  "nombre": "John Doe",
  "edad": 30,
  "ciudad": "Nueva York",
  "casado": "false",
  "hobbies": ["lectura", "ciclismo", "viajar"]
}

# Otro ejemplo:

{
  "numeros": [1, 2, 3, 4, 5],
  "otros_numeros": [10, 20, 30, 40, 50]
}


import json

try:
    with open('C:/Users/alvan/Downloads/countries_data.json', 'r', encoding='utf-8') as file:
        # Carga el contenido del archivo JSON en un diccionario
        datos = json.load(file)
except FileNotFoundError:
    print("El archivo no se encontró.")
    datos = []
languages_count = {}
# Recorrer cada país y contar el número de lenguas
for country in datos:
    for language in country['languages']:
        if language in languages_count:
            languages_count[language] += 1
        else:
            languages_count[language] = 1
# Ordenar el diccionario por cuenta de lengua
sorted_languages = sorted(languages_count.items(), key=lambda x: x[1], reverse=True)
# Imprimir las diez lenguas más habladas
print("Las diez lenguas más habladas en el mundo son:")
for language, count in sorted_languages[:10]:
    print(f"{language} ({count} países)")

"""
Las diez lenguas más habladas en el mundo son:
English (91 países)
French (45 países)
Arabic (25 países)
Spanish (24 países)
Portuguese (9 países)
Russian (9 países)
Dutch (8 países)
German (7 países)
Chinese (5 países)
Serbian (4 países)
"""
# Otra forma

import json

def encontrar_diez_idiomas_mas_hablados(archivo):
    try:
        with open('C:/Users/alvan/Downloads/countries_data.json', 'r', encoding='utf-8') as file:
            # Carga el contenido del archivo JSON en un diccionario
            data = json.load(file)

        # Crear un diccionario para contar la frecuencia de cada idioma
        idioma_contador = {}

        # Recorre los datos de los países
        for pais in data:
            idiomas = pais.get("languages", [])  # Obtén la lista de idiomas del país
            for idioma in idiomas:
                if idioma in idioma_contador:
                    idioma_contador[idioma] += 1
                else:
                    idioma_contador[idioma] = 1

        # Ordena el diccionario por la frecuencia de los idiomas en orden descendente
        idioma_mas_hablado = sorted(idioma_contador.items(), key=lambda x: x[1], reverse=True)

        # Devuelve los diez idiomas más hablados
        return idioma_mas_hablado[:10]

    except FileNotFoundError:
        return "El archivo JSON no se encontró."
    except json.JSONDecodeError:
        return "Error al decodificar el archivo JSON."
    except Exception as e:
        return f"Se produjo un error: {str(e)}"

# Llama a la función e imprime los diez idiomas más hablados
resultados = encontrar_diez_idiomas_mas_hablados('C:/Users/alvan/Downloads/countries_data.json')
if isinstance(resultados, list):
    print("Los diez idiomas más hablados son:")
    for idioma, frecuencia in resultados:
        print(f"{idioma}: {frecuencia} países")
else:
    print(resultados)

"""
Los diez idiomas más hablados son:
English: 91 países
French: 45 países
Arabic: 25 países
Spanish: 24 países
Portuguese: 9 países
Russian: 9 países
Dutch: 8 países
German: 7 países
Chinese: 5 países
Serbian: 4 países
"""
# Otra froma:

import json
from collections import Counter
import operator

def encontrar_diez_idiomas_mas_hablados(archivo):
    try:
        with open(archivo, 'r', encoding='utf-8') as file:
            # Carga el contenido del archivo JSON en un diccionario
            data = json.load(file)

        # Obtén la lista de idiomas de todos los países
        idiomas = [idioma for pais in data for idioma in pais.get("languages", [])]

        # Cuenta la frecuencia de cada idioma utilizando Counter()
        idioma_contador = Counter(idiomas)

        # Ordena el diccionario por la frecuencia de los idiomas en orden descendente
        idioma_mas_hablado = sorted(idioma_contador.items(), key=operator.itemgetter(1), reverse=True)

        # Devuelve los diez idiomas más hablados
        return idioma_mas_hablado[:10]

    except FileNotFoundError:
        return "El archivo JSON no se encontró."
    except json.JSONDecodeError:
        return "Error al decodificar el archivo JSON."
    except Exception as e:
        return f"Se produjo un error: {str(e)}"

# Llama a la función e imprime los diez idiomas más hablados
resultados = encontrar_diez_idiomas_mas_hablados('C:/Users/alvan/Downloads/countries_data.json')
if isinstance(resultados, list):
    print("Los diez idiomas más hablados son:")
    for idioma, frecuencia in resultados:
        print(f"{idioma}: {frecuencia} países")
else:
    print(resultados)
    
"""
Los diez idiomas más hablados son:
English: 91 países
French: 45 países
Arabic: 25 países
Spanish: 24 países
Portuguese: 9 países
Russian: 9 países
Dutch: 8 países
German: 7 países
Chinese: 5 países
Serbian: 4 países
"""

# Lea el archivo JSON ubicado en 'C:/Users/alvan/Downloads/countries_data.json' y cree una función que que cree una lista de los diez países 
# más poblados.

import json

def obtener_diez_paises_mas_poblados(archivo):
    try:
        with open('C:/Users/alvan/Downloads/countries_data.json', 'r', encoding='utf-8') as file:
            # Carga el contenido del archivo JSON en una lista de diccionarios
            data = json.load(file)

        # Ordena la lista de países por población en orden descendente
        paises_ordenados = sorted(data, key=lambda x: x.get("population", 0), reverse=True)

        # Obtiene los diez países más poblados
        diez_paises_mas_poblados = paises_ordenados[:10]

        return diez_paises_mas_poblados

    except FileNotFoundError:
        return "El archivo JSON no se encontró."
    except json.JSONDecodeError:
        return "Error al decodificar el archivo JSON."
    except Exception as e:
        return f"Se produjo un error: {str(e)}"

# Llama a la función e imprime los diez países más poblados
resultados = obtener_diez_paises_mas_poblados('C:/Users/alvan/Downloads/countries_data.json')
if isinstance(resultados, list):
    print("Los diez países más poblados son:")
    for pais in resultados:
        print(f"{pais['name']}: {pais['population']} habitantes")
else:
    print(resultados)

"""
Los diez países más poblados son:
China: 1377422166 habitantes
India: 1295210000 habitantes
United States of America: 323947000 habitantes
Indonesia: 258705000 habitantes
Brazil: 206135893 habitantes
Pakistan: 194125062 habitantes
Nigeria: 186988000 habitantes
Bangladesh: 161006790 habitantes
Russian Federation: 146599183 habitantes
Japan: 126960000 habitantes
"""
# Otra forma:

import json

def obtener_diez_paises_mas_poblados(archivo):
    try:
        with open('C:/Users/alvan/Downloads/countries_data.json', 'r', encoding='utf-8') as file:
            # Carga el contenido del archivo JSON en una lista de diccionarios
            data = json.load(file)

        # Ordena la lista de países por población en orden descendente
        paises_ordenados = sorted(data, key=lambda x: x.get("population", 0), reverse=True)

        return paises_ordenados[:10]

    except FileNotFoundError:
        return "El archivo JSON no se encontró."
    except json.JSONDecodeError:
        return "Error al decodificar el archivo JSON."
    except Exception as e:
        return f"Se produjo un error: {str(e)}"

# Llama a la función e imprime los diez países más poblados
resultados = obtener_diez_paises_mas_poblados('C:/Users/alvan/Downloads/countries_data.json')
if isinstance(resultados, list):
    print("Los diez países más poblados son:")
    for pais in resultados:
        print(f"{pais['name']}: {pais['population']} habitantes")
else:
    print(resultados)

"""
Los diez países más poblados son:
China: 1377422166 habitantes
India: 1295210000 habitantes
United States of America: 323947000 habitantes
Indonesia: 258705000 habitantes
Brazil: 206135893 habitantes
Pakistan: 194125062 habitantes
Nigeria: 186988000 habitantes
Bangladesh: 161006790 habitantes
Russian Federation: 146599183 habitantes
Japan: 126960000 habitantes
"""

# Extraiga todas las direcciones de correo electrónico entrantes en forma de lista del archivo ubicado en 
# C:/Users/alvan/Downloads/email_exchanges_big.txt

import re

def extraer_direcciones_de_correo(archivo):
    direcciones_de_correo = []
    try:
        with open('C:/Users/alvan/Downloads/email_exchanges_big.txt', 'r', encoding='utf-8') as file:
            # Leer el contenido del archivo de texto
            texto = file.read()

            # Utilizar una expresión regular para encontrar direcciones de correo electrónico
            # Esta expresión regular busca patrones típicos de direcciones de correo
            patron = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b'
            direcciones_de_correo = re.findall(patron, texto)

        return direcciones_de_correo

    except FileNotFoundError:
        return "El archivo de texto no se encontró."
    except Exception as e:
        return f"Se produjo un error: {str(e)}"

# Llama a la función e imprime la lista de direcciones de correo electrónico
resultados = extraer_direcciones_de_correo('C:/Users/alvan/Downloads/email_exchanges_big.txt')
if isinstance(resultados, list):
    print("Direcciones de correo electrónico encontradas:")
    for direccion in resultados:
        print(direccion)
else:
    print(resultados)
    
# Encontrar las palabras más comunes en el idioma Inglés, en el archivo ubicado en 'C:/Users/alvan/Downloads/email_exchanges_big.txt'. 
# Llame a su función palabras_mas_comunes, tomará dos parámetros - una cadena o un archivo y un número entero positivo, indicando el número 
# de palabras. Su función devolverá una matriz de tuplas en orden descendente.

import re
from collections import Counter

def palabras_mas_comunes(archivo, top_n):
    try:
        with open('C:/Users/alvan/Downloads/email_exchanges_big.txt', 'r', encoding='utf-8') as file:
            texto = file.read()
    except FileNotFoundError:
        return "El archivo no se encontró."

    # Tokenizar el texto en palabras usando expresiones regulares
    palabras = re.findall(r'\b\w+\b', texto.lower())
    
    # Obtener las palabras comunes en inglés
    with open('C:/Users/alvan/Downloads/email_exchanges_big.txt', 'r') as common_words_file:
        common_words = set(common_words_file.read().splitlines())

    # Filtrar las palabras que no son stop words
    palabras_filtradas = [palabra for palabra in palabras if palabra not in common_words]

    # Contar la frecuencia de cada palabra
    contador = Counter(palabras_filtradas)

    # Obtener las palabras más comunes en orden descendente
    palabras_comunes = contador.most_common(top_n)

    return palabras_comunes

# Llama a la función para obtener las 10 palabras más comunes en inglés
resultados = palabras_mas_comunes('C:/Users/alvan/Downloads/email_exchanges_big.txt', 10)

if isinstance(resultados, list):
    print(resultados)
else:
    print(resultados)

# [('edu', 31260), ('2007', 24480), ('org', 22456), ('sakaiproject', 21747), ('from', 21721), ('by', 18028), ('collab', 17970), ('x', 16677), ('received', 16176), ('0', 16061)]

# Si deseamos cambiar el orden de los elmentos de la tupla, debemos cambiar la linea que devuelve las palabras comunes:

import re
from collections import Counter

def palabras_mas_comunes(archivo, top_n):
    try:
        with open('C:/Users/alvan/Downloads/email_exchanges_big.txt', 'r', encoding='utf-8') as file:
            texto = file.read()
    except FileNotFoundError:
        return "El archivo no se encontró."

    # Tokenizar el texto en palabras usando expresiones regulares
    palabras = re.findall(r'\b\w+\b', texto.lower())
    
    # Obtener las palabras comunes en inglés
    with open('C:/Users/alvan/Downloads/email_exchanges_big.txt', 'r') as common_words_file:
        common_words = set(common_words_file.read().splitlines())

    # Filtrar las palabras que no son stop words
    palabras_filtradas = [palabra for palabra in palabras if palabra not in common_words]

    # Contar la frecuencia de cada palabra
    contador = Counter(palabras_filtradas)

    # Obtener las palabras más comunes en orden descendente
    palabras_comunes = contador.most_common(top_n)

    # Cambiar el orden de los elementos en la tupla
    palabras_comunes = [(frecuencia, palabra) for palabra, frecuencia in palabras_comunes]

    return palabras_comunes

# Llama a la función para obtener las 10 palabras más comunes en inglés
resultados = palabras_mas_comunes('C:/Users/alvan/Downloads/email_exchanges_big.txt', 10)

if isinstance(resultados, list):
    print(resultados)
else:
    print(resultados)

# [(31260, 'edu'), (24480, '2007'), (22456, 'org'), (21747, 'sakaiproject'), (21721, 'from'), (18028, 'by'), (17970, 'collab'), (16677, 'x'), (16176, 'received'), (16061, '0')]

# Utilice la función, palabras_más_comunes para encontrar: a) Las diez palabras más frecuentes utilizadas en el discurso de Obama "obama_speech.txt"
# b) Las diez palabras más frecuentes utilizadas en el discurso de Michelle "michelle_obama_speech.txt" c) Las diez palabras más frecuentes utilizadas
# en el discurso de Trump "donald_speech.txt" d) Las diez palabras más frecuentes utilizadas en el discurso de Melina "melina_trump_speech.txt"

# 10 palabras mas comunes discurso de Obama 

import re
from collections import Counter

def palabras_mas_comunes(archivo, top_n):
    try:
        with open('C:/Users/alvan/Downloads/obama_speech.txt', 'r', encoding='utf-8') as file:
            texto = file.read()
    except FileNotFoundError:
        return "El archivo no se encontró."

    # Tokenizar el texto en palabras usando expresiones regulares
    palabras = re.findall(r'\b\w+\b', texto.lower())
    
    # Obtener las palabras comunes en inglés
    with open('C:/Users/alvan/Downloads/obama_speech.txt', 'r') as common_words_file:
        common_words = set(common_words_file.read().splitlines())

    # Filtrar las palabras que no son stop words
    palabras_filtradas = [palabra for palabra in palabras if palabra not in common_words]

    # Contar la frecuencia de cada palabra
    contador = Counter(palabras_filtradas)

    # Obtener las palabras más comunes en orden descendente
    palabras_comunes = contador.most_common(top_n)

    # Cambiar el orden de los elementos en la tupla
    palabras_comunes = [(frecuencia, palabra) for palabra, frecuencia in palabras_comunes]

    return palabras_comunes

# Llama a la función para obtener las 10 palabras más comunes en inglés
resultados = palabras_mas_comunes('C:/Users/alvan/Downloads/obama_speech.txt', 10)

if isinstance(resultados, list):
    print(resultados)
else:
    print(resultados)
    
# [(129, 'the'), (113, 'and'), (81, 'of'), (70, 'to'), (67, 'our'), (62, 'we'), (50, 'that'), (48, 'a'), (36, 'is'), (25, 'in')]

# 10 palabras mas comunes discurso de Michelle

import re
from collections import Counter

def palabras_mas_comunes(archivo, top_n):
    try:
        with open('C:/Users/alvan/Downloads/michelle_obama_speech.txt', 'r', encoding='utf-8') as file:
            texto = file.read()
    except FileNotFoundError:
        return "El archivo no se encontró."

    # Tokenizar el texto en palabras usando expresiones regulares
    palabras = re.findall(r'\b\w+\b', texto.lower())
    
    # Obtener las palabras comunes en inglés
    with open('C:/Users/alvan/Downloads/michelle_obama_speech.txt', 'r') as common_words_file:
        common_words = set(common_words_file.read().splitlines())

    # Filtrar las palabras que no son stop words
    palabras_filtradas = [palabra for palabra in palabras if palabra not in common_words]

    # Contar la frecuencia de cada palabra
    contador = Counter(palabras_filtradas)

    # Obtener las palabras más comunes en orden descendente
    palabras_comunes = contador.most_common(top_n)

    # Cambiar el orden de los elementos en la tupla
    palabras_comunes = [(frecuencia, palabra) for palabra, frecuencia in palabras_comunes]

    return palabras_comunes

# Llama a la función para obtener las 10 palabras más comunes en inglés
resultados = palabras_mas_comunes('C:/Users/alvan/Downloads/michelle_obama_speech.txt', 10)

if isinstance(resultados, list):
    print(resultados)
else:
    print(resultados)
    
# [(96, 'and'), (85, 'the'), (84, 'to'), (50, 'that'), (46, 'of'), (41, 'a'), (37, 'he'), (36, 'in'), (28, 'my'), (28, 'i')#]

# 10 palabras mas comunes discurso de Trump

import re
from collections import Counter

def palabras_mas_comunes(archivo, top_n):
    try:
        with open('C:/Users/alvan/Downloads/donald_speech.txt', 'r', encoding='utf-8') as file:
            texto = file.read()
    except FileNotFoundError:
        return "El archivo no se encontró."

    # Tokenizar el texto en palabras usando expresiones regulares
    palabras = re.findall(r'\b\w+\b', texto.lower())
    
    # Obtener las palabras comunes en inglés
    with open('C:/Users/alvan/Downloads/donald_speech.txt', 'r') as common_words_file:
        common_words = set(common_words_file.read().splitlines())

    # Filtrar las palabras que no son stop words
    palabras_filtradas = [palabra for palabra in palabras if palabra not in common_words]

    # Contar la frecuencia de cada palabra
    contador = Counter(palabras_filtradas)

    # Obtener las palabras más comunes en orden descendente
    palabras_comunes = contador.most_common(top_n)

    # Cambiar el orden de los elementos en la tupla
    palabras_comunes = [(frecuencia, palabra) for palabra, frecuencia in palabras_comunes]

    return palabras_comunes

# Llama a la función para obtener las 10 palabras más comunes en inglés
resultados = palabras_mas_comunes('C:/Users/alvan/Downloads/donald_speech.txt', 10)

if isinstance(resultados, list):
    print(resultados)
else:
    print(resultados)
    
# [(65, 'the'), (59, 'and'), (44, 'we'), (40, 'will'), (38, 'of'), (32, 'to'), (30, 'our'), (20, 'is'), (17, 'america'), (13, 'for')]

# 10 palabras mas comunes discurso de Melina

import re
from collections import Counter

def palabras_mas_comunes(archivo, top_n):
    try:
        with open('C:/Users/alvan/Downloads/melina_trump_speech.txt', 'r', encoding='utf-8') as file:
            texto = file.read()
    except FileNotFoundError:
        return "El archivo no se encontró."

    # Tokenizar el texto en palabras usando expresiones regulares
    palabras = re.findall(r'\b\w+\b', texto.lower())
    
    # Obtener las palabras comunes en inglés
    with open('C:/Users/alvan/Downloads/melina_trump_speech.txt', 'r') as common_words_file:
        common_words = set(common_words_file.read().splitlines())

    # Filtrar las palabras que no son stop words
    palabras_filtradas = [palabra for palabra in palabras if palabra not in common_words]

    # Contar la frecuencia de cada palabra
    contador = Counter(palabras_filtradas)

    # Obtener las palabras más comunes en orden descendente
    palabras_comunes = contador.most_common(top_n)

    # Cambiar el orden de los elementos en la tupla
    palabras_comunes = [(frecuencia, palabra) for palabra, frecuencia in palabras_comunes]

    return palabras_comunes

# Llama a la función para obtener las 10 palabras más comunes en inglés
resultados = palabras_mas_comunes('C:/Users/alvan/Downloads/melina_trump_speech.txt', 10)

if isinstance(resultados, list):
    print(resultados)
else:
    print(resultados)
    
# [(77, 'and'), (55, 'to'), (52, 'the'), (29, 'is'), (28, 'i'), (27, 'for'), (25, 'of'), (24, 'that'), (22, 'a'), (21, 'you')]

# Escribe una aplicación python que compruebe la similitud entre dos textos. Toma un archivo o una cadena como parámetro y evaluará la similitud
# de los dos textos. Por ejemplo, comprueba la similitud entre las transcripciones del discurso de Michelle y Melina en los archivos 
# "michelle_obama_speech.txt" y "melina_trump_speech.txt". Es posible que necesite un par de funciones, una para limpiar el texto (clean_text), 
# otra para eliminar las palabras de parada y, por último, otra para comprobar la similitud. La lista de palabras de parada se encuentran en 
# C:/Users/alvan/Downloads/Stop words.txt"

import re

# Cargar palabras de parada
with open("C:/Users/alvan/Downloads/Stop words.txt", "r") as f:
    stop_words = set(f.read().splitlines())

# Limpie texto
def clean_text(text):
    """
    Limpie el texto eliminando la puntuacion, palabras de parada y numeros.

    Args:
        text: El texto a limpiar.

    Returns:
        El texto depurado.
    """

    text = re.sub(r"[^\w\s]", "", text)
    text = " ".join([word for word in text.split() if word not in stop_words])
    text = re.sub(r"\d+", "", text)

    return text

# Comprobar la similitud de los textos
def check_text_similarity(text1, text2):
    """
    Comprueba la similitud entre dos textos.

    Args:
        text1: El primer texto.
        text2: El segundo texto.

    Returns:
        Una puntuación de similitud entre los dos textos.
    """

    # Limpiar los textos
    text1 = clean_text(text1)
    text2 = clean_text(text2)

    # Convertir los textos en conjuntos de palabras
    words1 = set(text1.split())
    words2 = set(text2.split())

    # Calcule la puntuación de similitud
    similarity = len(words1 & words2) / len(words1 | words2)

    return similarity

# Comprueba la similitud de los discursos de Michelle y Melina
michelle_speech = ""
with open("C:/Users/alvan/Downloads/michelle_obama_speech.txt", "r") as f:
    michelle_speech = f.read()

melina_speech = ""
with open("C:/Users/alvan/Downloads/melina_trump_speech.txt", "r") as f:
    melina_speech = f.read()

similarity = check_text_similarity(michelle_speech, melina_speech)

print(f"La similitud entre los discursos de Michelle y Melina es de {similarity}")
# La similitud entre los discursos de Michelle y Melina es de 0.24148606811145512

# Encuentre las 10 palabras más repetidas en "C:/Users/alvan/Downloads/romeo_and_juliet.txt"

import re
from collections import Counter

romeo_julieta = ('C:/Users/alvan/Downloads/romeo_and_juliet.txt')
def palabras_mas_comunes(archivo, top_n):
    try:
        with open(romeo_julieta, 'r', encoding='utf-8') as file:
            texto = file.read()
    except FileNotFoundError:
        return "El archivo no se encontró."

    # Tokenizar el texto en palabras usando expresiones regulares
    palabras = re.findall(r'\b\w+\b', texto.lower())
    
    # Obtener las palabras comunes en inglés
    with open(romeo_julieta, 'r') as common_words_file:
        common_words = set(common_words_file.read().splitlines())

    # Filtrar las palabras que no son stop words
    palabras_filtradas = [palabra for palabra in palabras if palabra not in common_words]

    # Contar la frecuencia de cada palabra
    contador = Counter(palabras_filtradas)

    # Obtener las palabras más comunes en orden descendente
    palabras_comunes = contador.most_common(top_n)

    # Cambiar el orden de los elementos en la tupla
    palabras_comunes = [(frecuencia, palabra) for palabra, frecuencia in palabras_comunes]

    return palabras_comunes

# Llama a la función para obtener las 10 palabras más comunes
resultados = palabras_mas_comunes(romeo_julieta, 10)

if isinstance(resultados, list):
    print(resultados)
else:
    print(resultados)
# [(868, 'the'), (800, 'and'), (661, 'to'), (658, 'i'), (535, 'of'), (381, 'is'), (378, 'in'), (367, 'you'), (360, 'my'), (305, 's')]

# otra froma mas sencilla:

import re
from collections import Counter

# Abre el archivo y lee su contenido
with open('C:/Users/alvan/Downloads/romeo_and_juliet.txt', 'r', encoding='utf-8') as file:
    texto = file.read()

# Limpia el texto y divide en palabras
palabras = re.findall(r'\b\w+\b', texto.lower())

# Cuenta la frecuencia de cada palabra
frecuencia_palabras = Counter(palabras)

# Encuentra las 10 palabras más repetidas
palabras_mas_repetidas = frecuencia_palabras.most_common(10)

# Imprime las 10 palabras más repetidas
for palabra, frecuencia in palabras_mas_repetidas:
    print(f'{palabra}: {frecuencia}')
"""
the: 868
and: 800
to: 661
i: 658
of: 535
a: 530
is: 381
in: 378
that: 371
you: 367
"""

# Lee el archivo csv ubicado en 'C:/Users/alvan/Downloads/hacker_news.csv' y averigua: a) Cuenta el número de líneas que contienen "python" 
# o" Python" b) Cuenta el número de líneas que contienen "JavaScript", "javascript" o "Javascript" c) Cuenta el número de líneas que contienen 
# "Java" y no contienen "JavaScript"

import csv
import re

# Inicializa contadores
count_python = 0
count_javascript = 0
count_java_not_javascript = 0

# Abre el archivo CSV
with open('C:/Users/alvan/Downloads/hacker_news.csv', 'r', encoding='utf-8') as file:
    csv_reader = csv.reader(file)

    # Itera sobre las líneas del archivo
    for row in csv_reader:
        if len(row) > 2:  # Asegura que la fila tenga al menos 3 columnas
            text = row[2]  # Selecciona la tercera columna que contiene el texto

            # Conteo de líneas que contienen "Python" o "python"
            if re.search(r'python', text, re.IGNORECASE):
                count_python += 1

            # Conteo de líneas que contienen "JavaScript", "javascript" o "Javascript"
            if re.search(r'javascript', text, re.IGNORECASE):
                count_javascript += 1

            # Conteo de líneas que contienen "Java" y no contienen "JavaScript"
            if re.search(r'java', text, re.IGNORECASE) and not re.search(r'javascript', text, re.IGNORECASE):
                count_java_not_javascript += 1

# Imprime los resultados
print(f'Líneas que contienen "Python" o "python": {count_python}')
print(f'Líneas que contienen "JavaScript", "javascript" o "Javascript": {count_javascript}')
print(f'Líneas que contienen "Java" y no contienen "JavaScript": {count_java_not_javascript}')
"""
Líneas que contienen "Python" o "python": 101
Líneas que contienen "JavaScript", "javascript" o "Javascript": 70
Líneas que contienen "Java" y no contienen "JavaScript": 35
"""