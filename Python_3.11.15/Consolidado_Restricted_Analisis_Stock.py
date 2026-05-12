import pandas as pd
import numpy as np
import re
from pathlib import Path
from paddleocr import PPStructure
from openpyxl import load_workbook
from openpyxl.styles import numbers

# ============================================
# CONFIGURACIÓN
# ============================================

BASE_DIR = Path("/mnt/c/Users/Alvaro/Downloads")

IMAGES = [
    BASE_DIR / "Imagen 1.jpg",
    BASE_DIR / "Imagen 2.jpg",
    BASE_DIR / "Imagen 3.jpg"
]

OUTPUT_CSV = BASE_DIR / "Restricted1.csv"
OUTPUT_XLSX = BASE_DIR / "Restricted1.xlsx"

# ============================================
# OCR ENGINE
# ============================================

table_engine = PPStructure(
    show_log=False,
    lang='en'
)

# ============================================
# FUNCIONES
# ============================================

def limpiar_numero(valor):

    if pd.isna(valor):
        return "NA"

    valor = str(valor).strip()

    if valor in ["-", "", "—", "nan"]:
        return "NA"

    valor = valor.replace(",", "")

    try:

        # PORCENTAJE
        if "%" in valor:
            return float(valor.replace("%", "")) / 100

        # BILLIONS
        if "B" in valor:
            numero = float(valor.replace("B", "").strip())
            return numero * 1_000_000_000

        # MILLIONS
        if "M" in valor:
            numero = float(valor.replace("M", "").strip())
            return numero * 1_000_000

        # THOUSANDS
        if "K" in valor:
            numero = float(valor.replace("K", "").strip())
            return numero * 1_000

        return float(valor)

    except:
        return valor


def renombrar_columnas(df):

    rename_map = {

        "Revenue Forecast Count of":
            "Revenue Forecast Count of Estimates",

        "Change In Accounts":
            "Change In Accounts Receivable",

        "Net Income to Common Excl":
            "Net Income to Common Excl Extra Items",

        "Merger & Related Restructuring":
            "Merger & Related Restructuring Charges"
    }

    df.rename(columns=rename_map, inplace=True)

    return df


def limpiar_dataframe(df):

    # Full ticker uppercase
    if "Full Ticker" in df.columns:

        df["Full Ticker"] = (
            df["Full Ticker"]
            .astype(str)
            .str.upper()
        )

    # Convertir números financieros
    for col in df.columns:

        if col in ["Name", "Full Ticker"]:
            continue

        df[col] = df[col].apply(limpiar_numero)

    return df


def extraer_tabla(imagen_path):

    print(f"\nProcesando: {imagen_path.name}")

    result = table_engine(str(imagen_path))

    tablas = []

    for bloque in result:

        if bloque["type"] == "table":

            html = bloque["res"]["html"]

            try:

                dfs = pd.read_html(html)

                if len(dfs) > 0:

                    tablas.append(dfs[0])

            except:
                pass

    if len(tablas) == 0:
        raise Exception(f"No se detectó tabla en {imagen_path}")

    # Tomar tabla principal
    df = tablas[0]

    return df


# ============================================
# EXTRAER TABLAS
# ============================================

dfs = []

for imagen in IMAGES:

    df = extraer_tabla(imagen)

    df = renombrar_columnas(df)

    df = limpiar_dataframe(df)

    dfs.append(df)

# ============================================
# MERGE HORIZONTAL
# ============================================

df_final = dfs[0]

for df in dfs[1:]:

    columnas_nuevas = [

        c for c in df.columns

        if c not in df_final.columns
    ]

    columnas_merge = ["Name"] + columnas_nuevas

    df_final = df_final.merge(
        df[columnas_merge],
        on="Name",
        how="left"
    )

# ============================================
# ELIMINAR DUPLICADOS
# ============================================

df_final = df_final.loc[
    :,
    ~df_final.columns.duplicated()
]

# ============================================
# EXPORTAR CSV
# ============================================

df_final.to_csv(
    OUTPUT_CSV,
    index=False
)

# ============================================
# EXPORTAR XLSX
# ============================================

df_final.to_excel(
    OUTPUT_XLSX,
    index=False
)

# ============================================
# FORMATO NUMÉRICO EXCEL
# ============================================

wb = load_workbook(OUTPUT_XLSX)

ws = wb.active

for row in ws.iter_rows():

    for cell in row:

        if isinstance(cell.value, (int, float)):

            cell.number_format = '#,##0.00'

wb.save(OUTPUT_XLSX)

# ============================================
# FINAL
# ============================================

print("\n====================================")
print("ARCHIVOS GENERADOS")
print("====================================")

print(OUTPUT_CSV)
print(OUTPUT_XLSX)