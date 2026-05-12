import re
import math
import pandas as pd
import numpy as np
import yfinance as yf

from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# =========================================================
# CONFIGURACIÓN
# =========================================================

BASE_PATH = Path(
    "/mnt/c/Users/Alvaro/OneDrive/Documentos/Curso MIT McKinseay IFP/"
    "Inversiones Stocks/Analisis Morningstar Barbell - Quality Value Defensive 2026-05-11"
)

FILES = {
    "Vista 1": BASE_PATH / "Quality Value Defensive - Morningstar Barbell - Vista 1 - 2026-05-11.xlsx",
    "Vista 2": BASE_PATH / "Quality Value Defensive - Morningstar Barbell - Vista 2 - 2026-05-11.xlsx",
    "Vista 3": BASE_PATH / "Quality Value Defensive - Morningstar Barbell - Vista 3 - 2026-05-11.xlsx",
    "Vista 4": BASE_PATH / "Quality Value Defensive - Morningstar Barbell - Vista 4 - 2026-05-11.xlsx",
}

RESTRICTED_FILE = BASE_PATH / "Restricted_Final.xlsx"

OUTPUT_FILE = (
    BASE_PATH /
    "Quality Value Defensive - Morningstar Barbell 2026-05-11.xlsx"
)

# =========================================================
# FUNCIONES
# =========================================================

def financial_to_number(value):
    if pd.isna(value):
        return "NA"

    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return "NA"
        return value

    value = str(value).strip()

    if value in ["", "-", "NM", "NM-", "#RESTRICTED!"]:
        return "NA"

    multipliers = {
        "K": 1_000,
        "M": 1_000_000,
        "B": 1_000_000_000,
        "T": 1_000_000_000_000,
    }

    match = re.match(r"^([-+]?\d*\.?\d+)\s*([KMBT])$", value)

    if match:
        number = float(match.group(1))
        suffix = match.group(2)
        return number * multipliers[suffix]

    try:
        if "." in value:
            return float(value)
        return int(value)
    except:
        return value


def clean_dataframe(df):
    df = df.dropna(axis=1, how="all")

    df = df.replace(r"^\s*$", np.nan, regex=True)
    df = df.fillna("NA")
    df = df.replace("-", "NA")

    if "Full Ticker" not in df.columns:
        raise ValueError('❌ No existe columna "Full Ticker"')

    df["Full Ticker"] = (
        df["Full Ticker"]
        .astype(str)
        .str.upper()
        .str.strip()
    )

    if "Name" in df.columns:
        df["Name"] = df["Name"].astype(str).str.strip()

    if "Name" in df.columns:
        df = df[
            ~(
                (df["Name"] == "NA")
                &
                (df["Full Ticker"] == "NA")
            )
        ]

    invalid_tickers = df[
        (df["Full Ticker"].isna())
        |
        (df["Full Ticker"] == "NA")
        |
        (df["Full Ticker"] == "")
    ]

    if len(invalid_tickers) > 0:
        print(invalid_tickers)
        raise ValueError("Todas las empresas deben tener Full Ticker")

    for col in df.columns:
        df[col] = [financial_to_number(v) for v in df[col]]

    df.reset_index(drop=True, inplace=True)

    return df


def get_sector_industry(full_ticker):
    try:
        if full_ticker in ["NA", "", None]:
            return "NA", "NA"

        ticker = str(full_ticker).split(":")[-1].strip().upper()
        stock = yf.Ticker(ticker)
        info = stock.info

        sector = info.get("sector", "NA") or "NA"
        industry = info.get("industry", "NA") or "NA"

        return sector, industry

    except:
        return "NA", "NA"


def extract_ticker(full_ticker):
    if pd.isna(full_ticker):
        return "NA"

    full_ticker = str(full_ticker).strip().upper()

    if full_ticker in ["", "NA"]:
        return "NA"

    return full_ticker.split(":")[-1].strip().upper()


def reorder_identity_columns(df):
    first_cols = [
        "Name",
        "Full Ticker",
        "Ticker",
        "Sector",
        "Industry",
    ]

    existing_first = [
        c for c in first_cols
        if c in df.columns
    ]

    remaining_cols = [
        c for c in df.columns
        if c not in existing_first
    ]

    return df[existing_first + remaining_cols]


def prepare_restricted_dataframe(file_path, identity_map):
    if not file_path.exists():
        raise FileNotFoundError(
            f"❌ Archivo Restricted no encontrado:\n{file_path}"
        )

    df = pd.read_excel(file_path, engine="openpyxl")

    df = df.dropna(axis=1, how="all")
    df = df.replace(r"^\s*$", np.nan, regex=True)
    df = df.fillna("NA")
    df = df.replace("-", "NA")

    if "Full Ticker" not in df.columns:
        raise ValueError(
            '❌ Restricted_Final.xlsx no contiene columna "Full Ticker"'
        )

    df["Full Ticker"] = (
        df["Full Ticker"]
        .astype(str)
        .str.upper()
        .str.strip()
    )

    names = []
    tickers = []
    sectors = []
    industries = []

    for _, row in df.iterrows():
        full_ticker = str(row["Full Ticker"]).strip().upper()

        identity = identity_map.get(
            full_ticker,
            {
                "Name": "NA",
                "Ticker": extract_ticker(full_ticker),
                "Sector": "NA",
                "Industry": "NA",
            }
        )

        names.append(identity.get("Name", "NA"))
        tickers.append(identity.get("Ticker", extract_ticker(full_ticker)))
        sectors.append(identity.get("Sector", "NA"))
        industries.append(identity.get("Industry", "NA"))

    df["Name"] = names
    df["Ticker"] = tickers
    df["Sector"] = sectors
    df["Industry"] = industries

    df = reorder_identity_columns(df)

    return df

# =========================================================
# CARGA DE VISTAS 1-4
# =========================================================

dfs = {}

for sheet_name, file_path in FILES.items():
    print("\n================================================")
    print(f"Procesando: {sheet_name}")
    print("================================================")

    if not file_path.exists():
        raise FileNotFoundError(
            f"❌ Archivo no encontrado:\n{file_path}"
        )

    raw = pd.read_excel(
        file_path,
        header=None,
        engine="openpyxl"
    )

    raw = raw.iloc[7:, 2:]

    headers = raw.iloc[0]
    raw = raw[1:]

    raw.columns = headers.astype(str).str.strip()

    raw = raw.loc[
        :,
        ~raw.columns.str.contains(
            "^Unnamed",
            case=False,
            na=False
        )
    ]

    raw = raw.loc[:, ~raw.columns.duplicated()]
    raw.reset_index(drop=True, inplace=True)

    print("\nColumnas detectadas:")
    print(list(raw.columns))

    raw = clean_dataframe(raw)

    if sheet_name == "Vista 1":
        cols_remove = [
            "Fair Value (Analyst Target)",
            "Analyst Price Target Count",
        ]

        for c in cols_remove:
            if c in raw.columns:
                raw = raw.drop(columns=[c])

    dfs[sheet_name] = raw

# =========================================================
# ENRIQUECER VISTA 1
# =========================================================

print("\n================================================")
print("Investigando Sector e Industry...")
print("================================================")

vista1 = dfs["Vista 1"]

sectors = []
industries = []

for idx, row in vista1.iterrows():
    full_ticker = str(row.get("Full Ticker", "NA")).strip().upper()

    print(f"{idx + 1}/{len(vista1)} -> {full_ticker}")

    sector, industry = get_sector_industry(full_ticker)

    sectors.append(sector)
    industries.append(industry)

if "Ticker" not in vista1.columns:
    vista1.insert(
        2,
        "Ticker",
        vista1["Full Ticker"].apply(extract_ticker)
    )

if "Sector" in vista1.columns:
    vista1["Sector"] = sectors
else:
    vista1.insert(
        vista1.columns.get_loc("Ticker") + 1,
        "Sector",
        sectors
    )

if "Industry" in vista1.columns:
    vista1["Industry"] = industries
else:
    vista1.insert(
        vista1.columns.get_loc("Sector") + 1,
        "Industry",
        industries
    )

vista1 = reorder_identity_columns(vista1)

dfs["Vista 1"] = vista1

# =========================================================
# MAPA DE IDENTIDAD
# =========================================================

identity_map = {}

for _, row in vista1.iterrows():
    full_ticker = str(row.get("Full Ticker", "NA")).strip().upper()

    identity_map[full_ticker] = {
        "Name": row.get("Name", "NA"),
        "Ticker": row.get("Ticker", extract_ticker(full_ticker)),
        "Sector": row.get("Sector", "NA"),
        "Industry": row.get("Industry", "NA"),
    }

# =========================================================
# PROPAGAR A VISTAS 2-4
# =========================================================

for sheet_name in ["Vista 2", "Vista 3", "Vista 4"]:
    print(f"\nActualizando {sheet_name}")

    df = dfs[sheet_name]

    if "Ticker" not in df.columns:
        df.insert(
            2,
            "Ticker",
            df["Full Ticker"].apply(extract_ticker)
        )

    sectors = []
    industries = []

    for _, row in df.iterrows():
        full_ticker = str(row.get("Full Ticker", "NA")).strip().upper()

        identity = identity_map.get(
            full_ticker,
            {
                "Sector": "NA",
                "Industry": "NA",
            }
        )

        sectors.append(identity.get("Sector", "NA"))
        industries.append(identity.get("Industry", "NA"))

    if "Sector" in df.columns:
        df["Sector"] = sectors
    else:
        df.insert(
            df.columns.get_loc("Ticker") + 1,
            "Sector",
            sectors
        )

    if "Industry" in df.columns:
        df["Industry"] = industries
    else:
        df.insert(
            df.columns.get_loc("Sector") + 1,
            "Industry",
            industries
        )

    df = reorder_identity_columns(df)

    dfs[sheet_name] = df

# =========================================================
# INTEGRAR RESTRICTED_FINAL
# =========================================================

print("\n================================================")
print("Procesando Restricted_Final.xlsx")
print("================================================")

restricted_df = prepare_restricted_dataframe(
    RESTRICTED_FILE,
    identity_map
)

dfs["Restricted"] = restricted_df

# =========================================================
# CREAR EXCEL FINAL
# =========================================================

print("\n================================================")
print("Creando archivo maestro...")
print("================================================")

wb = Workbook()
default_sheet = wb.active
wb.remove(default_sheet)

for sheet_name, df in dfs.items():
    print(f"Escribiendo: {sheet_name}")

    ws = wb.create_sheet(title=sheet_name)

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            try:
                value = str(cell.value)

                if len(value) > max_length:
                    max_length = len(value)

                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"

            except:
                pass

        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

wb.save(OUTPUT_FILE)

print("\n================================================")
print("✅ ARCHIVO CREADO EXITOSAMENTE")
print("================================================")
print(f"\n📄 Archivo final:\n{OUTPUT_FILE}")

print("\n📌 Hojas:")
for sheet in wb.sheetnames:
    print(f" - {sheet}")

print("\n================================================")
print("FIN")
print("================================================")
