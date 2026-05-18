from pathlib import Path
import re
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURACIÓN PRINCIPAL
# ============================================================

BASE_DIR = Path(
    "/mnt/c/Users/Alvaro/OneDrive/Documentos/Curso MIT McKinseay IFP/"
    "Inversiones Stocks/Analisis Morningstar Barbell - Quality Value Defensive 2026-05-13/"
    "Arista ANE"
)

SOURCE_SHEETS_ORDER = [
    "Income Statement",
    "Balance Sheet",
    "Cash Flow Statement",
    "Financial Summary",
    "Growth",
    "Profitability and Efficiency",
    "Financial Health",
    "Ratio of Cash Flow",
    "Trailing Returns",
    "Valuation",
]

# IMPORTANTE:
# Archivos fuente .xls
SOURCE_FILES = {
    sheet_name: BASE_DIR / f"{sheet_name}.xls"
    for sheet_name in SOURCE_SHEETS_ORDER
}

# IMPORTANTE:
# Archivo final .xlsx
OUTPUT_FILE = BASE_DIR / "Arista ANET 2026-05-17.xlsx"

MISSING_LABEL = "NA_MISSING"
NOT_AVAILABLE_LABEL = "NA_NOT_AVAILABLE"

MIN_YEAR = 2000
MAX_YEAR = 2050

PERCENT_NUMBER_FORMAT = "0.00%"
NUMBER_FORMAT = "#,##0.00"
INTEGER_FORMAT = "#,##0"
YEAR_FORMAT = "0"


# ============================================================
# REGLAS DE RENOMBRAMIENTO DE VARIABLES
# ============================================================

GROWTH_PERCENT_VARIABLES = {
    "year over year": "Year Over Year %",
    "3-year average": "3-Year Average %",
    "5-year average": "5-Year Average %",
    "10-year average": "10-Year Average %",
    "year average": "Year Average %",
}

FINANCIAL_HEALTH_RENAMES = {
    "cap ex as a % of sales": "Cap Ex as a % of Sales %",
}

CASH_FLOW_RATIO_RENAMES = {
    "operating cash flow growth % yoy": "Operating Cash Flow Growth % YOY %",
    "free cash flow growth % yoy": "Free Cash Flow Growth % YOY %",
}


# ============================================================
# VARIABLES ABSOLUTAS QUE NO DEBEN SER PORCENTAJE
# ============================================================

ABSOLUTE_VALUE_VARIABLE_EXCEPTIONS = {
    "net cash flow from continuing operating activities, indirect",
    "cash from operations",
    "cash from operating activities",
    "operating cash flow",
    "levered free cash flow",
    "free cash flow",
    "capital expenditures",
    "net income",
    "net income to stockholders",
    "net income to company",
    "net income to common excl extra items",
    "revenue",
    "total revenue",
    "business revenue",
    "gross profit",
    "operating income",
    "operating income reported",
    "operating income adjusted",
    "ebit",
    "ebitda",
    "total assets",
    "total liabilities",
    "total debt",
    "net debt",
    "cash and equivalents",
    "cash and short term investments",
    "short term investments",
    "accounts receivable, net",
    "inventory",
    "total current assets",
    "total current liabilities",
    "property plant and equipment, net",
    "goodwill",
    "other intangibles",
    "common equity",
    "total equity",
    "repurchase of common stock",
    "dividends paid",
    "cash acquisitions",
    "cash from investing",
    "cash from financing",
    "beginning cash",
    "ending cash",
}


# ============================================================
# FUNCIONES DE TEXTO Y DETECCIÓN
# ============================================================

def normalize_text(value) -> str:
    """Normaliza textos: espacios, NBSP, saltos, etc."""

    if value is None:
        return ""

    text = str(value)
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)

    return text.strip()


def normalize_key(value) -> str:
    """Normaliza para comparar nombres de variables, columnas y hojas."""

    return normalize_text(value).lower()


def is_blank(value) -> bool:
    if value is None:
        return True

    if isinstance(value, str) and normalize_text(value) == "":
        return True

    return False


def is_na_like(value) -> bool:
    if not isinstance(value, str):
        return False

    text = normalize_key(value)

    return text in {
        "-",
        "—",
        "–",
        "na",
        "n/a",
        "n.a.",
        "nan",
        "null",
        "none",
    }


def is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def is_year_like(value) -> bool:
    """
    Detecta años:
    2026
    2026.00
    "2026"
    "2026.00"
    """

    if value is None:
        return False

    if isinstance(value, int):
        return MIN_YEAR <= value <= MAX_YEAR

    if isinstance(value, float):
        return value.is_integer() and MIN_YEAR <= int(value) <= MAX_YEAR

    if isinstance(value, str):
        text = normalize_text(value)
        match = re.fullmatch(r"(20\d{2})(\.0+)?", text)

        if match:
            year = int(match.group(1))
            return MIN_YEAR <= year <= MAX_YEAR

    return False


def to_year(value):
    if isinstance(value, (int, float)):
        return int(value)

    text = normalize_text(value)
    match = re.search(r"20\d{2}", text)

    if match:
        return int(match.group(0))

    return value


def is_growth_sheet(sheet_name) -> bool:
    return normalize_key(sheet_name) == "growth"


def is_financial_health_sheet(sheet_name) -> bool:
    return normalize_key(sheet_name) == "financial health"


def is_cash_flow_ratios_sheet(sheet_name) -> bool:
    """
    Detecta la hoja Ratios de Cash Flow.
    Tolera nombres parecidos.
    """

    key = normalize_key(sheet_name)

    return key in {
        "ratios de cash flow",
        "cash flow ratios",
        "ratios cash flow",
        "ratios of cash flow",
    }


def variable_name_ends_with_percent(variable_name) -> bool:
    return normalize_text(variable_name).endswith("%")


def is_absolute_value_exception(variable_name) -> bool:
    return normalize_key(variable_name) in ABSOLUTE_VALUE_VARIABLE_EXCEPTIONS


# ============================================================
# RENOMBRAMIENTO DE VARIABLES EN COLUMNA A
# ============================================================

def rename_variable_for_sheet(sheet_name, variable_name):
    """
    Renombra variables específicas en columna A según la hoja.
    Evita duplicar % si ya fue renombrada.
    """

    original = normalize_text(variable_name)
    key = normalize_key(original)

    if not original:
        return original

    key_without_final_percent = key[:-1].strip() if key.endswith("%") else key

    if is_growth_sheet(sheet_name):
        if key_without_final_percent in GROWTH_PERCENT_VARIABLES:
            return GROWTH_PERCENT_VARIABLES[key_without_final_percent]

    if is_financial_health_sheet(sheet_name):
        if key_without_final_percent in FINANCIAL_HEALTH_RENAMES:
            return FINANCIAL_HEALTH_RENAMES[key_without_final_percent]

    if is_cash_flow_ratios_sheet(sheet_name):
        if key_without_final_percent in CASH_FLOW_RATIO_RENAMES:
            return CASH_FLOW_RATIO_RENAMES[key_without_final_percent]

    return original


# ============================================================
# PARSEO NUMÉRICO
# ============================================================

def parse_numeric(value):
    """
    Convierte textos financieros a número.

    Casos soportados:
    "1,234.56"       -> 1234.56
    "$1,234.56"      -> 1234.56
    "(1,234.56)"     -> -1234.56
    "25%"            -> 25, status percent_text
    25               -> 25
    25.87            -> 25.87
    """

    if value is None:
        return None, "blank"

    if isinstance(value, (int, float)):
        return value, "numeric"

    if not isinstance(value, str):
        return value, "other"

    raw = normalize_text(value)

    if raw == "":
        return None, "blank"

    if is_na_like(raw):
        return None, "na_like"

    if is_formula(raw):
        return raw, "formula"

    negative = False

    if raw.startswith("(") and raw.endswith(")"):
        negative = True
        raw = raw[1:-1].strip()

    has_percent_symbol = "%" in raw

    cleaned = (
        raw.replace("%", "")
           .replace("$", "")
           .replace("USD", "")
           .replace("usd", "")
           .replace(",", "")
           .replace("\xa0", "")
           .strip()
    )

    cleaned = re.sub(r"\s+", "", cleaned)

    try:
        number = float(cleaned)

        if negative:
            number = -number

        if has_percent_symbol:
            return number, "percent_text"

        return number, "numeric_text"

    except ValueError:
        return value, "text"


def percent_to_decimal(number):
    """
    Convierte porcentaje a decimal interno:
    25.87 -> 0.2587
    0.2587 -> 0.2587
    -12.5 -> -0.125
    """

    if not isinstance(number, (int, float)):
        return number

    if abs(number) <= 1:
        return number

    return number / 100


def should_treat_as_percent(variable_name, parsed_status) -> bool:
    """
    Reglas:
    1. Si la variable está en excepciones absolutas, NO es porcentaje.
    2. Si la variable termina en %, SÍ es porcentaje.
    3. Si el valor trae explícitamente símbolo %, SÍ es porcentaje,
       salvo excepción absoluta.
    """

    if is_absolute_value_exception(variable_name):
        return False

    if variable_name_ends_with_percent(variable_name):
        return True

    if parsed_status == "percent_text":
        return True

    return False


# ============================================================
# FORMATO
# ============================================================

def apply_standard_number_format(cell, value):
    if isinstance(value, int):
        cell.number_format = INTEGER_FORMAT

    elif isinstance(value, float):
        if value.is_integer():
            cell.number_format = INTEGER_FORMAT
        else:
            cell.number_format = NUMBER_FORMAT


def clean_header_cell(cell):
    """
    Limpia encabezados y corrige años.
    """

    original = cell.value

    if is_year_like(original):
        cell.value = to_year(original)
        cell.number_format = YEAR_FORMAT

    elif isinstance(original, str):
        cell.value = normalize_text(original)

    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="D9EAF7")
    cell.alignment = Alignment(horizontal="center", vertical="center")


# ============================================================
# CONSOLIDACIÓN INICIAL DE ARCHIVOS .XLS
# ============================================================

def copy_source_sheet_to_target(source_file, target_wb, target_sheet_name):
    """
    Lee un archivo .xls con pandas/xlrd y lo copia como hoja
    dentro del workbook final .xlsx.
    """

    if not source_file.exists():
        raise FileNotFoundError(
            f"No se encontró el archivo fuente: {source_file}"
        )

    df = pd.read_excel(
        source_file,
        header=None,
        engine="xlrd"
    )

    target_ws = target_wb.create_sheet(
        title=target_sheet_name
    )

    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row, start=1):

            if pd.isna(value):
                value = None

            target_ws.cell(
                row=row_idx + 1,
                column=col_idx,
                value=value
            )

    return target_ws


def build_consolidated_workbook():
    """
    Construye un workbook en memoria consolidando los 10 archivos .xls
    en el orden definido por SOURCE_SHEETS_ORDER.
    """

    wb = Workbook()

    default_sheet = wb.active
    wb.remove(default_sheet)

    print("\n============================================================")
    print("CONSOLIDANDO ARCHIVOS FUENTE .XLS")
    print("============================================================")

    for sheet_name in SOURCE_SHEETS_ORDER:
        source_file = SOURCE_FILES[sheet_name]

        print(f"Agregando hoja: {sheet_name}")
        print(f"Archivo fuente: {source_file}")

        copy_source_sheet_to_target(
            source_file=source_file,
            target_wb=wb,
            target_sheet_name=sheet_name
        )

    return wb


# ============================================================
# LIMPIEZA DE HOJA
# ============================================================

def clean_sheet(ws):
    max_row = ws.max_row
    max_col = ws.max_column

    if max_row == 0 or max_col == 0:
        return

    # --------------------------------------------------------
    # 1. Limpiar encabezados de columnas
    # --------------------------------------------------------

    for col in range(1, max_col + 1):
        clean_header_cell(ws.cell(row=1, column=col))

    # --------------------------------------------------------
    # 2. Renombrar variables específicas en columna A ANTES
    #    de procesar valores numéricos
    # --------------------------------------------------------

    for row in range(1, max_row + 1):
        variable_cell = ws.cell(row=row, column=1)
        original_variable_name = variable_cell.value

        if isinstance(original_variable_name, str):
            renamed_variable = rename_variable_for_sheet(
                ws.title,
                original_variable_name
            )

            variable_cell.value = renamed_variable

    # --------------------------------------------------------
    # 3. Limpiar filas y valores
    # --------------------------------------------------------

    for row in range(1, max_row + 1):
        variable_cell = ws.cell(row=row, column=1)
        variable_name = variable_cell.value

        # Limpiar nombre de variable en columna A
        if isinstance(variable_name, str):
            variable_name = normalize_text(variable_name)
            variable_cell.value = variable_name

        # Corregir años en cualquier celda
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            original_value = cell.value

            if is_year_like(original_value):
                cell.value = to_year(original_value)
                cell.number_format = YEAR_FORMAT

        # Procesar valores desde columna 2
        for col in range(2, max_col + 1):
            cell = ws.cell(row=row, column=col)
            original_value = cell.value

            # No tocar fórmulas
            if is_formula(original_value):
                continue

            # Vacíos
            if is_blank(original_value):
                cell.value = MISSING_LABEL
                continue

            # NA / guiones
            if is_na_like(original_value):
                cell.value = NOT_AVAILABLE_LABEL
                continue

            # Años en cuerpo
            if is_year_like(original_value):
                cell.value = to_year(original_value)
                cell.number_format = YEAR_FORMAT
                continue

            parsed_value, status = parse_numeric(original_value)

            # Números
            if (
                status in {"numeric", "numeric_text", "percent_text"}
                and isinstance(parsed_value, (int, float))
            ):

                if should_treat_as_percent(variable_name, status):
                    new_value = percent_to_decimal(parsed_value)
                    cell.value = new_value
                    cell.number_format = PERCENT_NUMBER_FORMAT

                else:
                    cell.value = parsed_value
                    apply_standard_number_format(cell, parsed_value)

            # Texto normal
            elif status == "text":
                cell.value = normalize_text(parsed_value)

    # --------------------------------------------------------
    # 4. Formato visual básico
    # --------------------------------------------------------

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    ws.column_dimensions["A"].width = 60

    for col in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


# ============================================================
# PROCESO PRINCIPAL
# ============================================================

def main():
    wb = build_consolidated_workbook()

    # Eliminar hojas de logs si existieran de ejecuciones anteriores
    for sheet_name in [
        "Data_Cleaning_Log",
        "Missing_Data_Log",
        "Structure_Log",
    ]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    print("\n============================================================")
    print("LIMPIANDO Y ESTANDARIZANDO HOJAS")
    print("============================================================")

    for ws in wb.worksheets:
        print(f"Limpiando hoja: {ws.title}")
        clean_sheet(ws)

    wb.save(OUTPUT_FILE)

    print("\nLimpieza completada correctamente.")
    print(f"Archivo limpio: {OUTPUT_FILE}")
    print("No se generaron hojas de logs.")


if __name__ == "__main__":
    main()