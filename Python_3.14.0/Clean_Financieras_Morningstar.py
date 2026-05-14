from pathlib import Path
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURACIÓN PRINCIPAL
# ============================================================

INPUT_FILE = Path(
    "/mnt/c/Users/Alvaro/OneDrive/Documentos/Curso MIT McKinseay IFP/Inversiones Stocks/Analisis Morningstar Barbell - Quality Value Defensive 2026-05-13/Salesforce CRM 2026-05-14.xlsx"
)

OUTPUT_FILE = INPUT_FILE.with_name(INPUT_FILE.stem + " CLEAN STANDARDIZED.xlsx")

MISSING_LABEL = "NA_MISSING"
NOT_AVAILABLE_LABEL = "NA_NOT_AVAILABLE"

MIN_YEAR = 2000
MAX_YEAR = 2050

PERCENT_NUMBER_FORMAT = "0.00%"
NUMBER_FORMAT = "#,##0.00"
INTEGER_FORMAT = "#,##0"
YEAR_FORMAT = "0"

SKIP_SHEETS = {
    "Data_Cleaning_Log",
    "Missing_Data_Log",
    "Structure_Log",
}


# ============================================================
# REGLAS DE INTERPRETACIÓN
# ============================================================

# Variables de la hoja Growth que deben renombrarse agregando %
GROWTH_PERCENT_VARIABLES = {
    "year over year": "Year Over Year %",
    "3-year average": "3-Year Average %",
    "5-year average": "5-Year Average %",
    "10-year average": "10-Year Average %",
    "year average": "Year Average %",
}

# Variables absolutas que NO deben tratarse como porcentaje aunque Excel las muestre así.
ABSOLUTE_VALUE_VARIABLE_EXCEPTIONS = {
    "net cash flow from continuing operating activities, indirect",
    "cash from operations",
    "cash from operating activities",
    "operating cash flow",
    "levered free cash flow",
    "free cash flow",
    "capital expenditures",
    "net income",
    "net income to stockholders",
    "net income to company",
    "net income to common excl extra items",
    "revenue",
    "total revenue",
    "business revenue",
    "gross profit",
    "operating income",
    "operating income reported",
    "operating income adjusted",
    "ebit",
    "ebitda",
    "total assets",
    "total liabilities",
    "total debt",
    "net debt",
    "cash and equivalents",
    "cash and short term investments",
    "short term investments",
    "accounts receivable, net",
    "inventory",
    "total current assets",
    "total current liabilities",
    "property plant and equipment, net",
    "goodwill",
    "other intangibles",
    "common equity",
    "total equity",
    "repurchase of common stock",
    "dividends paid",
    "cash acquisitions",
    "cash from investing",
    "cash from financing",
    "beginning cash",
    "ending cash",
}


# ============================================================
# FUNCIONES DE TEXTO Y DETECCIÓN
# ============================================================

def normalize_text(value) -> str:
    """Normaliza textos: espacios, NBSP, saltos, etc."""
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def normalize_key(value) -> str:
    """Normaliza para comparar nombres de variables/columnas."""
    return normalize_text(value).lower()


def is_blank(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and normalize_text(value) == "":
        return True
    return False


def is_na_like(value) -> bool:
    if not isinstance(value, str):
        return False

    text = normalize_key(value)
    return text in {
        "-",
        "—",
        "–",
        "na",
        "n/a",
        "n.a.",
        "nan",
        "null",
        "none",
    }


def is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def is_year_like(value) -> bool:
    """
    Detecta años:
    2026
    2026.00
    "2026"
    "2026.00"
    """
    if value is None:
        return False

    if isinstance(value, int):
        return MIN_YEAR <= value <= MAX_YEAR

    if isinstance(value, float):
        return value.is_integer() and MIN_YEAR <= int(value) <= MAX_YEAR

    if isinstance(value, str):
        text = normalize_text(value)
        match = re.fullmatch(r"(20\d{2})(\.0+)?", text)
        if match:
            year = int(match.group(1))
            return MIN_YEAR <= year <= MAX_YEAR

    return False


def to_year(value):
    if isinstance(value, (int, float)):
        return int(value)

    text = normalize_text(value)
    match = re.search(r"20\d{2}", text)
    if match:
        return int(match.group(0))

    return value


def is_growth_sheet(sheet_name) -> bool:
    return normalize_key(sheet_name) == "growth"


def growth_variable_with_percent(value):
    """
    Si la variable de la hoja Growth es:
    Year Over Year
    3-Year Average
    5-Year Average
    10-Year Average
    Year Average

    devuelve el nombre con %.
    Si ya tiene %, no lo duplica.
    """
    text = normalize_text(value)

    if not text:
        return text

    key = normalize_key(text)

    # Si ya termina en %, comparar quitando %
    if key.endswith("%"):
        key_without_percent = key[:-1].strip()
        if key_without_percent in GROWTH_PERCENT_VARIABLES:
            return GROWTH_PERCENT_VARIABLES[key_without_percent]
        return text

    if key in GROWTH_PERCENT_VARIABLES:
        return GROWTH_PERCENT_VARIABLES[key]

    return text


def variable_name_ends_with_percent(variable_name) -> bool:
    return normalize_text(variable_name).endswith("%")


def is_absolute_value_exception(variable_name) -> bool:
    return normalize_key(variable_name) in ABSOLUTE_VALUE_VARIABLE_EXCEPTIONS


# ============================================================
# PARSEO NUMÉRICO
# ============================================================

def parse_numeric(value):
    """
    Convierte textos financieros a número.

    Casos soportados:
    "1,234.56"       -> 1234.56
    "$1,234.56"      -> 1234.56
    "(1,234.56)"     -> -1234.56
    "25%"            -> 25, status percent_text
    25               -> 25
    25.87            -> 25.87
    """

    if value is None:
        return None, "blank"

    if isinstance(value, (int, float)):
        return value, "numeric"

    if not isinstance(value, str):
        return value, "other"

    raw = normalize_text(value)

    if raw == "":
        return None, "blank"

    if is_na_like(raw):
        return None, "na_like"

    if is_formula(raw):
        return raw, "formula"

    negative = False

    if raw.startswith("(") and raw.endswith(")"):
        negative = True
        raw = raw[1:-1].strip()

    has_percent_symbol = "%" in raw

    cleaned = (
        raw.replace("%", "")
           .replace("$", "")
           .replace("USD", "")
           .replace("usd", "")
           .replace(",", "")
           .replace("\xa0", "")
           .strip()
    )

    cleaned = re.sub(r"\s+", "", cleaned)

    try:
        number = float(cleaned)
        if negative:
            number = -number

        if has_percent_symbol:
            return number, "percent_text"

        return number, "numeric_text"

    except ValueError:
        return value, "text"


def percent_to_decimal(number):
    """
    Convierte porcentaje a decimal interno:
    25.87 -> 0.2587
    0.2587 -> 0.2587
    -12.5 -> -0.125
    """
    if not isinstance(number, (int, float)):
        return number

    if abs(number) <= 1:
        return number

    return number / 100


def should_treat_as_percent(sheet_name, variable_name, parsed_status) -> bool:
    """
    Regla corregida:

    1. Si la variable está en excepciones absolutas, NO es porcentaje.
    2. Si la variable termina en %, SÍ es porcentaje.
       Esto cubre Growth porque primero renombramos:
       Year Over Year -> Year Over Year %
    3. Si el valor trae explícitamente símbolo %, SÍ es porcentaje,
       salvo excepción absoluta.
    4. NO usamos el formato visual de Excel como criterio.
    """

    if is_absolute_value_exception(variable_name):
        return False

    if variable_name_ends_with_percent(variable_name):
        return True

    if parsed_status == "percent_text":
        return True

    return False


# ============================================================
# FORMATO
# ============================================================

def apply_standard_number_format(cell, value):
    if isinstance(value, int):
        cell.number_format = INTEGER_FORMAT
    elif isinstance(value, float):
        if value.is_integer():
            cell.number_format = INTEGER_FORMAT
        else:
            cell.number_format = NUMBER_FORMAT


def clean_header_cell(cell, cleaning_log, sheet_name):
    original = cell.value

    if is_year_like(original):
        new_value = to_year(original)
        cell.value = new_value
        cell.number_format = YEAR_FORMAT
        cleaning_log.append([
            sheet_name,
            cell.coordinate,
            "header_year_standardized",
            original,
            new_value,
        ])

    elif isinstance(original, str):
        cleaned = normalize_text(original)
        if cleaned != original:
            cell.value = cleaned
            cleaning_log.append([
                sheet_name,
                cell.coordinate,
                "header_text_cleaned",
                original,
                cleaned,
            ])

    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="D9EAF7")
    cell.alignment = Alignment(horizontal="center", vertical="center")


# ============================================================
# LIMPIEZA DE HOJA
# ============================================================

def clean_sheet(ws, cleaning_log, missing_log, structure_log):
    max_row = ws.max_row
    max_col = ws.max_column

    if max_row == 0 or max_col == 0:
        structure_log.append([ws.title, max_row, max_col, "empty_or_invalid"])
        return

    # --------------------------------------------------------
    # 1. Limpiar encabezados de columnas
    # --------------------------------------------------------
    for col in range(1, max_col + 1):
        clean_header_cell(ws.cell(row=1, column=col), cleaning_log, ws.title)

    # --------------------------------------------------------
    # 2. PRIMERO renombrar variables de la hoja Growth en columna A
    #    antes de convertir formatos numéricos.
    # --------------------------------------------------------
    if is_growth_sheet(ws.title):
        for row in range(1, max_row + 1):
            variable_cell = ws.cell(row=row, column=1)
            original_value = variable_cell.value

            if isinstance(original_value, str):
                cleaned_value = normalize_text(original_value)
                renamed_value = growth_variable_with_percent(cleaned_value)

                if renamed_value != original_value:
                    variable_cell.value = renamed_value
                    cleaning_log.append([
                        ws.title,
                        variable_cell.coordinate,
                        "growth_variable_percent_suffix_added",
                        original_value,
                        renamed_value,
                    ])

    # --------------------------------------------------------
    # 3. Limpiar filas y valores
    # --------------------------------------------------------
    for row in range(1, max_row + 1):
        variable_cell = ws.cell(row=row, column=1)
        original_variable_name = variable_cell.value

        # Limpiar nombre de variable en primera columna
        if isinstance(original_variable_name, str):
            cleaned_variable_name = normalize_text(original_variable_name)
            if cleaned_variable_name != original_variable_name:
                variable_cell.value = cleaned_variable_name
                cleaning_log.append([
                    ws.title,
                    variable_cell.coordinate,
                    "variable_name_cleaned",
                    original_variable_name,
                    cleaned_variable_name,
                ])
        else:
            cleaned_variable_name = original_variable_name

        # Corregir años en cualquier celda del cuerpo
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            original_value = cell.value

            if is_year_like(original_value):
                new_value = to_year(original_value)
                if original_value != new_value or cell.number_format != YEAR_FORMAT:
                    cell.value = new_value
                    cell.number_format = YEAR_FORMAT
                    cleaning_log.append([
                        ws.title,
                        cell.coordinate,
                        "year_cell_standardized",
                        original_value,
                        new_value,
                    ])

        # Procesar valores de datos desde columna 2
        for col in range(2, max_col + 1):
            cell = ws.cell(row=row, column=col)
            original_value = cell.value

            # No tocar fórmulas
            if is_formula(original_value):
                continue

            # Vacíos
            if is_blank(original_value):
                cell.value = MISSING_LABEL
                missing_log.append([
                    ws.title,
                    cell.coordinate,
                    normalize_text(cleaned_variable_name),
                    "blank",
                    MISSING_LABEL,
                ])
                continue

            # NA / guiones
            if is_na_like(original_value):
                cell.value = NOT_AVAILABLE_LABEL
                missing_log.append([
                    ws.title,
                    cell.coordinate,
                    normalize_text(cleaned_variable_name),
                    "not_available_or_not_applicable",
                    NOT_AVAILABLE_LABEL,
                ])
                continue

            # Años en cuerpo
            if is_year_like(original_value):
                new_value = to_year(original_value)
                cell.value = new_value
                cell.number_format = YEAR_FORMAT
                cleaning_log.append([
                    ws.title,
                    cell.coordinate,
                    "year_body_standardized",
                    original_value,
                    new_value,
                ])
                continue

            parsed_value, status = parse_numeric(original_value)

            if status in {"numeric", "numeric_text", "percent_text"} and isinstance(parsed_value, (int, float)):

                percent_context = should_treat_as_percent(
                    sheet_name=ws.title,
                    variable_name=cleaned_variable_name,
                    parsed_status=status,
                )

                if percent_context:
                    new_value = percent_to_decimal(parsed_value)
                    cell.value = new_value
                    cell.number_format = PERCENT_NUMBER_FORMAT

                    cleaning_log.append([
                        ws.title,
                        cell.coordinate,
                        "percent_standardized_decimal",
                        original_value,
                        new_value,
                    ])

                else:
                    cell.value = parsed_value
                    apply_standard_number_format(cell, parsed_value)

                    if status == "percent_text":
                        action = "percent_text_kept_absolute_due_exception_or_context"
                    elif status == "numeric_text":
                        action = "numeric_text_to_number"
                    else:
                        action = "numeric_standardized"

                    cleaning_log.append([
                        ws.title,
                        cell.coordinate,
                        action,
                        original_value,
                        parsed_value,
                    ])

            elif status == "text":
                cleaned_text = normalize_text(parsed_value)
                if cleaned_text != original_value:
                    cell.value = cleaned_text
                    cleaning_log.append([
                        ws.title,
                        cell.coordinate,
                        "text_cleaned",
                        original_value,
                        cleaned_text,
                    ])

    # --------------------------------------------------------
    # 4. Formato visual básico
    # --------------------------------------------------------
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions

    ws.column_dimensions["A"].width = 55

    for col in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    structure_log.append([
        ws.title,
        max_row,
        max_col,
        "processed",
    ])


# ============================================================
# HOJAS DE LOG
# ============================================================

def create_log_sheet(wb, sheet_name, headers, rows, header_fill):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    ws.append(headers)

    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 34

    return ws


# ============================================================
# PROCESO PRINCIPAL
# ============================================================

def main():
    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"No se encontró el archivo: {INPUT_FILE}")

    wb = load_workbook(INPUT_FILE)

    # Eliminar logs anteriores si existen
    for sheet_name in list(SKIP_SHEETS):
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    cleaning_log = []
    missing_log = []
    structure_log = []

    for ws in wb.worksheets:
        if ws.title in SKIP_SHEETS:
            continue

        clean_sheet(ws, cleaning_log, missing_log, structure_log)

    create_log_sheet(
        wb,
        "Data_Cleaning_Log",
        ["Sheet", "Cell", "Action", "Original Value", "New Value"],
        cleaning_log,
        "1F4E78",
    )

    create_log_sheet(
        wb,
        "Missing_Data_Log",
        ["Sheet", "Cell", "Variable", "Missing Type", "Assigned Label"],
        missing_log,
        "9C0006",
    )

    create_log_sheet(
        wb,
        "Structure_Log",
        ["Sheet", "Rows", "Columns", "Status"],
        structure_log,
        "548235",
    )

    wb.save(OUTPUT_FILE)

    print("Limpieza completada correctamente.")
    print(f"Archivo original: {INPUT_FILE}")
    print(f"Archivo limpio:   {OUTPUT_FILE}")
    print(f"Cambios registrados: {len(cleaning_log)}")
    print(f"Datos faltantes / no disponibles registrados: {len(missing_log)}")
    print(f"Hojas procesadas: {len(structure_log)}")


if __name__ == "__main__":
    main()