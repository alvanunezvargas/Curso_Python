import pandas as pd
from sklearn import preprocessing
from sklearn.model_selection import train_test_split
from sklearn import svm
from sklearn.metrics import confusion_matrix
import openpyxl
from joblib import dump

# Cargar los datos desde un archivo Excel
df = pd.read_excel('ruta_al_archivo_excel')

# Asumiendo que la última columna es la columna de respuesta
Predictores = df.iloc[:, :-1]
Respuesta = df.iloc[:, -1]

# Normalizar los predictores utilizando la estandarización
X = preprocessing.StandardScaler().fit(Predictores).transform(Predictores)

# Dividir los datos en conjuntos de entrenamiento y prueba
X_train, X_test, y_train, y_test = train_test_split(X, Respuesta, test_size=0.2, random_state=4)

# Crear un clasificador SVM
clf = svm.SVC(gamma=0.001, C=100.)

# Entrenar el clasificador SVM en los datos de entrenamiento
clf.fit(X_train, y_train)

# Hacer predicciones en los datos de prueba
yhat = clf.predict(X_test)

# Imprimir la matriz de confusión
print(confusion_matrix(y_test, yhat, labels=[1, 0]))

# Guardar las predicciones en un archivo Excel
df_pred = pd.DataFrame(yhat, columns=['Predicciones'])
df_pred.to_excel('ruta_a_guardar_predicciones.xlsx', index=False)

# Guardar el modelo entrenado
dump(clf, 'ruta_a_guardar_modelo.joblib') 


"""
Ejercicio realizado con el conjunto de datos de Iris para clasificar las flores en tres clases. Este conjunto de datos fue recopilado por el estadístico británico Ronald Fisher en 1936

Este codigo imprime un tipo de gráfico llamado "Kernel Density Estimate" (KDE). Un KDE es una forma de estimar la función de densidad de probabilidad de una variable aleatoria.

En términos más simples, un KDE suaviza un histograma. En lugar de tener barras discretas como en un histograma, un KDE tiene una curva suave que da una idea de la "forma" de la distribución de datos.

Imprime la presicion obtenida con los modelos Análisis Discriminante Lineal (LinearDiscriminantAnalysis) y 
Análisis Discriminante Cuadrático (QuadraticDiscriminantAnalysis)

El Análisis Discriminante Lineal busca encontrar las mejores combinaciones lineales de estas características para separar los grupos de la manera más clara posible. En otras palabras, intenta dibujar líneas (o planos, si hay más de dos características) que dividan los grupos de forma óptima.

El Análisis Discriminante Cuadrático (QuadraticDiscriminantAnalysis) es similar al Análisis Discriminante Lineal, pero en lugar de líneas rectas, utiliza curvas (o superficies) para separar los grupos. Esto le da más flexibilidad para manejar situaciones en las que los grupos no están separados linealmente.
"""

import warnings
warnings.filterwarnings('ignore', category=FutureWarning)
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis, QuadraticDiscriminantAnalysis
from sklearn.metrics import accuracy_score
from sklearn import preprocessing
import seaborn as sns
import matplotlib.pyplot as plt
import matplotlib.pyplot as plt
import numpy as np

# Cargar los datos desde el archivo iris.csv
data = pd.read_csv('datasets/iris.csv')

# Dividir los datos en características (X) y etiquetas (y)
X = data[['sepal length', 'sepal width', 'petal length', 'petal width']]
y = data['class']

# Normalizar los predictores utilizando la estandarización. Esto reescala los predictores para que tengan media 0 y varianza 1.
# X = preprocessing.StandardScaler().fit(X).transform(X)

# Dividir los datos en conjuntos de entrenamiento y prueba
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# Inicializar y entrenar el modelo LDA
lda = LinearDiscriminantAnalysis()
lda.fit(X_train, y_train)

# Predecir las etiquetas en todo el conjunto de datos utilizando LDA
y_pred_lda = lda.predict(X)

# Calcular la precisión de LDA
accuracy_lda = accuracy_score(y, y_pred_lda)
print("Precisión de LDA:", accuracy_lda)

# Inicializar y entrenar el modelo QDA
qda = QuadraticDiscriminantAnalysis()
qda.fit(X_train, y_train)

# Predecir las etiquetas en todo el conjunto de datos utilizando QDA
y_pred_qda = qda.predict(X)

# Calcular la precisión de QDA
accuracy_qda = accuracy_score(y, y_pred_qda)
print("Precisión de QDA:", accuracy_qda)

# Crear un DataFrame con las predicciones
df_predictions = pd.DataFrame({
    'LDA Predictions': y_pred_lda,
    'QDA Predictions': y_pred_qda
})

# Guardar el DataFrame en un archivo de Excel
# df_predictions.to_excel('predictions.xlsx', index=False)

sns.pairplot(data, hue="class", height=3, diag_kind="kde")
plt.show()

# Otra forma:

import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.discriminant_analysis import LinearDiscriminantAnalysis, QuadraticDiscriminantAnalysis
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.naive_bayes import MultinomialNB, GaussianNB
from sklearn.ensemble import AdaBoostClassifier, GradientBoostingClassifier
from sklearn.svm import SVC, LinearSVC
from sklearn.neural_network import MLPClassifier
from sklearn.linear_model import LogisticRegression, SGDClassifier, Perceptron, RidgeClassifier
from sklearn.metrics import accuracy_score
import seaborn as sns
import matplotlib.pyplot as plt

# Cargar los datos desde el archivo iris.csv
data = pd.read_csv('datasets/iris.csv')

# Dividir los datos en características (X) y etiquetas (y)
X = data[['sepal length', 'sepal width', 'petal length', 'petal width']]
y = data['class']

# Dividir los datos en conjuntos de entrenamiento y prueba
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# Crear un pairplot de los datos, coloreando por la clase de la flor
sns.pairplot(data, hue="class", size=3, diag_kind="kde")
plt.show()

# Inicializar y entrenar los modelos
models = [
    ('LDA', LinearDiscriminantAnalysis()),
    ('QDA', QuadraticDiscriminantAnalysis()),
    ('Decision Tree', DecisionTreeClassifier()),
    ('Nearest Neighbors', KNeighborsClassifier()),
    ('Random Forest', RandomForestClassifier()),
    ('Naive Bayes', MultinomialNB()),
    ('AdaBoost', AdaBoostClassifier()),
    ('SVM', SVC(kernel='linear')),
    ('Neural Network', MLPClassifier(hidden_layer_sizes=(100, 100), max_iter=1000)),
    ('Gradient Boosting', GradientBoostingClassifier()),
    ('Logistic Regression', LogisticRegression(max_iter=1000)),
    ('Gaussian Naive Bayes', GaussianNB()),
    ('Linear SVM', LinearSVC(max_iter=10000, dual=False)),
    ('SGD', SGDClassifier(max_iter=1000)),
    ('Perceptron', Perceptron(max_iter=1000)),
    ('MLP', MLPClassifier(max_iter=1000)),
    ('Ridge', RidgeClassifier(max_iter=1000))    
]

# Predecir las etiquetas en todo el conjunto de datos y calcular la precisión para cada modelo
accuracies = {}
predictions = {}
for name, model in models:
    model.fit(X_train, y_train)
    accuracies[name] = model.score(X_test, y_test)
    predictions[name] = model.predict(X)

# Graficar la precisión de todos los modelos
plt.bar(accuracies.keys(), accuracies.values())
plt.ylabel('Precisión')
plt.show()
plt.close()

# Crear un DataFrame con las predicciones
df_predictions = pd.DataFrame(predictions)

# Guardar el DataFrame en un archivo de Excel
df_predictions.to_excel('predictions.xlsx', index=False)

# Realizar corrida de clasificacion en los datos de prueba desde el archivo 'prueba.xlsx'
test_data = pd.read_excel('prueba.xlsx')

# Asegúrate de que las columnas en test_data coincidan con las que usaste para entrenar los modelos
# Supongamos que las columnas son 'sepal length', 'sepal width', 'petal length', 'petal width'
X_test = test_data[['sepal length', 'sepal width', 'petal length', 'petal width']]

# Hacer predicciones en los datos de prueba con cada modelo
test_predictions = {name: model.predict(X_test) for name, model in models}

# Crear un DataFrame con las predicciones de prueba
df_test_predictions = pd.DataFrame(test_predictions)

# Guardar el DataFrame en un archivo de Excel
df_test_predictions.to_excel('test_predictions.xlsx', index=False)


# Evaluacion modelos para inferir Numero de cetano del Diesel de su espctro NIR

import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.utils import shuffle
from sklearn import metrics
from sklearn.ensemble import RandomForestRegressor
from sklearn.neighbors import KNeighborsRegressor
from sklearn.linear_model import Ridge
from sklearn.linear_model import Lasso
from sklearn.linear_model import LinearRegression
from sklearn.model_selection import GridSearchCV
from sklearn.model_selection import cross_val_score
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.feature_selection import SelectKBest
from sklearn.feature_selection import f_regression
from numpy import set_printoptions
import matplotlib.pyplot as plt


# Leer los datos del archivo Excel, omitiendo la primera fila
df = pd.read_excel('datasets/cnDieselTrain.xlsx', sheet_name='TrainX-Y')

# Dividir los datos en X e Y
cnTrainX = df.iloc[0:134, 1:402] 
cnTrainY = df.iloc[0:134, 402]

# Ver las primeras 5 filas de X
# print("Primeras 5 filas de X:")
# print(cnTrainX.head())

# Ver las últimas 5 filas de X
# print("\nÚltimas 5 filas de X:")
# print(cnTrainX.tail())

# Ver las primeras 5 filas de Y
# print("\nPrimeras 5 filas de Y:")
# print(cnTrainY.head())

# Ver las últimas 5 filas de Y
# print("\nÚltimas 5 filas de Y:")
# print(cnTrainY.tail())


feature_selection=SelectKBest(score_func=f_regression,k=5)
feature_fit=feature_selection
best_features=feature_fit.fit(cnTrainX,cnTrainY)
set_printoptions(precision=3)
features=feature_fit.transform(cnTrainX)


#Showing the results in DataFrame

dfscores=pd.DataFrame(best_features.scores_)
dfcolums=pd.DataFrame(cnTrainX.columns)
featureScores=pd.concat([dfcolums,dfscores],axis=1)
featureScores.columns=['Feature','Score']


#Plotting features vs score
test_f=featureScores.nlargest(50,'Score')
test_f.plot(x='Feature',y='Score',kind='bar',figsize=(15,15))
# plt.show()

# Transformar cnTrainX para solo contener las mejores características
cnTrainX_new = feature_fit.transform(cnTrainX)

# Convertir el resultado en un DataFrame
cnTrainX_new = pd.DataFrame(cnTrainX_new, columns=cnTrainX.columns[feature_fit.get_support()])

X_train,X_test,y_train,y_test=train_test_split(cnTrainX_new,cnTrainY,test_size=0.3,train_size=0.7,shuffle=True)

sc = StandardScaler()
X_train = sc.fit_transform(X_train)
X_test = sc.transform(X_test)

test=[(KNeighborsRegressor(n_neighbors=10)),(LinearRegression()),(RandomForestRegressor(n_estimators = 500, random_state = 0)),(Ridge(alpha=0.01)),(Lasso())]
for model in test: 
    cv_error=cross_val_score(model,X_train,y_train,cv=5,scoring='neg_mean_squared_error')
    print('Mean and SD of cross validation',np.mean(cv_error),np.std(cv_error))

print('')
regressor=KNeighborsRegressor(n_neighbors=10)
regressor.fit(X_train,y_train)
y_pred=regressor.predict(X_test)
print('\nMean Absolute Error KNeighborsRegressor:', metrics.mean_absolute_error(y_test, y_pred))
print('Mean Squared Error KNeighborsRegressor:', metrics.mean_squared_error(y_test, y_pred))
print('Root Mean Squared Error KNeighborsRegressor:', np.sqrt(metrics.mean_squared_error(y_test, y_pred)))


regressor=LinearRegression()
regressor.fit(X_train,y_train)
y_pred=regressor.predict(X_test)
print('\nMean Absolute Error LinearRegression:', metrics.mean_absolute_error(y_test, y_pred))
print('Mean Squared Error LinearRegression:', metrics.mean_squared_error(y_test, y_pred))
print('Root Mean Squared Error LinearRegression:', np.sqrt(metrics.mean_squared_error(y_test, y_pred)))


regressor=RandomForestRegressor(n_estimators = 500, random_state = 0)
regressor.fit(X_train,y_train)
y_pred=regressor.predict(X_test)
print('\nMean Absolute Error RandomForestRegressor:', metrics.mean_absolute_error(y_test, y_pred))
print('Mean Squared Error RandomForestRegressor:', metrics.mean_squared_error(y_test, y_pred))
print('Root Mean Squared Error RandomForestRegressor:', np.sqrt(metrics.mean_squared_error(y_test, y_pred)))


regressor=Lasso()
regressor.fit(X_train,y_train)
y_pred=regressor.predict(X_test)
print('\nMean Absolute Error Lasso:', metrics.mean_absolute_error(y_test, y_pred))
print('Mean Squared Error Lasso:', metrics.mean_squared_error(y_test, y_pred))
print('Root Mean Squared Error Lasso:', np.sqrt(metrics.mean_squared_error(y_test, y_pred)))


regressor=Ridge()
regressor.fit(X_train,y_train)
y_pred=regressor.predict(X_test)
print('\nMean Absolute Error Ridge:', metrics.mean_absolute_error(y_test, y_pred))
print('Mean Squared Error Ridge:', metrics.mean_squared_error(y_test, y_pred))
print('Root Mean Squared Ridge:', np.sqrt(metrics.mean_squared_error(y_test, y_pred)))
print('')

parameters={'alpha':[1e-15,1e-12,1e-5,0.1,5,10,20]}
rf_random = GridSearchCV(regressor,parameters,cv=5)
# Fit the random search model
rf_random.fit(X_train, y_train)
print(rf_random.best_params_)

regressor=Lasso(alpha=0.1)
regressor.fit(X_train,y_train)
y_pred=regressor.predict(X_test)
print('\nMean Absolute Error Lasso(alpha=0.1):', metrics.mean_absolute_error(y_test, y_pred))
print('Mean Squared Error Lasso(alpha=0.1):', metrics.mean_squared_error(y_test, y_pred))
print('Root Mean Squared Error Lasso(alpha=0.1):', np.sqrt(metrics.mean_squared_error(y_test, y_pred)))


# Otra forma de evaluar modelos

import pandas as pd
import numpy as np
from numpy import set_printoptions
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LinearRegression, Ridge, Lasso, ElasticNet, SGDRegressor
from sklearn.neighbors import KNeighborsRegressor
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor
from sklearn.svm import SVR
from sklearn.neural_network import MLPRegressor
from sklearn.ensemble import ExtraTreesRegressor
from sklearn.ensemble import AdaBoostRegressor, BaggingRegressor, ExtraTreesRegressor, GradientBoostingRegressor, HistGradientBoostingRegressor
from xgboost import XGBRegressor
from sklearn.model_selection import train_test_split  
from sklearn.feature_selection import SelectKBest
from sklearn.feature_selection import f_regression
from sklearn.metrics import mean_absolute_error, mean_squared_error
from sklearn.metrics import r2_score
import matplotlib.pyplot as plt
from pandas import ExcelWriter

# Cargar los datos desde el archivo iris.csv
df = pd.read_excel('datasets/cnDieselTrain.xlsx', sheet_name='TrainX-Y')

# Dividir los datos en características (X) y etiquetas (y)
cnTrainX = df.iloc[0:134, 1:402] 
cnTrainY = df.iloc[0:134, 402]

# realizar selección de características en un conjunto de datos
feature_selection=SelectKBest(score_func=f_regression,k=5)
feature_fit=feature_selection
best_features=feature_fit.fit(cnTrainX,cnTrainY)
set_printoptions(precision=3)
features=feature_fit.transform(cnTrainX)

#Showing the results in DataFrame

dfscores=pd.DataFrame(best_features.scores_)
dfcolums=pd.DataFrame(cnTrainX.columns)
featureScores=pd.concat([dfcolums,dfscores],axis=1)
featureScores.columns=['Feature','Score']

#Plotting features vs score
test_f=featureScores.nlargest(50,'Score')
test_f.plot(x='Feature',y='Score',kind='bar',figsize=(15,15))
# plt.show()

# Transformar cnTrainX para solo contener las mejores características
cnTrainX_new = feature_fit.transform(cnTrainX)

# Convertir el resultado en un DataFrame
cnTrainX_new = pd.DataFrame(cnTrainX_new, columns=cnTrainX.columns[feature_fit.get_support()])

# Dividir los datos en conjuntos de entrenamiento y prueba
X_train,X_test,y_train,y_test=train_test_split(cnTrainX_new,cnTrainY,test_size=0.3,train_size=0.7,shuffle=True)

sc = StandardScaler()
X_train = sc.fit_transform(X_train)
X_test = sc.transform(X_test)

# Inicializar y entrenar los modelos

models = [
    ('Linear Regression', LinearRegression()),
    ('Ridge Regression', Ridge()),
    ('Lasso Regression', Lasso()),
    ('Elastic Net', ElasticNet(max_iter=1000)),  
    ('K-Nearest Neighbors', KNeighborsRegressor()),
    ('Decision Tree', DecisionTreeRegressor()),
    ('Random Forest', RandomForestRegressor()),
    ('Support Vector Regression', SVR()),
    ('MLP (Multi-Layer Perceptron)', MLPRegressor(max_iter=10000)),
    ('Gradient Boosting', ExtraTreesRegressor()),  
    ('SGD (Stochastic Gradient Descent)', SGDRegressor()),
    ('AdaBoost', AdaBoostRegressor()),
    ('Bagging', BaggingRegressor()),
    ('Extra Trees', ExtraTreesRegressor()),
    ('Gradient Boosting', GradientBoostingRegressor()),
    ('Histogram Gradient Boosting', HistGradientBoostingRegressor()),
    ('XGBoost', XGBRegressor()),
    ]

# Inicializar diccionarios para almacenar los errores
mae_errors = {}
mse_errors = {}
rmse_errors = {}
r2_scores = {}

# Crear un DataFrame para almacenar las predicciones
predictions = pd.DataFrame()

# Crear un DataFrame para almacenar los errores
errors = pd.DataFrame()

for name, model in models:
    if model is not None:
        model.fit(X_train, y_train)  # Entrenar el modelo
        y_pred = model.predict(X_test)
        mae_errors[name] = mean_absolute_error(y_test, y_pred)
        mse_errors[name] = mean_squared_error(y_test, y_pred)
        rmse_errors[name] = np.sqrt(mean_squared_error(y_test, y_pred))
        r2_scores[name] = r2_score(y_test, y_pred)
        predictions[name] = y_pred

# Agregar los errores a su respectivo DataFrame
errors['Model'] = mae_errors.keys()
errors['MAE'] = mae_errors.values()
errors['MSE'] = mse_errors.values()
errors['RMSE'] = rmse_errors.values()
errors['R2 Score'] = r2_scores.values()

# Crear un escritor de Excel
writer = ExcelWriter('datasets\\NCPredictions.xlsx')

# Crear un escritor de Excel
with pd.ExcelWriter('datasets\\NCPredictions.xlsx') as writer:
    # Guardar los DataFrames en diferentes hojas
    predictions.to_excel(writer, sheet_name='Predictions', index=False)
    errors.to_excel(writer, sheet_name='Errors', index=False)
    
"""
# Numero de cetano en Diesel: "Mean CV Error": Es el error medio de validación cruzada. En este caso, se utiliza el error cuadrático medio negativo (`neg_mean_squared_error`). Es la media de los errores cuadráticos de cada pliegue en la validación cruzada. Los errores cuadráticos se calculan como la diferencia entre las predicciones del modelo y los valores reales al cuadrado. El valor negativo se utiliza porque `cross_val_score` considera que un valor más alto es mejor, lo que no es el caso para el error cuadrático medio.
# "SD CV Error": Es la desviación estándar del error de validación cruzada. Proporciona una medida de cuánto varía el error de validación cruzada entre los diferentes pliegues en la validación cruzada. Una desviación estándar más alta indica que el rendimiento del modelo varía más entre los diferentes pliegues.
# Estas métricas proporcionan una visión general del rendimiento del modelo. La media del error de validación cruzada da una idea de cuánto error se puede esperar en promedio, mientras que la desviación estándar del error de validación cruzada da una idea de cuánto puede variar este error.

Aplican algoritmos, se generan los modelos de aprendizaje y se imprimen en un archivo excel.

Los modelos se aplican a un set de espectros para determinar el numero de cetano.
"""

import pandas as pd
import numpy as np
from sklearn.feature_selection import SelectKBest, f_regression
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LinearRegression, Ridge, Lasso, ElasticNet, SGDRegressor, BayesianRidge
from sklearn.neighbors import KNeighborsRegressor
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor, ExtraTreesRegressor, AdaBoostRegressor, BaggingRegressor, GradientBoostingRegressor, HistGradientBoostingRegressor
from sklearn.svm import SVR
from sklearn.neural_network import MLPRegressor
from xgboost import XGBRegressor
from sklearn.model_selection import train_test_split, cross_val_score
import matplotlib.pyplot as plt

# Cargar los datos
df = pd.read_excel('datasets/cnDieselTrain.xlsx', sheet_name='TrainX-Y')

# Dividir los datos en características (X) y etiquetas (y)
cnTrainX = df.iloc[0:134, 1:402] 
cnTrainY = df.iloc[0:134, 402]

# Realizar selección de características
feature_selection = SelectKBest(score_func=f_regression, k=50)
features = feature_selection.fit_transform(cnTrainX, cnTrainY)

# Crear DataFrame con las puntuaciones de las características
feature_scores = pd.DataFrame({'Feature': cnTrainX.columns, 'Score': feature_selection.scores_})

# Ordenar el DataFrame por la columna 'Score' en orden descendente y seleccionar las 50 características principales
top_features = feature_scores.nlargest(50, 'Score')

# Seleccionar solo las características más relevantes de cnTrainX
cnTrainX_new = cnTrainX[top_features['Feature']]

# Trazar las puntuaciones de las características
top_features.plot(x='Feature', y='Score', kind='bar', figsize=(15, 15))
# plt.show()

# Escalar los datos
sc = StandardScaler()
cnTrainX_new = sc.fit_transform(cnTrainX_new)

# Convertir de nuevo a DataFrame y conservar los nombres de las características
cnTrainX_new = pd.DataFrame(cnTrainX_new, columns=top_features['Feature'])

# Dividir los datos en conjuntos de entrenamiento y prueba
X_train, X_test, y_train, y_test = train_test_split(cnTrainX_new, cnTrainY, test_size=0.3, train_size=0.7, shuffle=True)
# X_train, X_test, y_train, y_test = cnTrainX_new, pd.DataFrame(), cnTrainY, pd.DataFrame()

# Inicializar y entrenar los modelos
models = [
    ('Linear Regression', LinearRegression()),
    ('Ridge Regression', Ridge()),
    ('Lasso Regression', Lasso()),
    ('Elastic Net', ElasticNet(max_iter=1000)),  
    ('K-Nearest Neighbors', KNeighborsRegressor()),
    ('Decision Tree', DecisionTreeRegressor()),
    ('Random Forest', RandomForestRegressor()),
    ('Support Vector Regression', SVR()),
    ('MLP (Multi-Layer Perceptron)', MLPRegressor(max_iter=10000)),
    ('Gradient Boosting', ExtraTreesRegressor()),  
    ('SGD (Stochastic Gradient Descent)', SGDRegressor()),
    ('AdaBoost', AdaBoostRegressor()),
    ('Bagging', BaggingRegressor()),
    ('Extra Trees', ExtraTreesRegressor()),
    ('Gradient Boosting', GradientBoostingRegressor()),
    ('Histogram Gradient Boosting', HistGradientBoostingRegressor()),
    ('XGBoost', XGBRegressor()),
    ('Bayesian Ridge Regression', BayesianRidge())
    ]

# Cargar los datos de prueba
df_test = pd.read_excel('datasets/cnDieselTrain.xlsx', sheet_name='TestX')

# Seleccionar solo las características más relevantes de df_test
df_test_new = df_test[top_features['Feature']]

# Escalar los datos de prueba
sc = StandardScaler()
df_test_new = sc.fit_transform(df_test_new)

# Convertir de nuevo a DataFrame y conservar los nombres de las características
df_test_new = pd.DataFrame(df_test_new, columns=top_features['Feature'])

errors = []
predictions_train = pd.DataFrame()
predictions_test = pd.DataFrame()

for name, model in models:
    # Cross-validation error
    cv_error = cross_val_score(model, X_train, y_train, cv=5, scoring='neg_mean_squared_error')
    errors.append((name, np.mean(cv_error), np.std(cv_error)))
    
    # Fit model and make predictions
    model.fit(cnTrainX_new, cnTrainY)
    y_pred_train = model.predict(cnTrainX_new)
    y_pred_test = model.predict(df_test_new)
    predictions_train[name] = y_pred_train
    predictions_test[name] = y_pred_test

# Crear un DataFrame a partir de la lista de errores
errors = pd.DataFrame(errors, columns=['Model', 'Mean CV Error', 'SD CV Error'])

# Escribir el DataFrame en un archivo de Excel
errors.to_excel('datasets/cnErrorDiesel.xlsx', index=False)
predictions_train.to_excel('datasets/TrainPrediccionNC.xlsx', index=False)
predictions_test.to_excel('datasets/TestPrediccionNC.xlsx', index=False)


"""
Este codigo permite hacer la evaluacion del error y la prediccion de varias propiedades (etiquetas) de un conjunto de datos de Diesel. Se seleccionan las 50 caracteristicas mas relevantes para hacer la prediccion de las etiquetas. Se evaluan varios modelos de regresion y se guardan los errores y las predicciones en archivos de Excel.
"""

import pandas as pd
import numpy as np
from numpy import set_printoptions
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LinearRegression, Ridge, Lasso, ElasticNet, SGDRegressor, BayesianRidge
from sklearn.neighbors import KNeighborsRegressor
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor
from sklearn.svm import SVR
from sklearn.neural_network import MLPRegressor
from sklearn.ensemble import ExtraTreesRegressor
from sklearn.ensemble import AdaBoostRegressor, BaggingRegressor, ExtraTreesRegressor, GradientBoostingRegressor, HistGradientBoostingRegressor
from xgboost import XGBRegressor
from sklearn.model_selection import train_test_split, cross_val_score  
from sklearn.feature_selection import SelectKBest
from sklearn.feature_selection import f_regression
from sklearn.metrics import mean_absolute_error, mean_squared_error
from sklearn.metrics import r2_score
import openpyxl
import os

# Cargar los datos
df = pd.read_excel('datasets/Modelo Diesel.xlsx', sheet_name='SNV')

# Crear el archivo 'datasets/DieselSeleccion.xlsx' si no existe
file_path = 'datasets/DieselSeleccion.xlsx'
if not os.path.exists(file_path):
    wb = openpyxl.Workbook()
    wb.save(file_path)

# Bucle a través de las columnas que quieres usar como etiquetas
for i, column in enumerate(range(2354, 2335, -1), start=1): # Range(2354, 2335, -1) en excel es (2353(CLO), 2335(CKW)..
    
    # Dividir los datos en características (X) y etiquetas (y)
    # Por defecto si hay identificacion en el encabezado de columna, la indexacion 0 de python empieza
    # en la segunda fila de excel, por tanto 
    # df.iloc[0:361, 2:2336] en excel son las columnas [Filas 2:358, Columnas 3:2335]
    cnTrainX = df.iloc[0:361, 2:2336]
    cnTrainY = df.iloc[0:361, column]
    
    # Eliminar las filas con valores NaN en cnTrainY
    cnTrainX = cnTrainX[~cnTrainY.isna()]
    cnTrainY = cnTrainY.dropna()
    
    # Realizar selección de características
    feature_selection = SelectKBest(score_func=f_regression, k=50)
    features = feature_selection.fit_transform(cnTrainX, cnTrainY)

    # Crear DataFrame con las puntuaciones de las características
    feature_scores = pd.DataFrame({'Feature': cnTrainX.columns, 'Score': feature_selection.scores_})

    # Ordenar el DataFrame por la columna 'Score' en orden descendente y seleccionar las 50 características principales
    top_features = feature_scores.nlargest(50, 'Score')

    # Añadir las 50 características principales a una nueva hoja en el archivo Excel
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        top_features.to_excel(writer, sheet_name=df.columns[column], index=False)

# Eliminar la hoja temporal después de agregar la primera hoja con los datos
wb = openpyxl.load_workbook(file_path)
del wb['Sheet']
wb.save(file_path)

# Crear el archivo 'datasets/DieselErrorPrediccion.xlsx' si no existe
file_path = 'datasets/DieselErrorPrediccion.xlsx'
if not os.path.exists(file_path):
    wb = openpyxl.Workbook()
    wb.save(file_path)
    
# Crear el archivo 'datasets/DieselPrediccion.xlsx' si no existe
file_path_predictions = 'datasets/DieselPrediccion.xlsx'
if not os.path.exists(file_path_predictions):
    wb_predictions = openpyxl.Workbook()
    wb_predictions.save(file_path_predictions)

for i, column in enumerate(range(2354, 2335, -1), start=1): # Range(2354, 2335, -1) en excel es (2353(CLO), 2335(CKW)..
    
    # Dividir los datos en características (X) y etiquetas (y)
    # Por defecto si hay identificacion en el encabezado de columna, la indexacion 0 de python empieza
    # en la segunda fila de excel, por tanto 
    # df.iloc[0:361, 2:2336] en excel son las columnas [Filas 2:358, Columnas 3:2335]
    cnTrainX = df.iloc[0:361, 2:2336]
    cnTrainY = df.iloc[0:361, column]
    
    # Eliminar las filas con valores NaN en cnTrainY
    cnTrainX = cnTrainX[~cnTrainY.isna()]
    cnTrainY = cnTrainY.dropna()
    
    # Realizar selección de características
    feature_selection = SelectKBest(score_func=f_regression, k=50)
    features = feature_selection.fit_transform(cnTrainX, cnTrainY)

    # Crear DataFrame con las puntuaciones de las características
    feature_scores = pd.DataFrame({'Feature': cnTrainX.columns, 'Score': feature_selection.scores_})

    # Ordenar el DataFrame por la columna 'Score' en orden descendente y seleccionar las 50 características principales
    top_features = feature_scores.nlargest(50, 'Score')
    
    # Seleccionar solo las características más relevantes de cnTrainX
    cnTrainX_new = cnTrainX[top_features['Feature']]

    # Dividir los datos en conjuntos de entrenamiento y prueba
    X_train, X_test, y_train, y_test = train_test_split(cnTrainX_new, cnTrainY, test_size=0.3, train_size=0.7, shuffle=True)

    # Escalar las características
    sc = StandardScaler()
    X_train = sc.fit_transform(X_train)
    X_test = sc.transform(X_test)
    
    # Inicializar y entrenar los modelos
    models = [
    ('Linear Regression', LinearRegression()),
    ('Ridge Regression', Ridge()),
    ('Lasso Regression', Lasso()),
    ('Elastic Net', ElasticNet(max_iter=1000)),  
    ('K-Nearest Neighbors', KNeighborsRegressor()),
    ('Decision Tree', DecisionTreeRegressor()),
    ('Random Forest', RandomForestRegressor()),
    ('Support Vector Regression', SVR()),
    ('MLP (Multi-Layer Perceptron)', MLPRegressor(max_iter=10000)),
    ('Gradient Boosting', ExtraTreesRegressor()),  
    ('SGD (Stochastic Gradient Descent)', SGDRegressor()),
    ('AdaBoost', AdaBoostRegressor()),
    ('Bagging', BaggingRegressor()),
    ('Extra Trees', ExtraTreesRegressor()),
    ('Gradient Boosting', GradientBoostingRegressor()),
    ('Histogram Gradient Boosting', HistGradientBoostingRegressor()),
    ('XGBoost', XGBRegressor()),
    ('Bayesian Ridge Regression', BayesianRidge())
    ]
    
    errors = []
    predictions_train = pd.DataFrame()
    predictions_test = pd.DataFrame()

    for name, model in models:
    # Cross-validation error
      cv_error = cross_val_score(model, X_train, y_train, cv=5, scoring='neg_mean_squared_error')
      errors.append((name, np.mean(cv_error), np.std(cv_error)))
      model.fit(cnTrainX_new, cnTrainY)
      y_pred_train = model.predict(cnTrainX_new)
      predictions_train[name] = y_pred_train

    # Convertir la lista de errores en un DataFrame
    errors_df = pd.DataFrame(errors, columns=['Model', 'Mean Error', 'Standard Deviation'])

    # Añadir los errores a una hoja identificada con el nombre de las etiquetas de la columna
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        errors_df.to_excel(writer, sheet_name=df.columns[column], index=False)
    
    # Añadir las predicciones a una hoja en el nuevo archivo de Excel
    with pd.ExcelWriter(file_path_predictions, engine='openpyxl', mode='a') as writer:
        predictions_train.to_excel(writer, sheet_name=df.columns[column], index=False)
    
# Eliminar la hoja temporal después de agregar la primera hoja con los datos
wb = openpyxl.load_workbook(file_path)
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']
wb.save(file_path)

# Hacer lo mismo para el nuevo archivo de Excel
wb_predictions = openpyxl.load_workbook(file_path_predictions)
if 'Sheet' in wb_predictions.sheetnames:
    del wb_predictions['Sheet']
wb_predictions.save(file_path_predictions)

# Este codigo hace exactamente lo mismo que el anterior, se aplica a varias etiquetas de un conjunto de datos de Diesel. Se seleccionan las 50 caracteristicas mas relevantes para hacer la prediccion de las etiquetas. Se evaluan varios modelos de regresion y se guardan los errores y las predicciones en archivos de Excel.

import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LinearRegression, Ridge, Lasso, ElasticNet, SGDRegressor, BayesianRidge
from sklearn.neighbors import KNeighborsRegressor
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor
from sklearn.svm import SVR
from sklearn.neural_network import MLPRegressor
from sklearn.ensemble import ExtraTreesRegressor
from sklearn.ensemble import AdaBoostRegressor, BaggingRegressor, ExtraTreesRegressor, GradientBoostingRegressor, HistGradientBoostingRegressor
from xgboost import XGBRegressor
from sklearn.model_selection import train_test_split, cross_val_score  
from sklearn.feature_selection import SelectKBest
from sklearn.feature_selection import f_regression
import openpyxl
import os

# Cargar los datos
df = pd.read_excel('datasets/Modelo Diesel.xlsx', sheet_name='SNV')

# Crear el archivo 'datasets/DieselSeleccion.xlsx' si no existe
file_path = 'datasets/DieselSeleccion.xlsx'
if not os.path.exists(file_path):
    wb = openpyxl.Workbook()
    wb.save(file_path)

# Crear el archivo 'datasets/DieselErrorPrediccion.xlsx' si no existe
file_path_errors = 'datasets/DieselErrorPrediccion.xlsx'
if not os.path.exists(file_path_errors):
    wb_errors = openpyxl.Workbook()
    wb_errors.save(file_path_errors)

# Crear el archivo 'datasets/DieselPrediccion.xlsx' si no existe
file_path_predictions = 'datasets/DieselPrediccion.xlsx'
if not os.path.exists(file_path_predictions):
    wb_predictions = openpyxl.Workbook()
    wb_predictions.save(file_path_predictions)

# Inicializar los modelos
models = [
    ('Linear Regression', LinearRegression()),
    ('Ridge Regression', Ridge()),
    ('Lasso Regression', Lasso()),
    ('Elastic Net', ElasticNet(max_iter=1000)),  
    ('K-Nearest Neighbors', KNeighborsRegressor()),
    ('Decision Tree', DecisionTreeRegressor()),
    ('Random Forest', RandomForestRegressor()),
    ('Support Vector Regression', SVR()),
    ('MLP (Multi-Layer Perceptron)', MLPRegressor(max_iter=10000)),
    ('Gradient Boosting', ExtraTreesRegressor()),  
    ('SGD (Stochastic Gradient Descent)', SGDRegressor()),
    ('AdaBoost', AdaBoostRegressor()),
    ('Bagging', BaggingRegressor()),
    ('Extra Trees', ExtraTreesRegressor()),
    ('Gradient Boosting', GradientBoostingRegressor()),
    ('Histogram Gradient Boosting', HistGradientBoostingRegressor()),
    ('XGBoost', XGBRegressor()),
    ('Bayesian Ridge Regression', BayesianRidge())
]

# Bucle a través de las columnas que quieres usar como etiquetas
for i, column in enumerate(range(2354, 2335, -1), start=1): # Range(2354, 2335, -1) en excel es (2353(CLO), 2335(CKW)..
    
    # Dividir los datos en características (X) y etiquetas (y)
    # Por defecto si hay identificacion en el encabezado de columna, la indexacion 0 de python empieza
    # en la segunda fila de excel, por tanto 
    # df.iloc[0:361, 2:2336] en excel son las columnas [Filas 2:358, Columnas 3:2335]
    cnTrainX = df.iloc[0:361, 2:2336]
    cnTrainY = df.iloc[0:361, column]
    
    # Eliminar las filas con valores NaN en cnTrainY
    cnTrainX = cnTrainX[~cnTrainY.isna()]
    cnTrainY = cnTrainY.dropna()
    
    # Realizar selección de características
    feature_selection = SelectKBest(score_func=f_regression, k=50)
    features = feature_selection.fit_transform(cnTrainX, cnTrainY)

    # Crear DataFrame con las puntuaciones de las características
    feature_scores = pd.DataFrame({'Feature': cnTrainX.columns, 'Score': feature_selection.scores_})

    # Ordenar el DataFrame por la columna 'Score' en orden descendente y seleccionar las 50 características principales
    top_features = feature_scores.nlargest(50, 'Score')

    # Añadir las 50 características principales a una nueva hoja en el archivo Excel
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        top_features.to_excel(writer, sheet_name=df.columns[column], index=False)

    # Seleccionar solo las características más relevantes de cnTrainX
    cnTrainX_new = cnTrainX[top_features['Feature']]

    # Dividir los datos en conjuntos de entrenamiento y prueba
    X_train, X_test, y_train, y_test = train_test_split(cnTrainX_new, cnTrainY, test_size=0.3, train_size=0.7, shuffle=True)

    # Escalar las características
    sc = StandardScaler()
    X_train = sc.fit_transform(X_train)
    X_test = sc.transform(X_test)
    
    errors = []
    predictions_train = pd.DataFrame()
    predictions_test = pd.DataFrame()

    for name, model in models:
        # Cross-validation error
        cv_error = cross_val_score(model, X_train, y_train, cv=5, scoring='neg_mean_squared_error')
        errors.append((name, np.mean(cv_error), np.std(cv_error)))
        model.fit(cnTrainX_new, cnTrainY)
        y_pred_train = model.predict(cnTrainX_new)
        predictions_train[name] = y_pred_train

    # Convertir la lista de errores en un DataFrame
    errors_df = pd.DataFrame(errors, columns=['Model', 'Mean Error', 'Standard Deviation'])

    # Añadir los errores a una hoja identificada con el nombre de las etiquetas de la columna
    with pd.ExcelWriter(file_path_errors, engine='openpyxl', mode='a') as writer:
        errors_df.to_excel(writer, sheet_name=df.columns[column], index=False)
    
    # Añadir las predicciones a una hoja en el nuevo archivo de Excel
    with pd.ExcelWriter(file_path_predictions, engine='openpyxl', mode='a') as writer:
        predictions_train.to_excel(writer, sheet_name=df.columns[column], index=False)

# Eliminar la hoja temporal después de agregar la primera hoja con los datos
wb = openpyxl.load_workbook(file_path)
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']
wb.save(file_path)

# Hacer lo mismo para el nuevo archivo de Excel
wb_errors = openpyxl.load_workbook(file_path_errors)
if 'Sheet' in wb_errors.sheetnames:
    del wb_errors['Sheet']
wb_errors.save(file_path_errors)

wb_predictions = openpyxl.load_workbook(file_path_predictions)
if 'Sheet' in wb_predictions.sheetnames:
    del wb_predictions['Sheet']
wb_predictions.save(file_path_predictions)


"""
18/05/2024 13:39

Tiempo de corrida: 1 min 12 seg 49 centecimas con los modelos LinearRegression(), Ridge() y Lasso()

Corrida que genera los archivos 'datasets/DieselSeleccion2.xlsx', 'datasets/DieselErrorPrediccion2.xlsx' y 'datasets/DieselPrediccion2.xlsx'

Este codigo imprime los archivos 'datasets/DieselSeleccion.xlsx', 'datasets/DieselErrorPrediccion.xlsx' y 'datasets/DieselPrediccion.xlsx'. 'datasets/DieselSeleccion.xlsx' contiene las caracteristicas mas relevantes para hacer la prediccion de las etiquetas. Se evaluan varios modelos de regresion y se guardan los errores en 'datasets/DieselErrorPrediccion.xlsx' y las predicciones en 'datasets/DieselPrediccion.xlsx'. 
"""

import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LinearRegression, Ridge, Lasso, ElasticNet, SGDRegressor, BayesianRidge
from sklearn.neighbors import KNeighborsRegressor
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor
from sklearn.svm import SVR
from sklearn.neural_network import MLPRegressor
from sklearn.ensemble import ExtraTreesRegressor
from sklearn.ensemble import AdaBoostRegressor, BaggingRegressor, ExtraTreesRegressor, GradientBoostingRegressor, HistGradientBoostingRegressor
from xgboost import XGBRegressor
from sklearn.model_selection import train_test_split, cross_val_score  
from sklearn.feature_selection import SelectKBest
from sklearn.feature_selection import f_regression
from sklearn.metrics import mean_squared_error
import openpyxl
import os

# Cargar los datos
df = pd.read_excel('datasets/Modelo Diesel.xlsx', sheet_name='SNV')

# Crear el archivo 'datasets/DieselSeleccion.xlsx' si no existe
file_path = 'datasets/DieselSeleccion.xlsx'
if not os.path.exists(file_path):
    wb = openpyxl.Workbook()
    wb.save(file_path)

# Crear el archivo 'datasets/DieselPrediccion.xlsx' si no existe
file_path_predictions = 'datasets/DieselPrediccion.xlsx'
if not os.path.exists(file_path_predictions):
    wb_predictions = openpyxl.Workbook()
    wb_predictions.save(file_path_predictions)

# Crear el archivo 'datasets/DieselErrorPrediccion.xlsx' si no existe
file_path_errors = 'datasets/DieselErrorPrediccion.xlsx'
if not os.path.exists(file_path_errors):
    wb_errors = openpyxl.Workbook()
    wb_errors.save(file_path_errors)

# Inicializar los modelos
models = [LinearRegression(), Ridge(), Lasso(), ElasticNet(max_iter=1000), KNeighborsRegressor(),
          DecisionTreeRegressor(), RandomForestRegressor(), SVR(), MLPRegressor(max_iter=10000), 
          ExtraTreesRegressor(), SGDRegressor(), AdaBoostRegressor(), BaggingRegressor(), 
          GradientBoostingRegressor(), HistGradientBoostingRegressor(),
          XGBRegressor(), BayesianRidge()]
         
# Bucle a través de las columnas que quieres usar como etiquetas
for i, column in enumerate(range(2354, 2335, -1), start=1): # Range(2354, 2335, -1) en excel es (2353(CLO), 2335(CKW)..

    # Dividir los datos en características (X) y etiquetas (y)
    # Por defecto si hay identificacion en el encabezado de columna, la indexacion 0 de python empieza
    # en la segunda fila de excel, por tanto 
    # df.iloc[0:361, 2:2336] en excel son las columnas [Filas 2:358, Columnas 3:2335]
    cnTrainX = df.iloc[0:361, 2:2336]
    cnTrainY = df.iloc[0:361, column]
    
    # Eliminar las filas con valores NaN en cnTrainY
    cnTrainX = cnTrainX[~cnTrainY.isna()]
    cnTrainY = cnTrainY.dropna()
    
    # Realizar selección de características
    feature_selection = SelectKBest(score_func=f_regression, k=50)
    features = feature_selection.fit_transform(cnTrainX, cnTrainY)

    # Crear DataFrame con las puntuaciones de las características
    feature_scores = pd.DataFrame({'Feature': cnTrainX.columns, 'Score': feature_selection.scores_})

    # Ordenar el DataFrame por la columna 'Score' en orden descendente y seleccionar las características principales
    top_features = feature_scores.nlargest(50, 'Score')

    # Añadir las características principales a una nueva hoja en el archivo Excel
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        top_features.to_excel(writer, sheet_name=df.columns[column], index=False)

    # Seleccionar solo las características más relevantes de cnTrainX
    cnTrainX_new = cnTrainX[top_features['Feature']]

    # Trazar las puntuaciones de las características
    top_features.plot(x='Feature', y='Score', kind='bar', figsize=(15, 15))
    # plt.show()

    # Dividir los datos en conjuntos de entrenamiento y prueba
    X_train, X_test, y_train, y_test = train_test_split(cnTrainX_new, cnTrainY, test_size=0.3, train_size=0.7, shuffle=True)

    # Escalar las características
    sc = StandardScaler()
    X_train = sc.fit_transform(X_train)
    X_test = sc.transform(X_test)
    
    # Crear DataFrame vacío para almacenar los errores y los scores
    error_data = pd.DataFrame()
    all_predictions = []

    for model in models:
    # Entrenar el modelo
      model.fit(X_train, y_train)

      # Realizar predicciones
      y_pred_train = model.predict(X_train)
      y_pred_test = model.predict(X_test)

      # Convertir las predicciones en DataFrames
      y_pred_train_df = pd.DataFrame(y_pred_train, columns=['y_pred_train'])
      y_pred_test_df = pd.DataFrame(y_pred_test, columns=['y_pred_test'])

      # Resetear los índices de y_train y y_pred_train_df para asegurar la alineación correcta
      y_train_reset = y_train.reset_index(drop=True)
      y_pred_train_df_reset = y_pred_train_df.reset_index(drop=True)
      
      # Resetear los índices de y_test y y_pred_test_df para asegurar la alineación correcta
      y_test_reset = y_test.reset_index(drop=True)
      y_pred_test_df_reset = y_pred_test_df.reset_index(drop=True)

      # Calcular el error cuadrático medio para las predicciones
      mse_train = mean_squared_error(y_train, y_pred_train)
      mse_test = mean_squared_error(y_test, y_pred_test)

      # Realizar la validación cruzada
      cv_scores = cross_val_score(model, cnTrainX_new, cnTrainY, cv=5, scoring='neg_mean_squared_error')

      # Los scores de la validación cruzada en scikit-learn se dan en negativo para los errores.
      # Tomamos el negativo de los scores para obtener los errores positivos, y luego tomamos la raíz cuadrada para obtener el RMSE.
      cv_rmse = np.sqrt(-cv_scores)

      # Concatenar y_train y y_pred_train_df, y_test y y_pred_test a lo largo del eje de las columnas
      y_train_and_pred = pd.concat([y_train_reset, y_pred_train_df_reset], axis=1)
      y_test_and_pred = pd.concat([y_test_reset, y_pred_test_df_reset], axis=1)

      # Añadir una columna para indicar el tipo de datos
      y_train_and_pred['Type'] = 'Train'
      y_test_and_pred['Type'] = 'Test'

      # Concatenar los DataFrames de entrenamiento y prueba
      all_data = pd.concat([y_train_and_pred, y_test_and_pred], axis=0)
      
      # Cambiar el nombre de las columnas para incluir el nombre del modelo
      all_data.columns = [f'{type(model).__name__}_{col}' if col != 'Type' else col for col in all_data.columns]
      
      # Añadir el DataFrame al listado de todas las predicciones
      all_predictions.append(all_data)
      
      # Concatenar todas las predicciones de todos los modelos
      final_predictions = pd.concat(all_predictions, axis=1)
      
      # Añadir las predicciones a una hoja en el nuevo archivo de Excel
      with pd.ExcelWriter(file_path_predictions, engine='openpyxl', mode='a') as writer:
          if df.columns[column] in writer.book.sheetnames:
              del writer.book[df.columns[column]]
          final_predictions.to_excel(writer, sheet_name=df.columns[column], index=False)
          
      # Agregar los errores y los scores al DataFrame
      # Crear un diccionario con los datos
      error_dict = {
      'Modelo': [type(model).__name__],
      'MSE Train': [mse_train],
      'MSE Test': [mse_test],
      'CV Scores': [cv_scores],
      'CV RMSE': [cv_rmse]
       }
    
      # Convertir el diccionario en un DataFrame
      model_error_data = pd.DataFrame(error_dict)
    
      # Agregar los errores y los scores al DataFrame
      error_data = pd.concat([error_data, model_error_data], ignore_index=True)

      # Añadir los errores a una hoja identificada con el nombre de las etiquetas de la columna
      with pd.ExcelWriter(file_path_errors, engine='openpyxl', mode='a') as writer:
          if df.columns[column] in writer.book.sheetnames:
           del writer.book[df.columns[column]]
          error_data.to_excel(writer, sheet_name=df.columns[column], index=False)
       

# Eliminar la hoja temporal después de agregar la primera hoja con los datos
wb = openpyxl.load_workbook(file_path)
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']
wb.save(file_path)

# Hacer lo mismo para el nuevo archivo de Excel
wb_errors = openpyxl.load_workbook(file_path_errors)
if 'Sheet' in wb_errors.sheetnames:
    del wb_errors['Sheet']
wb_errors.save(file_path_errors)

wb_predictions = openpyxl.load_workbook(file_path_predictions)
if 'Sheet' in wb_predictions.sheetnames:
    del wb_predictions['Sheet']
wb_predictions.save(file_path_predictions)


"""
18/05/2024 17:57

Tiempo de corrida: 26 seg 26 centecimas con los modelos LinearRegression(), Ridge() y Lasso()

Corrida que genera los archivos 'datasets/DieselSeleccion3.xlsx', 'datasets/DieselErrorPrediccion3.xlsx' y 'datasets/DieselPrediccion3.xlsx'

Este codigo imprime los archivos 'datasets/DieselSeleccion.xlsx', 'datasets/DieselErrorPrediccion.xlsx' y 'datasets/DieselPrediccion.xlsx'. 'datasets/DieselSeleccion.xlsx' contiene las caracteristicas mas relevantes para hacer la prediccion de las etiquetas. Se evaluan varios modelos de regresion y se guardan los errores en 'datasets/DieselErrorPrediccion.xlsx' y las predicciones en 'datasets/DieselPrediccion.xlsx'. 
"""

import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LinearRegression, Ridge, Lasso, ElasticNet, SGDRegressor, BayesianRidge
from sklearn.neighbors import KNeighborsRegressor
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor, ExtraTreesRegressor, AdaBoostRegressor, BaggingRegressor, GradientBoostingRegressor, HistGradientBoostingRegressor
from sklearn.svm import SVR
from sklearn.neural_network import MLPRegressor
from xgboost import XGBRegressor
from sklearn.model_selection import train_test_split, cross_val_score  
from sklearn.feature_selection import SelectKBest, f_regression
from sklearn.metrics import mean_squared_error
import openpyxl
import os

# Función para crear archivo si no existe
def create_file_if_not_exists(file_path):
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        wb.save(file_path)

# Cargar los datos
df = pd.read_excel('datasets/Modelo Diesel.xlsx', sheet_name='SNV')

# Crear los archivos necesarios si no existen
create_file_if_not_exists('datasets/DieselSeleccion.xlsx')
create_file_if_not_exists('datasets/DieselPrediccion.xlsx')
create_file_if_not_exists('datasets/DieselErrorPrediccion.xlsx')

# Inicializar los modelos
models = [LinearRegression(), Ridge(), Lasso(), ElasticNet(max_iter=1000), KNeighborsRegressor(),
          DecisionTreeRegressor(), RandomForestRegressor(), SVR(), MLPRegressor(max_iter=10000), 
          ExtraTreesRegressor(), SGDRegressor(), AdaBoostRegressor(), BaggingRegressor(), 
          GradientBoostingRegressor(), HistGradientBoostingRegressor(),
          XGBRegressor(), BayesianRidge()]

# Bucle a través de las columnas que quieres usar como etiquetas
for column in range(2354, 2335, -1):  # Range(2354, 2335, -1) en excel es (2353(CLO), 2335(CKW)..
    column_name = df.columns[column]

    # Dividir los datos en características (X) y etiquetas (y)
    # Por defecto si hay identificacion en el encabezado de columna, la indexacion 0 de python empieza
    # en la segunda fila de excel, por tanto 
    # df.iloc[0:361, 2:2336] en excel son las columnas [Filas 2:358, Columnas 3:2335]
    cnTrainX = df.iloc[0:361, 2:2336]
    cnTrainY = df.iloc[0:361, column]

    # Eliminar las filas con valores NaN en cnTrainY
    cnTrainX = cnTrainX[~cnTrainY.isna()]
    cnTrainY = cnTrainY.dropna()

    # Realizar selección de características
    feature_selection = SelectKBest(score_func=f_regression, k=50)
    cnTrainX_new = feature_selection.fit_transform(cnTrainX, cnTrainY)

    # Crear DataFrame con las puntuaciones de las características
    feature_scores = pd.DataFrame({'Feature': cnTrainX.columns, 'Score': feature_selection.scores_}).nlargest(50, 'Score')

    # Añadir las características principales a una nueva hoja en el archivo Excel
    with pd.ExcelWriter('datasets/DieselSeleccion.xlsx', engine='openpyxl', mode='a') as writer:
        feature_scores.to_excel(writer, sheet_name=column_name, index=False)

    # Dividir los datos en conjuntos de entrenamiento y prueba
    X_train, X_test, y_train, y_test = train_test_split(cnTrainX_new, cnTrainY, test_size=0.3, train_size=0.7, shuffle=True)

    # Escalar las características
    sc = StandardScaler()
    X_train = sc.fit_transform(X_train)
    X_test = sc.transform(X_test)

    # DataFrames para almacenar resultados y errores
    all_predictions = pd.DataFrame()
    error_data = []

    for model in models:
        model_name = type(model).__name__

        # Entrenar el modelo
        model.fit(X_train, y_train)

        # Realizar predicciones
        y_pred_train = model.predict(X_train)
        y_pred_test = model.predict(X_test)

        # Calcular el error cuadrático medio para las predicciones
        mse_train = mean_squared_error(y_train, y_pred_train)
        mse_test = mean_squared_error(y_test, y_pred_test)

        # Realizar la validación cruzada
        cv_scores = cross_val_score(model, cnTrainX_new, cnTrainY, cv=5, scoring='neg_mean_squared_error')
        cv_rmse = np.sqrt(-cv_scores)

        # Concatenar resultados
        train_results = pd.DataFrame({'y_train': y_train, f'y_pred_train_{model_name}': y_pred_train})
        test_results = pd.DataFrame({'y_test': y_test, f'y_pred_test_{model_name}': y_pred_test})
        all_predictions = pd.concat([all_predictions, train_results, test_results], axis=1)

        # Guardar errores
        error_data.append({
            'Modelo': model_name,
            'MSE Train': mse_train,
            'MSE Test': mse_test,
            'CV Scores': cv_scores,
            'CV RMSE': cv_rmse
        })

    # Guardar las predicciones y los errores en sus respectivos archivos de Excel
    with pd.ExcelWriter('datasets/DieselPrediccion.xlsx', engine='openpyxl', mode='a') as writer:
        all_predictions.to_excel(writer, sheet_name=column_name, index=False)

    with pd.ExcelWriter('datasets/DieselErrorPrediccion.xlsx', engine='openpyxl', mode='a') as writer:
        pd.DataFrame(error_data).to_excel(writer, sheet_name=column_name, index=False)

# Eliminar la hoja temporal después de agregar la primera hoja con los datos
for file_path in ['datasets/DieselSeleccion.xlsx', 'datasets/DieselPrediccion.xlsx', 'datasets/DieselErrorPrediccion.xlsx']:
    wb = openpyxl.load_workbook(file_path)
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    wb.save(file_path)
    

"""
18/05/2024 18:50

Tiempo de corrida: 26 seg 26 centecimas con los modelos LinearRegression(), Ridge() y Lasso()

Corrida que genera los archivos 'datasets/DieselSeleccion3.xlsx', 'datasets/DieselErrorPrediccion3.xlsx' y 'datasets/DieselPrediccion3.xlsx'

Este codigo imprime los archivos 'datasets/DieselSeleccion.xlsx', 'datasets/DieselErrorPrediccion.xlsx' y 'datasets/DieselPrediccion.xlsx'. 'datasets/DieselSeleccion.xlsx' contiene las caracteristicas mas relevantes para hacer la prediccion de las etiquetas. Se evaluan varios modelos de regresion y se guardan los errores en 'datasets/DieselErrorPrediccion.xlsx' y las predicciones en 'datasets/DieselPrediccion.xlsx'. Este código incluye la creación de la carpeta models y el almacenamiento de los modelos quimiometricos en ella.
"""


import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LinearRegression, Ridge, Lasso, ElasticNet, SGDRegressor, BayesianRidge
from sklearn.neighbors import KNeighborsRegressor
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor, ExtraTreesRegressor, AdaBoostRegressor, BaggingRegressor, GradientBoostingRegressor, HistGradientBoostingRegressor
from sklearn.svm import SVR
from sklearn.neural_network import MLPRegressor
from xgboost import XGBRegressor
from sklearn.model_selection import train_test_split, cross_val_score  
from sklearn.feature_selection import SelectKBest, f_regression
from sklearn.metrics import mean_squared_error
import openpyxl
import os
import joblib

# Función para crear archivo si no existe
def create_file_if_not_exists(file_path):
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        wb.save(file_path)

# Cargar los datos
df = pd.read_excel('datasets/Modelo Diesel.xlsx', sheet_name='SNV')

# Crear los archivos necesarios si no existen
create_file_if_not_exists('datasets/DieselSeleccion.xlsx')
create_file_if_not_exists('datasets/DieselPrediccion.xlsx')
create_file_if_not_exists('datasets/DieselErrorPrediccion.xlsx')

# Inicializar los modelos
models = [LinearRegression(), Ridge(), Lasso(), ElasticNet(max_iter=1000), KNeighborsRegressor(),
          DecisionTreeRegressor(), RandomForestRegressor(), SVR(), MLPRegressor(max_iter=10000), 
          ExtraTreesRegressor(), SGDRegressor(), AdaBoostRegressor(), BaggingRegressor(), 
          GradientBoostingRegressor(), HistGradientBoostingRegressor(),
          XGBRegressor(), BayesianRidge()]

# Crear el directorio models en caso de que no exista.
models_dir = 'models'
if not os.path.exists(models_dir):
    os.makedirs(models_dir)

# Bucle a través de las columnas que quieres usar como etiquetas
for column in range(2354, 2335, -1):  # Range(2354, 2335, -1) en excel es (2353(CLO), 2335(CKW)..
    column_name = df.columns[column]

    # Dividir los datos en características (X) y etiquetas (y)
    # Por defecto si hay identificacion en el encabezado de columna, la indexacion 0 de python empieza
    # en la segunda fila de excel, por tanto 
    # df.iloc[0:361, 2:2336] en excel son las columnas [Filas 2:358, Columnas 3:2335]
    cnTrainX = df.iloc[0:361, 2:2336]
    cnTrainY = df.iloc[0:361, column]

    # Eliminar las filas con valores NaN en cnTrainY
    cnTrainX = cnTrainX[~cnTrainY.isna()]
    cnTrainY = cnTrainY.dropna()

    # Realizar selección de características
    feature_selection = SelectKBest(score_func=f_regression, k=50)
    cnTrainX_new = feature_selection.fit_transform(cnTrainX, cnTrainY)

    # Crear DataFrame con las puntuaciones de las características
    feature_scores = pd.DataFrame({'Feature': cnTrainX.columns, 'Score': feature_selection.scores_}).nlargest(50, 'Score')

    # Añadir las características principales a una nueva hoja en el archivo Excel
    with pd.ExcelWriter('datasets/DieselSeleccion.xlsx', engine='openpyxl', mode='a') as writer:
        feature_scores.to_excel(writer, sheet_name=column_name, index=False)

    # Dividir los datos en conjuntos de entrenamiento y prueba
    X_train, X_test, y_train, y_test = train_test_split(cnTrainX_new, cnTrainY, test_size=0.3, train_size=0.7, shuffle=True)

    # Escalar las características
    sc = StandardScaler()
    X_train = sc.fit_transform(X_train)
    X_test = sc.transform(X_test)

    # DataFrames para almacenar resultados y errores
    all_predictions = pd.DataFrame()
    error_data = []

    for model in models:
        model_name = type(model).__name__

        # Entrenar el modelo
        model.fit(X_train, y_train)

        # Realizar predicciones
        y_pred_train = model.predict(X_train)
        y_pred_test = model.predict(X_test)
        
        # Guardar el modelo entrenado
        model_file_path = os.path.join(models_dir, f'{model_name}.joblib')
        joblib.dump(model, model_file_path)

        # Calcular el error cuadrático medio para las predicciones
        mse_train = mean_squared_error(y_train, y_pred_train)
        mse_test = mean_squared_error(y_test, y_pred_test)

        # Realizar la validación cruzada
        cv_scores = cross_val_score(model, cnTrainX_new, cnTrainY, cv=5, scoring='neg_mean_squared_error')
        cv_rmse = np.sqrt(-cv_scores)

        # Concatenar resultados
        train_results = pd.DataFrame({'y_train': y_train, f'y_pred_train_{model_name}': y_pred_train})
        test_results = pd.DataFrame({'y_test': y_test, f'y_pred_test_{model_name}': y_pred_test})
        all_predictions = pd.concat([all_predictions, train_results, test_results], axis=1)

        # Guardar errores
        error_data.append({
            'Modelo': model_name,
            'MSE Train': mse_train,
            'MSE Test': mse_test,
            'CV Scores': cv_scores,
            'CV RMSE': cv_rmse
        })

    # Guardar las predicciones y los errores en sus respectivos archivos de Excel
    with pd.ExcelWriter('datasets/DieselPrediccion.xlsx', engine='openpyxl', mode='a') as writer:
        all_predictions.to_excel(writer, sheet_name=column_name, index=False)

    with pd.ExcelWriter('datasets/DieselErrorPrediccion.xlsx', engine='openpyxl', mode='a') as writer:
        pd.DataFrame(error_data).to_excel(writer, sheet_name=column_name, index=False)

# Eliminar la hoja temporal después de agregar la primera hoja con los datos
for file_path in ['datasets/DieselSeleccion.xlsx', 'datasets/DieselPrediccion.xlsx', 'datasets/DieselErrorPrediccion.xlsx']:
    wb = openpyxl.load_workbook(file_path)
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    wb.save(file_path)
    


"""
19/05/2024 7:42

Tiempo de corrida: 16 seg 14 centecimas con todos los modelos aqui descritos

Corrida que genera los archivos 'datasets/3DieselSeleccion.xlsx', 'datasets/3DieselErrorPrediccion.xlsx' y 'datasets/3DieselPrediccion.xlsx'

Este codigo imprime los archivos 'datasets/DieselSeleccion.xlsx', 'datasets/DieselErrorPrediccion.xlsx' y 'datasets/DieselPrediccion.xlsx'. 'datasets/DieselSeleccion.xlsx' contiene las características mas relevantes para hacer la prediccion de las etiquetas. Se evaluan varios modelos de regresion y se guardan los errores en 'datasets/DieselErrorPrediccion.xlsx' y las predicciones en 'datasets/DieselPrediccion.xlsx'. Este código incluye la creación de la carpeta models y el almacenamiento de los modelos quimiometricos en ella, el promedio de CV Scores y CV RMSE.
"""


import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LinearRegression, Ridge, Lasso, ElasticNet, SGDRegressor, BayesianRidge
from sklearn.neighbors import KNeighborsRegressor
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor, ExtraTreesRegressor, AdaBoostRegressor, BaggingRegressor, GradientBoostingRegressor, HistGradientBoostingRegressor
from sklearn.svm import SVR
from sklearn.neural_network import MLPRegressor
from xgboost import XGBRegressor
from sklearn.model_selection import train_test_split, cross_val_score  
from sklearn.feature_selection import SelectKBest, f_regression
from sklearn.metrics import mean_squared_error
import openpyxl
import os
import joblib

# Función para crear archivo si no existe
def create_file_if_not_exists(file_path):
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        wb.save(file_path)

# Cargar los datos
df = pd.read_excel('datasets/Modelo Diesel.xlsx', sheet_name='SNV')

# Crear los archivos necesarios si no existen
create_file_if_not_exists('datasets/DieselSeleccion.xlsx')
create_file_if_not_exists('datasets/DieselPrediccion.xlsx')
create_file_if_not_exists('datasets/DieselErrorPrediccion.xlsx')

# Inicializar los modelos
models = [LinearRegression(), Ridge(), Lasso(), ElasticNet(max_iter=1000), KNeighborsRegressor(),
          DecisionTreeRegressor(), RandomForestRegressor(), SVR(), MLPRegressor(max_iter=10000), 
          ExtraTreesRegressor(), SGDRegressor(), AdaBoostRegressor(), BaggingRegressor(), 
          GradientBoostingRegressor(), HistGradientBoostingRegressor(),
          XGBRegressor(), BayesianRidge()]

# Create models directory if it doesn't exist
models_dir = 'models'
if not os.path.exists(models_dir):
    os.makedirs(models_dir)

# Bucle a través de las columnas que quieres usar como etiquetas
for column in range(2354, 2335, -1):  # Range(2354, 2335, -1) en excel es (2353(CLO), 2335(CKW)..
    column_name = df.columns[column]

    # Dividir los datos en características (X) y etiquetas (y)
    # Por defecto si hay identificacion en el encabezado de columna, la indexacion 0 de python empieza
    # en la segunda fila de excel, por tanto 
    # df.iloc[0:361, 2:2336] en excel son las columnas [Filas 2:358, Columnas 3:2335]
    cnTrainX = df.iloc[0:361, 2:2336]
    cnTrainY = df.iloc[0:361, column]

    # Eliminar las filas con valores NaN en cnTrainY
    cnTrainX = cnTrainX[~cnTrainY.isna()]
    cnTrainY = cnTrainY.dropna()

    # Realizar selección de características
    feature_selection = SelectKBest(score_func=f_regression, k=50)
    cnTrainX_new = feature_selection.fit_transform(cnTrainX, cnTrainY)

    # Crear DataFrame con las puntuaciones de las características
    feature_scores = pd.DataFrame({'Feature': cnTrainX.columns, 'Score': feature_selection.scores_}).nlargest(50, 'Score')

    # Añadir las características principales a una nueva hoja en el archivo Excel
    with pd.ExcelWriter('datasets/DieselSeleccion.xlsx', engine='openpyxl', mode='a') as writer:
        feature_scores.to_excel(writer, sheet_name=column_name, index=False)

    # Dividir los datos en conjuntos de entrenamiento y prueba
    X_train, X_test, y_train, y_test = train_test_split(cnTrainX_new, cnTrainY, test_size=0.3, train_size=0.7, shuffle=True)

    # Escalar las características
    sc = StandardScaler()
    X_train = sc.fit_transform(X_train)
    X_test = sc.transform(X_test)

    # DataFrames para almacenar resultados y errores
    all_predictions = pd.DataFrame()
    error_data = []

    for model in models:
        model_name = type(model).__name__

        # Entrenar el modelo
        model.fit(X_train, y_train)

        # Realizar predicciones
        y_pred_train = model.predict(X_train)
        y_pred_test = model.predict(X_test)
        
        # Guardar el modelo entrenado
        model_file_path = os.path.join(models_dir, f'{column_name}_{model_name}.joblib')
        joblib.dump(model, model_file_path)

        # Calcular el error cuadrático medio para las predicciones
        mse_train = mean_squared_error(y_train, y_pred_train)
        mse_test = mean_squared_error(y_test, y_pred_test)

        # Realizar la validación cruzada
        cv_scores = cross_val_score(model, cnTrainX_new, cnTrainY, cv=5, scoring='neg_mean_squared_error')
        cv_rmse = np.sqrt(-cv_scores)
        
        # Calcular el promedio de cv_scores y cv_rmse
        cv_scores_mean = np.mean(cv_scores)
        cv_rmse_mean = np.mean(cv_rmse)

        # Concatenar resultados
        train_results = pd.DataFrame({'y_train': y_train, f'y_pred_train_{model_name}': y_pred_train})
        test_results = pd.DataFrame({'y_test': y_test, f'y_pred_test_{model_name}': y_pred_test})
        all_predictions = pd.concat([all_predictions, train_results, test_results], axis=1)

        # Guardar errores
        # Guardar errores
        error_data.append({
        'Modelo': model_name,
        'MSE Train': mse_train,
        'MSE Test': mse_test,
        'CV Score 1': cv_scores[0],
        'CV Score 2': cv_scores[1],
        'CV Score 3': cv_scores[2],
        'CV Score 4': cv_scores[3],
        'CV Score 5': cv_scores[4],
        'μ CV Score': cv_scores_mean,
        'CV RMSE 1': cv_rmse[0],
        'CV RMSE 2': cv_rmse[1],
        'CV RMSE 3': cv_rmse[2],
        'CV RMSE 4': cv_rmse[3],
        'CV RMSE 5': cv_rmse[4],
        'μ CV RMSE': cv_rmse_mean
         })

    # Guardar las predicciones y los errores en sus respectivos archivos de Excel
    with pd.ExcelWriter('datasets/DieselPrediccion.xlsx', engine='openpyxl', mode='a') as writer:
        all_predictions.to_excel(writer, sheet_name=column_name, index=False)

    with pd.ExcelWriter('datasets/DieselErrorPrediccion.xlsx', engine='openpyxl', mode='a') as writer:
        pd.DataFrame(error_data).to_excel(writer, sheet_name=column_name, index=False)

# Eliminar la hoja temporal después de agregar la primera hoja con los datos
for file_path in ['datasets/DieselSeleccion.xlsx', 'datasets/DieselPrediccion.xlsx', 'datasets/DieselErrorPrediccion.xlsx']:
    wb = openpyxl.load_workbook(file_path)
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    wb.save(file_path)
    

"""
19/05/2024 9:46

Tiempo de corrida: 16 seg 14 centecimas con todos los modelos aqui descritos

Corrida que genera los archivos 'datasets/3DieselSeleccion.xlsx', 'datasets/3DieselErrorPrediccion.xlsx' y 'datasets/3DieselPrediccion.xlsx'

Este codigo imprime los archivos 'datasets/DieselSeleccion.xlsx', 'datasets/DieselErrorPrediccion.xlsx' y 'datasets/DieselPrediccion.xlsx'. 'datasets/DieselSeleccion.xlsx' contiene las características mas relevantes para hacer la prediccion de las etiquetas. Se evaluan varios modelos de regresion y se guardan los errores en 'datasets/DieselErrorPrediccion.xlsx' y las predicciones en 'datasets/DieselPrediccion.xlsx'. Este código incluye la creación de la carpeta models, el almacenamiento de los modelos quimiometricos en ella, el promedio de CV Scores y CV RMSE y la creación de los archivos
.joblib con las mejores características o longitudes de onda con mayor influencia en la etiqueta  
"""

import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LinearRegression, Ridge, Lasso, ElasticNet, SGDRegressor, BayesianRidge
from sklearn.neighbors import KNeighborsRegressor
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor, ExtraTreesRegressor, AdaBoostRegressor, BaggingRegressor, GradientBoostingRegressor, HistGradientBoostingRegressor
from sklearn.svm import SVR
from sklearn.neural_network import MLPRegressor
from xgboost import XGBRegressor
from sklearn.model_selection import train_test_split, cross_val_score  
from sklearn.feature_selection import SelectKBest, f_regression
from sklearn.metrics import mean_squared_error
import openpyxl
import os
import joblib

# Función para crear archivo si no existe
def create_file_if_not_exists(file_path):
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        wb.save(file_path)

# Cargar los datos
df = pd.read_excel('datasets/Modelo Diesel.xlsx', sheet_name='SNV')

# Crear los archivos necesarios si no existen
create_file_if_not_exists('datasets/DieselSeleccion.xlsx')
create_file_if_not_exists('datasets/DieselPrediccion.xlsx')
create_file_if_not_exists('datasets/DieselErrorPrediccion.xlsx')

# Inicializar los modelos
models = [LinearRegression(), Ridge(), Lasso(), ElasticNet(max_iter=1000), KNeighborsRegressor(),
          DecisionTreeRegressor(), RandomForestRegressor(), SVR(), MLPRegressor(max_iter=10000), 
          ExtraTreesRegressor(), SGDRegressor(), AdaBoostRegressor(), BaggingRegressor(), 
          GradientBoostingRegressor(), HistGradientBoostingRegressor(),
          XGBRegressor(), BayesianRidge()]

# Create models directory if it doesn't exist
models_dir = 'models'
if not os.path.exists(models_dir):
    os.makedirs(models_dir)

# Bucle a través de las columnas que quieres usar como etiquetas
for column in range(2354, 2335, -1):  # Range(2354, 2335, -1) en excel es (2353(CLO), 2335(CKW)..
    column_name = df.columns[column]

    # Dividir los datos en características (X) y etiquetas (y)
    # Por defecto si hay identificacion en el encabezado de columna, la indexacion 0 de python empieza
    # en la segunda fila de excel, por tanto 
    # df.iloc[0:361, 2:2336] en excel son las columnas [Filas 2:358, Columnas 3:2335]

    cnTrainX = df.iloc[0:361, 2:2336]
    cnTrainY = df.iloc[0:361, column]

    # Eliminar las filas con valores NaN en cnTrainY
    cnTrainX = cnTrainX[~cnTrainY.isna()]
    cnTrainY = cnTrainY.dropna()

    # Realizar selección de características
    feature_selection = SelectKBest(score_func=f_regression, k=50)
    cnTrainX_new = feature_selection.fit_transform(cnTrainX, cnTrainY)
    selected_features = cnTrainX.columns[feature_selection.get_support()]
    
    # Guardar las características seleccionadas en un archivo .joblib
    features_file_path = os.path.join(models_dir, f'SelectKBest_{column_name}_features.joblib')
    joblib.dump(selected_features, features_file_path)


    # Crear DataFrame con las puntuaciones de las características
    feature_scores = pd.DataFrame({'Feature': cnTrainX.columns, 'Score': feature_selection.scores_}).nlargest(50, 'Score')

    # Añadir las características principales a una nueva hoja en el archivo Excel
    with pd.ExcelWriter('datasets/DieselSeleccion.xlsx', engine='openpyxl', mode='a') as writer:
        feature_scores.to_excel(writer, sheet_name=column_name, index=False)

    # Dividir los datos en conjuntos de entrenamiento y prueba
    X_train, X_test, y_train, y_test = train_test_split(cnTrainX_new, cnTrainY, test_size=0.3, train_size=0.7, shuffle=True)
    
    # Guardar el escalador ajustado
    scaler_file_path = os.path.join(models_dir, f'Scaler_{column_name}.joblib')
    joblib.dump(sc, scaler_file_path)

    # Escalar las características
    sc = StandardScaler()
    X_train = sc.fit_transform(X_train)
    X_test = sc.transform(X_test)

    # DataFrames para almacenar resultados y errores
    all_predictions = pd.DataFrame()
    error_data = []

    for model in models:
        model_name = type(model).__name__

        # Entrenar el modelo
        model.fit(X_train, y_train)

        # Realizar predicciones
        y_pred_train = model.predict(X_train)
        y_pred_test = model.predict(X_test)
        
        # Guardar el modelo entrenado
        model_file_path = os.path.join(models_dir, f'{column_name}_{model_name}.joblib')
        joblib.dump(model, model_file_path)

        # Calcular el error cuadrático medio para las predicciones
        mse_train = mean_squared_error(y_train, y_pred_train)
        mse_test = mean_squared_error(y_test, y_pred_test)

        # Realizar la validación cruzada
        cv_scores = cross_val_score(model, cnTrainX_new, cnTrainY, cv=5, scoring='neg_mean_squared_error')
        cv_rmse = np.sqrt(-cv_scores)
        
        # Calcular el promedio de cv_scores y cv_rmse
        cv_scores_mean = np.mean(cv_scores)
        cv_rmse_mean = np.mean(cv_rmse)

        # Concatenar resultados
        train_results = pd.DataFrame({'y_train': y_train, f'y_pred_train_{model_name}': y_pred_train})
        test_results = pd.DataFrame({'y_test': y_test, f'y_pred_test_{model_name}': y_pred_test})
        all_predictions = pd.concat([all_predictions, train_results, test_results], axis=1)

        # Guardar errores
        # Guardar errores
        error_data.append({
        'Modelo': model_name,
        'MSE Train': mse_train,
        'MSE Test': mse_test,
        'CV Score 1': cv_scores[0],
        'CV Score 2': cv_scores[1],
        'CV Score 3': cv_scores[2],
        'CV Score 4': cv_scores[3],
        'CV Score 5': cv_scores[4],
        'μ CV Score': cv_scores_mean,
        'CV RMSE 1': cv_rmse[0],
        'CV RMSE 2': cv_rmse[1],
        'CV RMSE 3': cv_rmse[2],
        'CV RMSE 4': cv_rmse[3],
        'CV RMSE 5': cv_rmse[4],
        'μ CV RMSE': cv_rmse_mean
         })

    # Guardar las predicciones y los errores en sus respectivos archivos de Excel
    with pd.ExcelWriter('datasets/DieselPrediccion.xlsx', engine='openpyxl', mode='a') as writer:
        all_predictions.to_excel(writer, sheet_name=column_name, index=False)

    with pd.ExcelWriter('datasets/DieselErrorPrediccion.xlsx', engine='openpyxl', mode='a') as writer:
        pd.DataFrame(error_data).to_excel(writer, sheet_name=column_name, index=False)

# Eliminar la hoja temporal después de agregar la primera hoja con los datos
for file_path in ['datasets/DieselSeleccion.xlsx', 'datasets/DieselPrediccion.xlsx', 'datasets/DieselErrorPrediccion.xlsx']:
    wb = openpyxl.load_workbook(file_path)
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    wb.save(file_path)


"""
20/05/2024 3:24

Tiempo de corrida: 16 seg 14 centecimas con todos los modelos aqui descritos

Corrida que genera los archivos 'datasets/3DieselSeleccion.xlsx', 'datasets/3DieselErrorPrediccion.xlsx' y 'datasets/3DieselPrediccion.xlsx'

Este codigo imprime los archivos 'datasets/DieselSeleccion.xlsx', 'datasets/DieselErrorPrediccion.xlsx' y 'datasets/DieselPrediccion.xlsx'. 'datasets/DieselSeleccion.xlsx' contiene las características mas relevantes para hacer la prediccion de las etiquetas. Se evaluan varios modelos de regresion y se guardan los errores en 'datasets/DieselErrorPrediccion.xlsx' y las predicciones en 'datasets/DieselPrediccion.xlsx'. Este código incluye la creación de la carpeta models, el almacenamiento de los modelos quimiometricos en ella, el promedio de CV Scores y CV RMSE, ponderado con pesos del error cuadratico medio (μ_AllMseRMSE = mse_train*0.1 + mse_test*0.4 + cv_scores_mean*-1*0.5) y la creación de los archivos .joblib con las mejores características o longitudes de onda con mayor influencia en la etiqueta. OJO Para contar las columnas y filas en excel, pilas que solo cuanta las celdas acupadas, las vacias no las cuenta  
"""

# Modificación creación del modelo 20/05/2024 10:55. Corrijio sobreescritura de archivo de modelo de regresion. Tiempo corrida 21 min 12 seg 81 cet.

import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LinearRegression, Ridge, Lasso, ElasticNet, SGDRegressor, BayesianRidge
from sklearn.neighbors import KNeighborsRegressor
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor, ExtraTreesRegressor, AdaBoostRegressor, BaggingRegressor, GradientBoostingRegressor, HistGradientBoostingRegressor
from sklearn.svm import SVR
from sklearn.neural_network import MLPRegressor
from xgboost import XGBRegressor
from sklearn.model_selection import train_test_split, cross_val_score  
from sklearn.feature_selection import SelectKBest, f_regression
from sklearn.metrics import mean_squared_error
import openpyxl
import os
import joblib

# Función para crear archivo si no existe
def create_file_if_not_exists(file_path):
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        wb.save(file_path)

# Cargar los datos
df = pd.read_excel('datasets/Modelo Diesel.xlsx', sheet_name='SNV')

# Crear los archivos necesarios si no existen
create_file_if_not_exists('datasets/DieselSeleccion.xlsx')
create_file_if_not_exists('datasets/DieselPrediccion.xlsx')
create_file_if_not_exists('datasets/DieselErrorPrediccion.xlsx')

# Inicializar los modelos
models = [LinearRegression(), Ridge(), Lasso(), ElasticNet(max_iter=1000), KNeighborsRegressor(),
          DecisionTreeRegressor(), RandomForestRegressor(), SVR(), MLPRegressor(max_iter=10000), 
          ExtraTreesRegressor(), SGDRegressor(), AdaBoostRegressor(), BaggingRegressor(), 
          GradientBoostingRegressor(), HistGradientBoostingRegressor(),
          XGBRegressor(), BayesianRidge()]

# Crear el directorio models si no existe
models_dir = 'models'
if not os.path.exists(models_dir):
    os.makedirs(models_dir)

# Etiquetas que quieres usar
labels = ['NCGRB', 'ICASTM', 'POL', 'MON', 'ATW', 'ATV', 'PNU', 'VIS', 'FLU', 'PI', 
          'API', 'BIO', 'S', 'PFE', 'X95', 'X90', 'X50', 'X10', 'PIE']

# Bucle a través de las etiquetas
for label in labels:
    # Dividir los datos en características (X) y etiquetas (y)
    cnTrainX = df.iloc[0:361, 0:2334]  
    cnTrainY = df[label]

    # Eliminar las filas con valores NaN en cnTrainY
    cnTrainX = cnTrainX[~cnTrainY.isna()]
    cnTrainY = cnTrainY.dropna()

    # Realizar selección de características
    feature_selection = SelectKBest(score_func=f_regression, k=50)
    cnTrainX_new = feature_selection.fit_transform(cnTrainX, cnTrainY)
    selected_features = cnTrainX.columns[feature_selection.get_support()]
    
    # Guardar las características seleccionadas en un archivo .joblib
    features_file_path = os.path.join(models_dir, f'SelectKBest_{label}_features.joblib')
    joblib.dump(selected_features, features_file_path)

    # Crear DataFrame con las puntuaciones de las características
    feature_scores = pd.DataFrame({'Feature': cnTrainX.columns, 'Score': feature_selection.scores_}).nlargest(50, 'Score')

    # Añadir las características principales a una nueva hoja en el archivo Excel
    with pd.ExcelWriter('datasets/DieselSeleccion.xlsx', engine='openpyxl', mode='a') as writer:
        feature_scores.to_excel(writer, sheet_name=label, index=False)

    # Dividir los datos en conjuntos de entrenamiento y prueba
    X_train, X_test, y_train, y_test = train_test_split(cnTrainX_new, cnTrainY, test_size=0.3, train_size=0.7, shuffle=True)

    # Escalar las características
    sc = StandardScaler()
    X_train = sc.fit_transform(X_train)
    X_test = sc.transform(X_test)
    
    # Guardar el escalador ajustado
    scaler_file_path = os.path.join(models_dir, f'Scaler_{label}.joblib')
    joblib.dump(sc, scaler_file_path)

    # DataFrames para almacenar resultados y errores
    all_predictions = pd.DataFrame()
    error_data = []

    for model in models:
        model_name = type(model).__name__

        # Entrenar el modelo
        model.fit(X_train, y_train)

        # Realizar predicciones
        y_pred_train = model.predict(X_train)
        y_pred_test = model.predict(X_test)
        
        # Guardar el modelo entrenado
        model_file_path = os.path.join(models_dir, f'{label}_{model_name}.joblib')
        joblib.dump(model, model_file_path)

        # Calcular el error cuadrático medio para las predicciones
        mse_train = mean_squared_error(y_train, y_pred_train)
        mse_test = mean_squared_error(y_test, y_pred_test)

        # Realizar la validación cruzada
        cv_scores = cross_val_score(model, cnTrainX_new, cnTrainY, cv=5, scoring='neg_mean_squared_error')
        cv_rmse = np.sqrt(-cv_scores)
        
        # Calcular el promedio de cv_scores y cv_rmse
        cv_scores_mean = np.mean(cv_scores)
        cv_rmse_mean = np.mean(cv_rmse)

        # Concatenar resultados
        train_results = pd.DataFrame({'y_train': y_train, f'y_pred_train_{model_name}': y_pred_train})
        test_results = pd.DataFrame({'y_test': y_test, f'y_pred_test_{model_name}': y_pred_test})
        all_predictions = pd.concat([all_predictions, train_results, test_results], axis=1)

        μ_AllMseRMSE = mse_train * 0.1 + mse_test * 0.4 + cv_scores_mean * -1 * 0.5
        
        # Guardar errores
        error_data.append({
            'Modelo': model_name,
            'MSE Train': mse_train,
            'MSE Test': mse_test,
            'CV Score 1': cv_scores[0],
            'CV Score 2': cv_scores[1],
            'CV Score 3': cv_scores[2],
            'CV Score 4': cv_scores[3],
            'CV Score 5': cv_scores[4],
            'μ CV Score': cv_scores_mean,
            'CV RMSE 1': cv_rmse[0],
            'CV RMSE 2': cv_rmse[1],
            'CV RMSE 3': cv_rmse[2],
            'CV RMSE 4': cv_rmse[3],
            'CV RMSE 5': cv_rmse[4],
            'μ CV RMSE': cv_rmse_mean,
            'μ All MSE RMSE': μ_AllMseRMSE
        })

    # Guardar las predicciones y los errores en sus respectivos archivos de Excel
    with pd.ExcelWriter('datasets/DieselPrediccion.xlsx', engine='openpyxl', mode='a') as writer:
        all_predictions.to_excel(writer, sheet_name=label, index=False)

    with pd.ExcelWriter('datasets/DieselErrorPrediccion.xlsx', engine='openpyxl', mode='a') as writer:
        pd.DataFrame(error_data).to_excel(writer, sheet_name=label, index=False)

# Eliminar la hoja temporal después de agregar la primera hoja con los datos
for file_path in ['datasets/DieselSeleccion.xlsx', 'datasets/DieselPrediccion.xlsx', 'datasets/DieselErrorPrediccion.xlsx']:
    wb = openpyxl.load_workbook(file_path)
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    wb.save(file_path)


# Despues de correr el codigo anterior de entrenamiento, se puede predecir con este codigo.

import os
import pandas as pd
import joblib
import warnings
from sklearn.exceptions import DataConversionWarning
from sklearn.preprocessing import StandardScaler

# Ignorar las advertencias de DataConversionWarning
warnings.filterwarnings(action='ignore', category=DataConversionWarning)

# Define los directorios y archivos
models_dir = 'models/diesel'
Scal_Featu_dir = 'models'
sample_file_path = 'datasets/Muestra Diesel.xlsx'
output_file_path = 'datasets/Predicciones Diesel.xlsx'

# Lista de etiquetas
labels = ['NCGRB', 'ICASTM', 'POL', 'MON', 'ATW', 'ATV', 'PNU', 'VIS', 'FLU', 'PI', 
          'API', 'BIO', 'S', 'PFE', 'X95', 'X90', 'X50', 'X10', 'PIE']

# Cargar la muestra
sample_df = pd.read_excel(sample_file_path)

# Diccionario para almacenar las predicciones
predictions = {}

for label in labels:
    # Cargar las características seleccionadas
    features_file_path = os.path.join(Scal_Featu_dir, f'SelectKBest_{label}_features.joblib')
    selected_features = joblib.load(features_file_path)
    
    # Cargar el escalador
    scaler_file_path = os.path.join(Scal_Featu_dir, f'Scaler_{label}.joblib')
    scaler = joblib.load(scaler_file_path)
    
    # Seleccionar las características de la muestra
    X_sample = sample_df[selected_features]
    
    # Escalar las características
    X_sample_scaled = scaler.transform(X_sample.values)  # Use .values to ignore feature names
    
    # Cargar el modelo
    model_file_path = os.path.join(models_dir, f'{label}.joblib')
    model = joblib.load(model_file_path)
    
    # Realizar la predicción
    predictions[label] = model.predict(X_sample_scaled)

# Convertir las predicciones a un DataFrame
predictions_df = pd.DataFrame(predictions)

# Guardar las predicciones en un archivo Excel
predictions_df.to_excel(output_file_path, index=False)

print(f"Predicciones guardadas en {output_file_path}")


# Otra forma de predecir las etiquetas de una muestra:

import os
import pandas as pd
import joblib
import warnings
from sklearn.exceptions import DataConversionWarning
from sklearn.preprocessing import StandardScaler

# Ignorar las advertencias de DataConversionWarning
warnings.filterwarnings(action='ignore', category=DataConversionWarning)

# Define los directorios y archivos
models_dir = 'models/diesel'
Scal_Featu_dir = 'models'
sample_file_path = 'datasets/Muestra Diesel.xlsx'
output_file_path = 'datasets/Predicciones Diesel.xlsx'

# Lista de etiquetas
labels = ['NCGRB', 'ICASTM', 'POL', 'MON', 'ATW', 'ATV', 'PNU', 'VIS', 'FLU', 'PI', 
          'API', 'BIO', 'S', 'PFE', 'X95', 'X90', 'X50', 'X10', 'PIE']

# Cargar la muestra
sample_df = pd.read_excel(sample_file_path)

# Diccionario para almacenar las predicciones
predictions = {}

for label in labels:
    # Cargar las características seleccionadas
    features_file_path = os.path.join(Scal_Featu_dir, f'SelectKBest_{label}_features.joblib')
    selected_features = joblib.load(features_file_path)
    
    # Cargar el escalador
    scaler_file_path = os.path.join(Scal_Featu_dir, f'Scaler_{label}.joblib')
    scaler = joblib.load(scaler_file_path)
    
    # Seleccionar las características de la muestra
    X_sample = sample_df[selected_features]
    
    # Escalar las características
    X_sample_scaled = scaler.transform(X_sample.values)  # Use .values to ignore feature names
    
    # Cargar el modelo
    model_file_path = os.path.join(models_dir, f'{label}.joblib')
    model = joblib.load(model_file_path)
    
    # Realizar la predicción
    predictions[label] = model.predict(X_sample_scaled)

# Convertir las predicciones a un DataFrame
predictions_df = pd.DataFrame(predictions)

# Guardar las predicciones en un archivo Excel
predictions_df.to_excel(output_file_path, index=False)

print(f"Predicciones guardadas en {output_file_path}")



"""
Esta línea de código estamos realizando una agrupación mediante aprendizaje NO supervisado de 359 espectros de Diesel que contiene 2334 absorbancias a diferentes longitudes de onda por muestra. Este código imprime Varianza explicada por cada componente, la Varianza explicada acumulada, Influencia de las variables en el Componente Principal 1, Influencia de las variables en el Componente Principal 2, Valores propios (varianza explicada) por cada componente principal, Valor propio para el Componente Principal 1, Valor propio para el Componente Principal 2 y Visualiza los resultados de los componentes principales 1 y 2 vs 3 y 4.

Para entender qué variables tienen más influencia en cada componente principal, podemos examinar los vectores propios (eigenvectors) que resultan del análisis PCA. Cada componente principal es una combinación lineal de tus variables originales, y los coeficientes de esta combinación lineal son los elementos del vector propio correspondiente. Los componentes principales en el análisis PCA (Principal Component Analysis) se obtienen a través de una transformación lineal de las variables originales. Esta transformación se realiza de tal manera que el primer componente principal captura la mayor variación posible en los datos, el segundo componente principal captura la mayor variación restante y es ortogonal al primero, y así sucesivamente.

La combinación lineal se realiza utilizando los vectores propios (eigenvectors) y los valores propios (eigenvalues) de la matriz de covarianza de los datos. Los vectores propios representan las direcciones de los nuevos ejes (componentes principales), y los valores propios representan la cantidad de varianza que se captura en cada nuevo eje.

La fórmula para calcular el i-ésimo componente principal (PCi) a partir de las variables originales x1, x2, ..., xn es:

PCi = e1i*x1 + e2i*x2 + ... + eni*xn

donde eji es el j-ésimo elemento del i-ésimo vector propio.

Por ejemplo, si tienes dos variables originales x1 y x2, y los elementos del primer vector propio son e11 y e12, entonces el primer componente principal se calcula como:

PC1 = e11*x1 + e12*x2

Y si los elementos del segundo vector propio son e21 y e22, entonces el segundo componente principal se calcula como:

PC2 = e21*x1 + e22*x2

Y así sucesivamente para los componentes principales restantes.

Los valores propios (eigenvalues) en el análisis PCA representan la varianza explicada por cada componente principal.

Para determinar cuánta varianza explica cada componente principal, podemos dividir los valores propios por la suma total de los valores propios. Esto nos da la proporción de varianza explicada por cada componente principal.

"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA

# Cargar los datos
df = pd.read_excel('datasets/Modelo Diesel.xlsx', sheet_name='SNV')
data = df.iloc[0:361, 0:2336]

# Escalar los datos
scaler = StandardScaler()
data_scaled = scaler.fit_transform(data)

# Reducir la dimensionalidad a 4 componentes principales
pca = PCA(n_components=4)
data_pca = pca.fit_transform(data_scaled)

# Definir el número de clusters
num_clusters = 3  # Cambia este número según sea necesario

# Realizar el clustering con KMeans en los datos reducidos
kmeans_pca = KMeans(n_clusters=num_clusters, random_state=42)
kmeans_pca.fit(data_pca)
labels_pca = kmeans_pca.labels_

# Imprimir la varianza explicada por cada componente principal
print('Varianza explicada por cada componente:')
print(pca.explained_variance_ratio_)

# Imprimir la varianza explicada acumulada
print('\nVarianza explicada acumulada:')
print(np.cumsum(pca.explained_variance_ratio_))

# Obtener los nombres de las columnas originales
features = df.columns[:2336]

# Crear un DataFrame para los primeros dos componentes principales
components_df = pd.DataFrame(pca.components_[:2], columns=features)

# Para el Componente Principal 1
print("Influencia de las variables en el Componente Principal 1:")
print(components_df.loc[0].sort_values(ascending=False))

# Para el Componente Principal 2
print("\nInfluencia de las variables en el Componente Principal 2:")
print(components_df.loc[1].sort_values(ascending=False))

# Imprimir los valores propios
print("Valores propios (varianza explicada) por cada componente principal:")
print(pca.explained_variance_)

# Para el Componente Principal 1
print("\nValor propio para el Componente Principal 1:")
print(pca.explained_variance_[0])

# Para el Componente Principal 2
print("\nValor propio para el Componente Principal 2:")
print(pca.explained_variance_[1])

# Visualizar los resultados
plt.figure(figsize=(15, 5))

# Componente Principal 1 vs Componente Principal 2
plt.subplot(2, 3, 1)
scatter = plt.scatter(data_pca[:, 0], data_pca[:, 1], c=labels_pca, cmap='viridis', alpha=0.6)
plt.title('CP1 vs CP2')
plt.xlabel('Componente Principal 1')
plt.ylabel('Componente Principal 2')

# Componente Principal 1 vs Componente Principal 3
plt.subplot(2, 3, 2)
scatter = plt.scatter(data_pca[:, 0], data_pca[:, 2], c=labels_pca, cmap='viridis', alpha=0.6)
plt.title('CP1 vs CP3')
plt.xlabel('Componente Principal 1')
plt.ylabel('Componente Principal 3')

# Componente Principal 1 vs Componente Principal 4
plt.subplot(2, 3, 3)
scatter = plt.scatter(data_pca[:, 0], data_pca[:, 3], c=labels_pca, cmap='viridis', alpha=0.6)
plt.title('CP1 vs CP4')
plt.xlabel('Componente Principal 1')
plt.ylabel('Componente Principal 4')

# Componente Principal 2 vs Componente Principal 3
plt.subplot(2, 3, 4)
scatter = plt.scatter(data_pca[:, 1], data_pca[:, 2], c=labels_pca, cmap='viridis', alpha=0.6)
plt.title('CP2 vs CP3')
plt.xlabel('Componente Principal 2')
plt.ylabel('Componente Principal 3')

# Componente Principal 2 vs Componente Principal 4
plt.subplot(2, 3, 5)
scatter = plt.scatter(data_pca[:, 1], data_pca[:, 3], c=labels_pca, cmap='viridis', alpha=0.6)
plt.title('CP2 vs CP4')
plt.xlabel('Componente Principal 2')
plt.ylabel('Componente Principal 4')

# Componente Principal 3 vs Componente Principal 4
plt.subplot(2, 3, 6)
scatter = plt.scatter(data_pca[:, 2], data_pca[:, 3], c=labels_pca, cmap='viridis', alpha=0.6)
plt.title('CP3 vs CP4')
plt.xlabel('Componente Principal 3')
plt.ylabel('Componente Principal 4')

plt.tight_layout()
plt.show()



# Imprimir las primeras filas y las últimas filas del rango de columnas

import pandas as pd

# Leer el archivo Excel y seleccionar la hoja 
df = pd.read_excel('datasets/Modelo Diesel 29-05-2024.xlsx', sheet_name='prueba')

# Seleccionar el rango de columnas
column_range = df.iloc[0:352, 1:6067]  # pandas utiliza indexación basada en 0

# Mostrar las primeras filas del rango de columnas
print("Primeras filas del rango de columnas:")
print(column_range.head())

# Mostrar las últimas filas del rango de columnas
print("\nÚltimas filas del rango de columnas:")
print(column_range.tail())


# Verificar si hay celdas vacias en el rango selecconado

import pandas as pd

# Leer el archivo Excel y seleccionar la hoja
df = pd.read_excel('datasets/Modelo Diesel 29-05-2024.xlsx', sheet_name='prueba')

# Seleccionar el rango de columnas
column_range = df.iloc[0:352, 1:6067]  # pandas utiliza indexación basada en 0

# Verificar si hay celdas vacías en el rango seleccionado
empty_cells = column_range.isnull().any().any()

if empty_cells:
    print("Hay celdas vacías en el rango seleccionado.")
else:
    print("No hay celdas vacías en el rango seleccionado.")
    
nan_cells = column_range.isna().any().any()

if nan_cells:
    print("Hay celdas que contienen NaN en el rango seleccionado.")
else:
    print("No hay celdas que contienen NaN en el rango seleccionado.")
    
    
# Verificar si en el rango hay celdas vacias o NaN

import pandas as pd

# Leer el archivo Excel y seleccionar la hoja
df = pd.read_excel('datasets/Modelo Diesel 29-05-2024.xlsx', sheet_name='prueba')

labels = ['NCGRB', 'ICASTM', 'POL', 'MON', 'ATW', 'ATV', 'PNU', 'VIS', 'FLU', 'PI', 
          'API', 'BIO', 'S', 'PFE', 'X95', 'X90', 'X50', 'X10', 'PIE']

# Verificar si hay celdas vacías o con NaN en las columnas especificadas en labels
empty_cells = df[labels].isnull().any()

for label, has_nan in empty_cells.items():
    if has_nan:
        print(f"La columna '{label}' contiene celdas vacías o con NaN.")
    else:
        print(f"La columna '{label}' no contiene celdas vacías ni NaN.")
        


# Corregir valores de celdas en este caso celdas de valores de absorvancia con formatos string '1,2358E-5' a float 0.000012358. Se corre el siguiente codigo y al archivo corregido se abre en excel se seleccionan el rango de celdas correspondeintes a las absorvancias y se cambia el formato de las celdas a numero con 6 decimales.

import openpyxl

# Cargar el archivo Excel
archivo_excel = "datasets/Modelo Diesel 29-05-2024.xlsx"
wb = openpyxl.load_workbook(archivo_excel)

# Seleccionar la hoja de cálculo deseada
hoja = wb.active

# Definir el rango de celdas
rango_filas = range(1, 353)
rango_columnas = range(1, 6068)

# Iterar sobre cada celda en el rango
for fila in rango_filas:
    for columna in rango_columnas:
        celda = hoja.cell(row=fila, column=columna)
        
        # Verificar si el contenido de la celda es una cadena
        if isinstance(celda.value, str):
            # Reemplazar la coma por un punto
            celda.value = celda.value.replace(',', '.')
            
# Guardar los cambios en el archivo Excel
wb.save("datasets/Modelo Diesel 29-05-2024_corregido.xlsx")


# Evaluacion de modelos de regresion con validacion cruzada y seleccion de caracteristicas

import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import Ridge, BayesianRidge
from sklearn.neighbors import KNeighborsRegressor
from sklearn.ensemble import RandomForestRegressor, ExtraTreesRegressor, AdaBoostRegressor, HistGradientBoostingRegressor
from sklearn.svm import SVR
from sklearn.model_selection import train_test_split, cross_val_score  
from sklearn.feature_selection import SelectKBest, f_regression
from sklearn.metrics import mean_squared_error
import openpyxl
import os
import joblib

# Función para crear archivo si no existe
def create_file_if_not_exists(file_path):
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        wb.save(file_path)

# Cargar los datos
df = pd.read_excel('datasets/Modelo Diesel 29-05-2024.xlsx', sheet_name='prueba')

# Crear los archivos necesarios si no existen
create_file_if_not_exists('datasets/DieselSeleccion.xlsx')
create_file_if_not_exists('datasets/DieselPrediccion.xlsx')
create_file_if_not_exists('datasets/DieselErrorPrediccion.xlsx')

# Inicializar los modelos
models = [Ridge(), KNeighborsRegressor(), RandomForestRegressor(), SVR(), 
          ExtraTreesRegressor(), AdaBoostRegressor(), HistGradientBoostingRegressor(),
          BayesianRidge()]

# Crear el directorio models si no existe
models_dir = 'models'
if not os.path.exists(models_dir):
    os.makedirs(models_dir)

# Etiquetas que quieres usar
labels = ['NCGRB', 'ICASTM', 'POL', 'MON', 'ATW', 'ATV', 'PNU', 'VIS', 'FLU', 'PI', 
          'API', 'BIO', 'S', 'PFE', 'X95', 'X90', 'X50', 'X10', 'PIE']

# Bucle a través de las etiquetas
for label in labels:
    
    # Eliminar las filas con valores NaN en la columna de la etiqueta actual
    df = df[~df[label].isna()]
    
    # Dividir los datos en características (X) y etiquetas (y)
    cnTrainX = df.iloc[0:352, 1:6067]  
    cnTrainY = df[label]

    # Eliminar las filas con valores NaN en cnTrainY
    cnTrainX = cnTrainX[~cnTrainY.isna()]
    cnTrainY = cnTrainY.dropna()

    # Realizar selección de características
    feature_selection = SelectKBest(score_func=f_regression, k=50)
    cnTrainX_new = feature_selection.fit_transform(cnTrainX, cnTrainY)
    selected_features = cnTrainX.columns[feature_selection.get_support()]
    
    # Guardar las características seleccionadas en un archivo .joblib
    features_file_path = os.path.join(models_dir, f'Best_{label}.joblib')
    joblib.dump(selected_features, features_file_path)

    # Crear DataFrame con las puntuaciones de las características
    feature_scores = pd.DataFrame({'Feature': cnTrainX.columns, 'Score': feature_selection.scores_}).nlargest(50, 'Score')

    # Añadir las características principales a una nueva hoja en el archivo Excel
    with pd.ExcelWriter('datasets/DieselSeleccion.xlsx', engine='openpyxl', mode='a') as writer:
        feature_scores.to_excel(writer, sheet_name=label, index=False)

    # Dividir los datos en conjuntos de entrenamiento y prueba
    X_train, X_test, y_train, y_test = train_test_split(cnTrainX_new, cnTrainY, test_size=0.3, train_size=0.7, shuffle=True)

    # Escalar las características
    sc = StandardScaler()
    X_train = sc.fit_transform(X_train)
    X_test = sc.transform(X_test)
    
    # Guardar el escalador ajustado
    scaler_file_path = os.path.join(models_dir, f'Scaler_{label}.joblib')
    joblib.dump(sc, scaler_file_path)

    # DataFrames para almacenar resultados y errores
    all_predictions = pd.DataFrame()
    error_data = []

    for model in models:
        model_name = type(model).__name__

        # Entrenar el modelo
        model.fit(X_train, y_train)

        # Realizar predicciones
        y_pred_train = model.predict(X_train)
        y_pred_test = model.predict(X_test)
        
        # Guardar el modelo entrenado
        model_file_path = os.path.join(models_dir, f'{label}_{model_name}.joblib')
        joblib.dump(model, model_file_path)

        # Calcular el error cuadrático medio para las predicciones
        mse_train = mean_squared_error(y_train, y_pred_train)
        mse_test = mean_squared_error(y_test, y_pred_test)

        # Realizar la validación cruzada
        cv_scores = cross_val_score(model, cnTrainX_new, cnTrainY, cv=5, scoring='neg_mean_squared_error')
        cv_rmse = np.sqrt(-cv_scores)
        
        # Concatenar resultados
        train_results = pd.DataFrame({'y_train': y_train, f'y_pred_train_{model_name}': y_pred_train})
        test_results = pd.DataFrame({'y_test': y_test, f'y_pred_test_{model_name}': y_pred_test})
        all_predictions = pd.concat([all_predictions, train_results, test_results], axis=1)

        μ_AllMseRMSE = mse_train * 0.1 + mse_test * 0.4 + cv_scores[0]* -1 * 0.1 + cv_scores[1]* -1 * 0.1 + cv_scores[2]* -1 * 0.1 + cv_scores[3]* -1 * 0.1 + cv_scores[4]* -1 * 0.1
        
        # Guardar errores
        error_data.append({
            'Modelo': model_name,
            'MSE Train': mse_train,
            'MSE Test': mse_test,
            'CV Score 1': cv_scores[0],
            'CV Score 2': cv_scores[1],
            'CV Score 3': cv_scores[2],
            'CV Score 4': cv_scores[3],
            'CV Score 5': cv_scores[4],
            'CV RMSE 1': cv_rmse[0],
            'CV RMSE 2': cv_rmse[1],
            'CV RMSE 3': cv_rmse[2],
            'CV RMSE 4': cv_rmse[3],
            'CV RMSE 5': cv_rmse[4],
            'μ All MSE RMSE': μ_AllMseRMSE
        })

    # Guardar las predicciones y los errores en sus respectivos archivos de Excel
    with pd.ExcelWriter('datasets/DieselPrediccion.xlsx', engine='openpyxl', mode='a') as writer:
        all_predictions.to_excel(writer, sheet_name=label, index=False)

    with pd.ExcelWriter('datasets/DieselErrorPrediccion.xlsx', engine='openpyxl', mode='a') as writer:
        pd.DataFrame(error_data).to_excel(writer, sheet_name=label, index=False)

# Eliminar la hoja temporal después de agregar la primera hoja con los datos
for file_path in ['datasets/DieselSeleccion.xlsx', 'datasets/DieselPrediccion.xlsx', 'datasets/DieselErrorPrediccion.xlsx']:
    wb = openpyxl.load_workbook(file_path)
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    wb.save(file_path)
    

# Prediccion de muestras con cifras significativas

import os
import pandas as pd
import joblib
import warnings
from sklearn.exceptions import DataConversionWarning
from sklearn.preprocessing import StandardScaler

# Ignorar las advertencias de DataConversionWarning
warnings.filterwarnings(action='ignore', category=DataConversionWarning)

# Define los directorios y archivos
models_dir = 'models1'
Scal_Featu_dir = 'models1'
sample_file_path = 'datasets/Muestra Diesel.xlsx'
output_file_path = 'datasets/Predicciones Diesel.xlsx'

# Lista de etiquetas
labels = ['NCGRB', 'ICASTM', 'POL', 'MON', 'ATW', 'ATV', 'PNU', 'VIS', 'FLU', 'PI', 
          'API', 'BIO', 'S', 'PFE', 'X95', 'X90', 'X50', 'X10', 'PIE']

# Cifras significativas para cada etiqueta
sig_figs = {'NCGRB': 1, 'ICASTM': 1, 'POL': 2, 'MON': 2, 'ATW': 2, 'ATV': 2, 'PNU': 2, 
            'VIS': 2, 'FLU': 0, 'PI': 1, 'API': 1, 'BIO': 2, 'S': 1, 'PFE': 1, 'X95': 1, 
            'X90': 1, 'X50': 1, 'X10': 1, 'PIE': 1}

# Cargar la muestra
sample_df = pd.read_excel(sample_file_path)

# Diccionario para almacenar las predicciones
predictions = {}

for label in labels:
    # Cargar las características seleccionadas
    features_file_path = os.path.join(Scal_Featu_dir, f'Best_{label}.joblib')
    selected_features = joblib.load(features_file_path)
    
    # Cargar el escalador
    scaler_file_path = os.path.join(Scal_Featu_dir, f'Scaler_{label}.joblib')
    scaler = joblib.load(scaler_file_path)
    
    # Seleccionar las características de la muestra
    X_sample = sample_df[selected_features]
    
    # Escalar las características
    X_sample_scaled = scaler.transform(X_sample.values)  # Use .values to ignore feature names
    
    # Cargar el modelo
    model_file_path = os.path.join(models_dir, f'{label}.joblib')
    model = joblib.load(model_file_path)
    
    # Realizar la predicción
    predictions[label] = model.predict(X_sample_scaled)

# Convertir las predicciones a un DataFrame
predictions_df = pd.DataFrame(predictions)

# Redondear las predicciones a las cifras significativas especificadas
for label in labels:
    predictions_df[label] = predictions_df[label].round(sig_figs[label])

# Guardar las predicciones en un archivo Excel
predictions_df.to_excel(output_file_path, index=False)

print(f"Predicciones guardadas en {output_file_path}")



""" 2/03/2025 5:37 PM Sugerencia de correccion del codigo para modelar propiedades de Diesel, visto arriba 20/05/2024 3:24. Aun no se ha probado el codigo
pero se sugiere que se pruebe y se corrija si es necesario. """



import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LinearRegression, Ridge, Lasso, ElasticNet, BayesianRidge
from sklearn.neighbors import KNeighborsRegressor
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor, AdaBoostRegressor
from sklearn.svm import SVR
from sklearn.neural_network import MLPRegressor
from xgboost import XGBRegressor
from sklearn.model_selection import train_test_split, cross_val_score, GridSearchCV
from sklearn.feature_selection import SelectKBest, f_regression, RFE
from sklearn.metrics import mean_squared_error, r2_score, mean_absolute_error
import joblib
import os
import openpyxl

# Función para crear archivo si no existe
def create_file_if_not_exists(file_path):
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        wb.save(file_path)

# Cargar los datos
df = pd.read_excel('datasets/Modelo Diesel.xlsx', sheet_name='SNV')

# Manejo de valores nulos
df.fillna(df.mean(), inplace=True)  # Imputación con la media

# Crear archivos si no existen
for file in ['datasets/DieselSeleccion.csv', 'datasets/DieselPrediccion.csv', 'datasets/DieselErrorPrediccion.csv']:
    if not os.path.exists(file):
        open(file, 'w').close()

# Modelos con optimización de hiperparámetros
models = {
    'LinearRegression': LinearRegression(),
    'Ridge': Ridge(alpha=1.0),
    'Lasso': Lasso(alpha=0.1),
    'ElasticNet': ElasticNet(alpha=0.1, max_iter=1000),
    'KNeighborsRegressor': KNeighborsRegressor(n_neighbors=5),
    'DecisionTreeRegressor': DecisionTreeRegressor(max_depth=10),
    'RandomForestRegressor': RandomForestRegressor(n_estimators=100, max_depth=10),
    'GradientBoostingRegressor': GradientBoostingRegressor(n_estimators=100, max_depth=5),
    'SVR': SVR(kernel='rbf', C=1.0, epsilon=0.1),
    'MLPRegressor': MLPRegressor(hidden_layer_sizes=(100,), max_iter=5000),
    'XGBRegressor': XGBRegressor(n_estimators=100, max_depth=5),
    'BayesianRidge': BayesianRidge()
}

# Directorio para guardar modelos
models_dir = 'models'
os.makedirs(models_dir, exist_ok=True)

labels = ['NCGRB', 'ICASTM', 'POL', 'MON', 'ATW', 'ATV', 'PNU', 'VIS', 'FLU', 'PI', 
          'API', 'BIO', 'S', 'PFE', 'X95', 'X90', 'X50', 'X10', 'PIE']

for label in labels:
    cnTrainX = df.iloc[:, :-len(labels)]
    cnTrainY = df[label]
    
    # Selección de características con RFE
    selector = RFE(estimator=RandomForestRegressor(n_estimators=50), n_features_to_select=50)
    cnTrainX_new = selector.fit_transform(cnTrainX, cnTrainY)
    selected_features = cnTrainX.columns[selector.support_]
    joblib.dump(selected_features, os.path.join(models_dir, f'SelectedFeatures_{label}.joblib'))
    
    # Guardar selección de características en CSV
    pd.DataFrame({'Feature': cnTrainX.columns, 'Selected': selector.support_}).to_csv('datasets/DieselSeleccion.csv', mode='a', index=False)
    
    # División de datos
    X_train, X_test, y_train, y_test = train_test_split(cnTrainX_new, cnTrainY, test_size=0.3, random_state=42)
    
    # Escalado de datos
    sc = StandardScaler()
    X_train = sc.fit_transform(X_train)
    X_test = sc.transform(X_test)
    joblib.dump(sc, os.path.join(models_dir, f'Scaler_{label}.joblib'))
    
    all_predictions = []
    error_data = []
    
    for name, model in models.items():
        model.fit(X_train, y_train)
        y_pred_train = model.predict(X_train).astype(np.float32)
        y_pred_test = model.predict(X_test).astype(np.float32)
        
        # Guardar modelo entrenado
        joblib.dump(model, os.path.join(models_dir, f'{label}_{name}.joblib'))
        
        # Evaluación del modelo
        mse_train = mean_squared_error(y_train, y_pred_train)
        mse_test = mean_squared_error(y_test, y_pred_test)
        r2_test = r2_score(y_test, y_pred_test)
        mae_test = mean_absolute_error(y_test, y_pred_test)
        
        cv_scores = cross_val_score(model, cnTrainX_new, cnTrainY, cv=5, scoring='neg_mean_squared_error')
        cv_rmse = np.sqrt(-cv_scores).mean()
        
        # Guardar errores
        error_data.append({
            'Modelo': name, 'MSE Train': mse_train, 'MSE Test': mse_test,
            'R2 Test': r2_test, 'MAE Test': mae_test, 'CV RMSE': cv_rmse
        })
        
        # Guardar predicciones
        all_predictions.append(pd.DataFrame({'y_test': y_test, f'y_pred_{name}': y_pred_test}))
    
    # Guardar predicciones en CSV
    pd.concat(all_predictions, axis=1).to_csv('datasets/DieselPrediccion.csv', mode='a', index=False)
    
    # Guardar errores en CSV
    pd.DataFrame(error_data).to_csv('datasets/DieselErrorPrediccion.csv', mode='a', index=False)

print("Entrenamiento y evaluación completados.")